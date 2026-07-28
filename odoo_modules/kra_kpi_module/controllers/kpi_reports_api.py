from odoo import http, fields
from odoo.http import request
from datetime import timedelta
import json
import re

# Idle-check timings live with the model that owns them (kpi.work.session) so the
# cron and these routes can never drift apart on the numbers.
from ..models.kpi_work_session import IDLE_CHECK_MIN, NO_TASK_NUDGE_MIN, _fmt_hms
from ..models.kpi_nontask_block import NOTE_MAX, NONTASK_REASON_CODES


def _parse_task_array(text):
    """Try to extract a JSON array of strings from an LLM response. Falls back to
    line-splitting if the model produced free text.

    Accepts shapes:
      ["task a", "task b"]
      {"tasks": ["task a", ...]}   (Groq json_object mode, OpenAI-style)
      {"items": ["task a", ...]}   (other models that pick different keys)
      {"<anything>": [list of strings]}
    """
    if not text:
        return []

    def _from_value(val):
        if isinstance(val, list):
            return [str(x).strip() for x in val if str(x).strip()][:50]
        if isinstance(val, dict):
            # Prefer well-known keys first
            for key in ('tasks', 'items', 'results', 'task_names', 'data'):
                if isinstance(val.get(key), list):
                    return [str(x).strip() for x in val[key] if str(x).strip()][:50]
            # Otherwise: first list-of-strings value we find
            for v in val.values():
                if isinstance(v, list) and v and all(isinstance(x, str) for x in v):
                    return [x.strip() for x in v if x.strip()][:50]
        return None

    # First attempt: parse as JSON directly
    try:
        got = _from_value(json.loads(text))
        if got is not None:
            return got
    except Exception:
        pass
    # Second attempt: find a JSON array inside the text
    m = re.search(r'\[(?:[^\[\]]|\n)*\]', text, re.DOTALL)
    if m:
        try:
            got = _from_value(json.loads(m.group(0)))
            if got is not None:
                return got
        except Exception:
            pass
    # Third attempt: split lines, strip bullets/numbers
    out = []
    for line in text.splitlines():
        s = line.strip().lstrip('-*•').strip()
        s = re.sub(r'^\d+[\.\)]\s*', '', s)
        s = s.strip('"\' ')
        if 5 <= len(s) <= 200:
            out.append(s)
    return out[:50]


class KpiReportsAPI(http.Controller):
    
    @http.route('/kpi_reports/unit/generate', type='json', auth='user', methods=['POST'], csrf=False)
    def generate_unit_report(self, **params):
        """Generate KPI Unit Report data"""
        try:
            from_date = params.get('from_date')
            to_date = params.get('to_date')
            employee_id = params.get('employee_id')
            time_frame = params.get('time_frame')
            
            report = request.env['kpi.unit.report'].create({
                'from_date': from_date,
                'to_date': to_date,
                'employee_id': int(employee_id) if employee_id else False,
                'time_frame': time_frame or False,
            })
            
            report_data = report.get_report_data()
            
            return {
                'status': True,
                'report_id': report.id,
                'data': report_data,
            }
        except Exception as e:
            return {
                'status': False,
                'message': str(e),
            }
    
    @http.route('/kpi_reports/unit/employee_tasks', type='json', auth='user', methods=['POST'], csrf=False)
    def get_employee_tasks(self, **params):
        """Get detailed task breakdown for an employee"""
        try:
            employee_id = params.get('employee_id')
            from_date = params.get('from_date')
            to_date = params.get('to_date')
            
            if not employee_id:
                return {'status': False, 'message': 'employee_id is required'}
            
            domain = [('user_id', '=', int(employee_id))]
            
            if from_date and to_date:
                domain.append('|')
                domain.append('&')
                domain.append(('deadline', '>=', from_date))
                domain.append(('deadline', '<=', to_date))
                domain.append('&')
                domain.append(('create_date', '>=', from_date + ' 00:00:00'))
                domain.append(('create_date', '<=', to_date + ' 23:59:59'))
            elif from_date:
                domain.append('|')
                domain.append(('deadline', '>=', from_date))
                domain.append(('create_date', '>=', from_date + ' 00:00:00'))
            elif to_date:
                domain.append('|')
                domain.append(('deadline', '<=', to_date))
                domain.append(('create_date', '<=', to_date + ' 23:59:59'))
            
            tasks = request.env['kra.kpi'].sudo().search(domain, order='id')
            
            task_list = []
            for task in tasks:
                estimate_hrs = (task.estimate_hours or 0) + ((task.estimate_minutes or 0) / 60.0)
                hrs_spent = (task.timer_total_seconds or 0) / 3600.0
                
                task_list.append({
                    'id': task.id,
                    'seq_no': task.id,
                    'name': task.name,
                    'estimated_hrs': estimate_hrs,
                    'hrs_spent': (task.timer_total_seconds or 0) / 3600.0,
                    'estimated_unit': estimate_hrs,
                    'unit_spent': hrs_spent * 100,
                    'kra_name': task.kra_id.name if task.kra_id else '',
                    'priority': task.priority or '',
                    'task_state': task.task_state or '',
                })
            
            return {
                'status': True,
                'tasks': task_list,
                'employee_name': request.env['res.users'].browse(int(employee_id)).name,
            }
        except Exception as e:
            import logging
            _logger = logging.getLogger(__name__)
            _logger.error(f"Error in get_employee_tasks: {str(e)}")
            return {
                'status': False,
                'message': str(e),
            }
    
    @http.route('/kpi_reports/performance/get_employees', type='json', auth='user', methods=['POST'], csrf=False)
    def get_employees_with_tasks(self, **params):
        """Get list of employees who have tasks in the date range with saved salary data"""
        try:
            from_date = params.get('from_date')
            to_date = params.get('to_date')
            employee_id = params.get('employee_id')
            
            domain = [('user_id', '!=', False)]
            
            if employee_id:
                domain.append(('user_id', '=', int(employee_id)))
            
            if from_date and to_date:
                domain.append('|')
                domain.append('&')
                domain.append(('deadline', '>=', from_date))
                domain.append(('deadline', '<=', to_date))
                domain.append('&')
                domain.append(('create_date', '>=', from_date + ' 00:00:00'))
                domain.append(('create_date', '<=', to_date + ' 23:59:59'))
            elif from_date:
                domain.append('|')
                domain.append(('deadline', '>=', from_date))
                domain.append(('create_date', '>=', from_date + ' 00:00:00'))
            elif to_date:
                domain.append('|')
                domain.append(('deadline', '<=', to_date))
                domain.append(('create_date', '<=', to_date + ' 23:59:59'))
            
            kpi_tasks = request.env['kra.kpi'].search(domain)
            employee_ids = kpi_tasks.mapped('user_id')
            
            employees = []
            for emp in employee_ids:
                saved_salary_config = request.env['kpi.employee.salary.config'].search([
                    ('employee_id', '=', emp.id)
                ], limit=1)
                
                employees.append({
                    'id': emp.id,
                    'name': emp.name,
                    'saved_salary': saved_salary_config.salary if saved_salary_config else 0
                })
            
            return {
                'status': True,
                'employees': employees
            }
        except Exception as e:
            import logging
            _logger = logging.getLogger(__name__)
            _logger.error(f"Error in get_employees_with_tasks: {str(e)}")
            return {
                'status': False,
                'message': str(e),
                'employees': []
            }
    
    @http.route('/kpi_reports/performance/generate', type='json', auth='user', methods=['POST'], csrf=False)
    def generate_performance_report(self, **params):
        """Generate KPI Performance Report data with salary inputs"""
        try:
            from_date = params.get('from_date')
            to_date = params.get('to_date')
            employee_id = params.get('employee_id')
            time_frame = params.get('time_frame')
            employee_salary_data = params.get('employee_salary_data', [])
            
            report = request.env['kpi.performance.report'].create({
                'from_date': from_date,
                'to_date': to_date,
                'employee_id': int(employee_id) if employee_id else False,
                'time_frame': time_frame or False,
            })
            
            for emp_data in employee_salary_data:
                request.env['kpi.performance.employee.salary'].create({
                    'report_id': report.id,
                    'employee_id': emp_data['employee_id'],
                    'salary': emp_data['salary'],
                    'attendance_days': emp_data['attendance_days']
                })
                
                salary_config = request.env['kpi.employee.salary.config'].search([
                    ('employee_id', '=', emp_data['employee_id'])
                ], limit=1)
                
                if salary_config:
                    salary_config.write({
                        'salary': emp_data['salary'],
                        'attendance_days': emp_data['attendance_days'],
                        'last_updated': request.env.cr.now()
                    })
                else:
                    request.env['kpi.employee.salary.config'].create({
                        'employee_id': emp_data['employee_id'],
                        'salary': emp_data['salary'],
                        'attendance_days': emp_data['attendance_days']
                    })
            
            report_data = report.get_report_data()
            
            return {
                'status': True,
                'report_id': report.id,
                'data': report_data,
            }
        except Exception as e:
            import logging
            _logger = logging.getLogger(__name__)
            _logger.error(f"Error in generate_performance_report: {str(e)}")
            return {
                'status': False,
                'message': str(e),
            }
    
    @http.route('/kpi_reports/performance/generate_direct', type='json', auth='user', methods=['POST'], csrf=False)
    def generate_performance_report_direct(self, **params):
        """Generate KPI Performance Report directly from Employee Salary Master"""
        try:
            from_date = params.get('from_date')
            to_date = params.get('to_date')
            employee_id = params.get('employee_id')
            
            domain = [('user_id', '!=', False)]
            
            if employee_id:
                domain.append(('user_id', '=', int(employee_id)))
            
            if from_date and to_date:
                domain.append('|')
                domain.append('&')
                domain.append(('deadline', '>=', from_date))
                domain.append(('deadline', '<=', to_date))
                domain.append('&')
                domain.append(('create_date', '>=', from_date + ' 00:00:00'))
                domain.append(('create_date', '<=', to_date + ' 23:59:59'))
            elif from_date:
                domain.append('|')
                domain.append(('deadline', '>=', from_date))
                domain.append(('create_date', '>=', from_date + ' 00:00:00'))
            elif to_date:
                domain.append('|')
                domain.append(('deadline', '<=', to_date))
                domain.append(('create_date', '<=', to_date + ' 23:59:59'))
            
            kpi_tasks = request.env['kra.kpi'].search(domain)
            
            employee_ids = kpi_tasks.mapped('user_id')
            
            reassignment_domain = [('previous_assignee_id', '!=', False)]
            
            if from_date:
                reassignment_domain.append(('reassignment_date', '>=', from_date + ' 00:00:00'))
            if to_date:
                reassignment_domain.append(('reassignment_date', '<=', to_date + ' 23:59:59'))
            
            if employee_id:
                reassignment_domain.append(('previous_assignee_id', '=', int(employee_id)))
            
            reassignment_records = request.env['kpi.reassignment.history'].search(reassignment_domain)
            reassignment_employee_ids = reassignment_records.mapped('previous_assignee_id')
            
            all_employee_ids = employee_ids | reassignment_employee_ids
            
            salary_configs = request.env['kpi.employee.salary.config'].search([
                ('employee_id', 'in', all_employee_ids.ids)
            ])
            
            salary_map = {
                config.employee_id.id: {
                    'salary': config.salary,
                    'working_days': config.working_days,
                    'available_hours': config.available_hours,
                    'available_units': config.available_units,
                    'unit_cost': config.unit_cost
                }
                for config in salary_configs
            }
            
            missing_employees = []
            
            employee_data = {}
            for kpi in kpi_tasks:
                emp = kpi.user_id
                emp_id = emp.id
                
                if emp_id not in employee_data:
                    dept = ''
                    if hasattr(emp, 'employee_id') and emp.employee_id:
                        dept = emp.employee_id.department_id.name or ''
                    
                    employee_data[emp_id] = {
                        'employee_id': emp_id,
                        'employee_name': emp.name,
                        'department': dept,
                        'total_estimated_hours': 0,
                        'total_actual_hours': 0,
                    }
                
                estimate_hrs = (kpi.estimate_hours or 0) + ((kpi.estimate_minutes or 0) / 60.0)
                actual_hrs = (kpi.timer_total_seconds or 0) / 3600.0
                
                employee_data[emp_id]['total_estimated_hours'] += estimate_hrs
                employee_data[emp_id]['total_actual_hours'] += actual_hrs
            
            for reassignment in reassignment_records:
                prev_emp_id = reassignment.previous_assignee_id.id
                
                if prev_emp_id not in employee_data:
                    emp = reassignment.previous_assignee_id
                    dept = ''
                    if hasattr(emp, 'employee_id') and emp.employee_id:
                        dept = emp.employee_id.department_id.name or ''
                    
                    employee_data[prev_emp_id] = {
                        'employee_id': prev_emp_id,
                        'employee_name': emp.name,
                        'department': dept,
                        'total_estimated_hours': 0,
                        'total_actual_hours': 0,
                    }
                
                reassignment_hours = (reassignment.time_spent_seconds or 0) / 3600.0
                employee_data[prev_emp_id]['total_actual_hours'] += reassignment_hours
                
                kpi_task = reassignment.kpi_id
                if kpi_task:
                    estimate_hrs = (kpi_task.estimate_hours or 0) + ((kpi_task.estimate_minutes or 0) / 60.0)
                    employee_data[prev_emp_id]['total_estimated_hours'] += estimate_hrs
            
            result = []
            for emp_id, data in employee_data.items():
                salary_info = salary_map.get(emp_id)
                
                if not salary_info or salary_info['salary'] <= 0 or salary_info['working_days'] <= 0:
                    missing_employees.append({
                        'id': emp_id,
                        'name': data['employee_name']
                    })
                    continue
                
                salary = salary_info['salary']
                working_days = salary_info['working_days']
                available_hours = salary_info['available_hours']
                available_units = salary_info['available_units']
                unit_cost = salary_info['unit_cost']
                
                estimated_hrs = data['total_estimated_hours']
                actual_hrs = data['total_actual_hours']
                
                estimated_minutes = round(estimated_hrs * 60)
                actual_minutes = round(actual_hrs * 60)
                
                estimated_hrs_rounded = estimated_minutes / 60.0
                actual_hrs_rounded = actual_minutes / 60.0
                
                estimated_units = estimated_hrs_rounded * 100
                productivity_units = actual_hrs_rounded * 100
                performance_deviation = estimated_units - productivity_units
                productivity_cost = productivity_units * unit_cost
                productivity_cost_pct = (actual_hrs_rounded / available_hours * 100) if available_hours > 0 else 0

                # ✅ FIX: Set deviation to 0 for not started tasks
                if estimated_hrs_rounded == 0 and actual_hrs_rounded == 0:
                    remark = 'No Tasks Assigned'
                    remark_class = 'text-warning'
                    performance_deviation = 0  # ✅ Force deviation to 0
                elif estimated_hrs_rounded > 0 and actual_hrs_rounded == 0:
                    remark = 'Not Started'
                    remark_class = 'text-primary'
                    performance_deviation = 0  # ✅ Force deviation to 0 for not started
                else:
                    if performance_deviation > 0:
                        remark = 'Excellent Performance'
                        remark_class = 'text-success'
                    else:
                        remark = 'Low Performance'
                        remark_class = 'text-danger'
                
                result.append({
                    'employee_id': emp_id,
                    'employee_name': data['employee_name'],
                    'department': data.get('department', ''),
                    'salary': round(salary, 2),
                    'working_days': working_days,
                    'estimated_hours': round(estimated_hrs_rounded, 2),
                    'estimated_units': round(estimated_units, 2),
                    'actual_hours': round(actual_hrs_rounded, 2),
                    'productivity_units': round(productivity_units, 2),
                    'performance_deviation': round(performance_deviation, 2),
                    'available_hours': round(available_hours, 2),
                    'available_units': round(available_units, 2),
                    'unit_cost': round(unit_cost, 4),
                    'productivity_cost': round(productivity_cost, 2),
                    'productivity_cost_pct': round(productivity_cost_pct, 2),
                    'remarks': remark,
                    'remarks_class': remark_class,
                })
            
            return {
                'status': True,
                'data': result,
                'missing_employees': missing_employees
            }
        except Exception as e:
            import logging
            _logger = logging.getLogger(__name__)
            _logger.error(f"Error in generate_performance_report_direct: {str(e)}")
            return {
                'status': False,
                'message': str(e),
                'data': [],
                'missing_employees': []
            }
    
    # ============================================================
    #  ✅ NEW: Individual Employee Performance Detail Endpoint
    # ============================================================
    @http.route('/kpi_reports/performance/employee_detail', type='json', auth='user', methods=['POST'], csrf=False)
    def get_employee_performance_detail(self, **params):
        """Get task-level performance breakdown for a specific employee"""
        try:
            employee_id = params.get('employee_id')
            from_date = params.get('from_date')
            to_date = params.get('to_date')
            
            if not employee_id:
                return {'status': False, 'message': 'employee_id is required'}
            
            emp_id = int(employee_id)
            employee = request.env['res.users'].sudo().browse(emp_id)
            
            if not employee.exists():
                return {'status': False, 'message': 'Employee not found'}
            
            # Build domain for current tasks
            current_tasks_domain = [('user_id', '=', emp_id)]
            
            if from_date and to_date:
                current_tasks_domain.append('|')
                current_tasks_domain.append('&')
                current_tasks_domain.append(('deadline', '>=', from_date))
                current_tasks_domain.append(('deadline', '<=', to_date))
                current_tasks_domain.append('&')
                current_tasks_domain.append(('create_date', '>=', from_date + ' 00:00:00'))
                current_tasks_domain.append(('create_date', '<=', to_date + ' 23:59:59'))
            elif from_date:
                current_tasks_domain.append('|')
                current_tasks_domain.append(('deadline', '>=', from_date))
                current_tasks_domain.append(('create_date', '>=', from_date + ' 00:00:00'))
            elif to_date:
                current_tasks_domain.append('|')
                current_tasks_domain.append(('deadline', '<=', to_date))
                current_tasks_domain.append(('create_date', '<=', to_date + ' 23:59:59'))
            
            # Get current tasks
            current_tasks = request.env['kra.kpi'].sudo().search(current_tasks_domain)
            
            # Build domain for reassignment history
            reassignment_domain = [('previous_assignee_id', '=', emp_id)]
            
            if from_date:
                reassignment_domain.append(('reassignment_date', '>=', from_date + ' 00:00:00'))
            if to_date:
                reassignment_domain.append(('reassignment_date', '<=', to_date + ' 23:59:59'))
            
            # Get reassignment history
            reassignment_records = request.env['kpi.reassignment.history'].sudo().search(reassignment_domain)
            
            # Process tasks
            task_details = []
            
            # Process current tasks
            for task in current_tasks:
                estimate_hrs = (task.estimate_hours or 0) + ((task.estimate_minutes or 0) / 60.0)
                actual_hrs = (task.timer_total_seconds or 0) / 3600.0
                
                # Apply same calculation logic as main report
                estimated_minutes = round(estimate_hrs * 60)
                actual_minutes = round(actual_hrs * 60)
                
                estimated_hrs_rounded = estimated_minutes / 60.0
                actual_hrs_rounded = actual_minutes / 60.0
                
                estimated_units = estimated_hrs_rounded * 100
                productivity_units = actual_hrs_rounded * 100
                performance_deviation = estimated_units - productivity_units

                # ✅ FIX: Set deviation to 0 for not started tasks
                if actual_hrs_rounded == 0 and estimated_hrs_rounded > 0:
                    # Task not started - set deviation to 0
                    performance_deviation = 0
                    display_status = 'not_started'
                else:
                    display_status = task.task_state or 'not_started'
                
                task_details.append({
                    'task_id': task.id,
                    'task_name': task.name,
                    'status': task.task_state or 'not_started',
                    'estimated_hrs': round(estimated_hrs_rounded, 2),
                    'estimated_units': round(estimated_units, 2),
                    'actual_hrs': round(actual_hrs_rounded, 2),
                    'productivity_units': round(productivity_units, 2),
                    'performance_deviation': round(performance_deviation, 2),
                    'kra_name': task.kra_id.name if task.kra_id else '',
                    'priority': task.priority or 'regular',
                    'assignment_type': 'Current Assignee',
                })
            
            # Process reassigned tasks
            for reassignment in reassignment_records:
                task = reassignment.kpi_id
                if not task:
                    continue
                
                estimate_hrs = (task.estimate_hours or 0) + ((task.estimate_minutes or 0) / 60.0)
                
                # Actual hours from reassignment history (time spent by previous assignee)
                actual_hrs = (reassignment.time_spent_seconds or 0) / 3600.0
                
                # Apply same calculation logic
                estimated_minutes = round(estimate_hrs * 60)
                actual_minutes = round(actual_hrs * 60)
                
                estimated_hrs_rounded = estimated_minutes / 60.0
                actual_hrs_rounded = actual_minutes / 60.0
                
                estimated_units = estimated_hrs_rounded * 100
                productivity_units = actual_hrs_rounded * 100
                performance_deviation = estimated_units - productivity_units

                if actual_hrs_rounded == 0 and estimated_hrs_rounded > 0:
                    performance_deviation = 0
                
                task_details.append({
                    'task_id': task.id,
                    'task_name': task.name,
                    'status': 'reassigned',
                    'estimated_hrs': round(estimated_hrs_rounded, 2),
                    'estimated_units': round(estimated_units, 2),
                    'actual_hrs': round(actual_hrs_rounded, 2),
                    'productivity_units': round(productivity_units, 2),
                    'performance_deviation': round(performance_deviation, 2),
                    'kra_name': task.kra_id.name if task.kra_id else '',
                    'priority': task.priority or 'regular',
                    'assignment_type': 'Previous Assignee (Reassigned)',
                })
            
            return {
                'status': True,
                'employee_name': employee.name,
                'tasks': task_details,
                'total_tasks': len(task_details)
            }
            
        except Exception as e:
            import logging
            _logger = logging.getLogger(__name__)
            _logger.error(f"Error in get_employee_performance_detail: {str(e)}")
            return {
                'status': False,
                'message': str(e),
                'tasks': []
            }
        
    @http.route('/kpi_reports/task/details', type='json', auth='user', methods=['POST'], csrf=False)
    def get_task_details(self, **params):
        """Get detailed information about a specific KPI task"""
        try:
            task_id = params.get('task_id')
            
            if not task_id:
                return {'status': False, 'message': 'task_id is required'}
            
            task = request.env['kra.kpi'].sudo().browse(int(task_id))
            
            if not task.exists():
                return {'status': False, 'message': 'Task not found'}
            
            estimate = (task.estimate_hours or 0) + ((task.estimate_minutes or 0) / 60.0)
            
            return {
                'status': True,
                'task': {
                    'id': task.id,
                    'name': task.name,
                    'kra_name': task.kra_id.name if task.kra_id else '',
                    'priority': task.priority or '',
                    'estimate': estimate,
                    'points': task.points or 0,
                    'assignee': task.user_id.name if task.user_id else '',
                    'user_group': task.user_group_id.name if task.user_group_id else '',
                    'deadline': str(task.deadline) if task.deadline else '',
                    'reminder_days': task.reminder_days or 0,
                    'description': task.description or '',
                    'checklist': task.checklist or '',
                    'guidelines': task.guidelines or '',
                    'task_state': task.task_state or '',
                    'timer_total_seconds': task.timer_total_seconds or 0,
                    'paused_reason': task.paused_reason or '',
                }
            }
        except Exception as e:
            return {
                'status': False,
                'message': str(e),
            }
    
    @http.route('/kpi_reports/attention/generate', type='json', auth='user', methods=['POST'], csrf=False)
    def generate_attention_report(self, **params):
        """Generate KPI Attention Report data for paused tasks"""
        try:
            employee_id = params.get('employee_id')
            from_date = params.get('from_date')
            to_date = params.get('to_date')
            
            report = request.env['kpi.attention.report'].create({
                'employee_id': int(employee_id) if employee_id else False,
                'from_date': from_date or False,
                'to_date': to_date or False,
            })
            
            report_data = report.get_report_data()
            
            return {
                'status': True,
                'report_id': report.id,
                'data': report_data,
            }
        except Exception as e:
            import logging
            _logger = logging.getLogger(__name__)
            _logger.error(f"Error in generate_attention_report: {str(e)}")
            return {
                'status': False,
                'message': str(e),
            }
    
    # ============================================================
    #                     KRA REPORT APIs
    # ============================================================
    
    @http.route('/kpi_reports/kra/get_filters', type='json', auth='user', methods=['POST'], csrf=False)
    def get_kra_report_filters(self, **params):
        """Get filter options (KRAs and Employees) for KRA Report"""
        try:
            kra_report = request.env['kpi.kra.report']
            
            kras = kra_report.get_all_kras_for_filter()
            employees = kra_report.get_all_employees_for_filter()
            
            return {
                'status': True,
                'kras': kras,
                'employees': employees,
            }
        except Exception as e:
            import logging
            _logger = logging.getLogger(__name__)
            _logger.error(f"Error in get_kra_report_filters: {str(e)}")
            return {
                'status': False,
                'message': str(e),
                'kras': [],
                'employees': [],
            }
    
    @http.route('/kpi_reports/kra/generate', type='json', auth='user', methods=['POST'], csrf=False)
    def generate_kra_report(self, **params):
        """Generate KRA Report with hierarchical structure"""
        try:
            from_date = params.get('from_date')
            to_date = params.get('to_date')
            employee_id = params.get('employee_id')
            kra_id = params.get('kra_id')
            time_frame = params.get('time_frame')
            
            report = request.env['kpi.kra.report'].create({
                'from_date': from_date,
                'to_date': to_date,
                'employee_id': int(employee_id) if employee_id else False,
                'kra_id': int(kra_id) if kra_id else False,
                'time_frame': time_frame or False,
            })
            
            report_data = report.get_report_data()
            summary_data = report.get_kra_summary()
            
            return {
                'status': True,
                'report_id': report.id,
                'data': report_data,
                'summary': summary_data,
                'filters': {
                    'from_date': from_date,
                    'to_date': to_date,
                    'employee_name': request.env['res.users'].browse(int(employee_id)).name if employee_id else 'All Employees',
                    'kra_name': request.env['kra.master'].browse(int(kra_id)).name if kra_id else 'All KRAs',
                }
            }
        except Exception as e:
            import logging
            _logger = logging.getLogger(__name__)
            _logger.error(f"Error in generate_kra_report: {str(e)}")
            return {
                'status': False,
                'message': str(e),
            }
    
    @http.route('/kpi_reports/kra/summary', type='json', auth='user', methods=['POST'], csrf=False)
    def get_kra_report_summary(self, **params):
        """Get KRA Report summary grouped by KRA hierarchy"""
        try:
            from_date = params.get('from_date')
            to_date = params.get('to_date')
            employee_id = params.get('employee_id')
            kra_id = params.get('kra_id')
            
            report = request.env['kpi.kra.report'].create({
                'from_date': from_date,
                'to_date': to_date,
                'employee_id': int(employee_id) if employee_id else False,
                'kra_id': int(kra_id) if kra_id else False,
            })
            
            summary_data = report.get_kra_summary()
            
            return {
                'status': True,
                'summary': summary_data,
            }
        except Exception as e:
            import logging
            _logger = logging.getLogger(__name__)
            _logger.error(f"Error in get_kra_report_summary: {str(e)}")
            return {
                'status': False,
                'message': str(e),
            }
    
    # ============================================================
    #              KRA EMPLOYEE REPORT APIs
    # ============================================================
    
    @http.route('/kpi_reports/employee/get_employees', type='json', auth='user', methods=['POST'], csrf=False)
    def get_employees_for_employee_report(self, **params):
        """Get list of all employees who have KPI tasks assigned"""
        try:
            report = request.env['kpi.employee.report']
            employees = report.get_all_employees()
            
            return {
                'status': True,
                'employees': employees,
            }
        except Exception as e:
            import logging
            _logger = logging.getLogger(__name__)
            _logger.error(f"Error in get_employees_for_employee_report: {str(e)}")
            return {
                'status': False,
                'message': str(e),
                'employees': [],
            }
    
    @http.route('/kpi_reports/employee/generate', type='json', auth='user', methods=['POST'], csrf=False)
    def generate_employee_report(self, **params):
        """Generate KRA Employee Report with detailed task breakdown"""
        try:
            from_date = params.get('from_date')
            to_date = params.get('to_date')
            employee_ids = params.get('employee_ids', [])
            time_frame = params.get('time_frame')
            
            # Create report record
            report_vals = {
                'from_date': from_date,
                'to_date': to_date,
                'time_frame': time_frame or False,
            }
            
            report = request.env['kpi.employee.report'].create(report_vals)
            
            # Set employee_ids if provided
            if employee_ids:
                employee_ids_int = [int(eid) for eid in employee_ids if eid]
                report.write({'employee_ids': [(6, 0, employee_ids_int)]})
            
            # Generate report data
            report_data = report.get_report_data()
            
            return {
                'status': True,
                'report_id': report.id,
                'data': report_data,
                'filters': {
                    'from_date': from_date,
                    'to_date': to_date,
                    'employees': ', '.join(report.employee_ids.mapped('name')) if report.employee_ids else 'All Employees',
                    'time_frame': time_frame or 'Custom',
                }
            }
        except Exception as e:
            import logging
            _logger = logging.getLogger(__name__)
            _logger.error(f"Error in generate_employee_report: {str(e)}")
            return {
                'status': False,
                'message': str(e),
            }
    
    @http.route('/kpi_reports/employee/pdf_data', type='json', auth='user', methods=['POST'], csrf=False)
    def get_employee_report_pdf_data(self, **params):
        """Get formatted data for PDF generation of KRA Employee Report"""
        try:
            from_date = params.get('from_date')
            to_date = params.get('to_date')
            employee_ids = params.get('employee_ids', [])
            time_frame = params.get('time_frame')
            
            # Create report record
            report_vals = {
                'from_date': from_date,
                'to_date': to_date,
                'time_frame': time_frame or False,
            }
            
            report = request.env['kpi.employee.report'].create(report_vals)
            
            # Set employee_ids if provided
            if employee_ids:
                employee_ids_int = [int(eid) for eid in employee_ids if eid]
                report.write({'employee_ids': [(6, 0, employee_ids_int)]})
            
            # Get PDF formatted data
            pdf_data = report.get_pdf_report_data()
            
            return {
                'status': True,
                'pdf_data': pdf_data,
            }
        except Exception as e:
            import logging
            _logger = logging.getLogger(__name__)
            _logger.error(f"Error in get_employee_report_pdf_data: {str(e)}")
            return {
                'status': False,
                'message': str(e),
            }
    
    @http.route('/kpi_reports/employee/task_details', type='json', auth='user', methods=['POST'], csrf=False)
    def get_employee_task_details(self, **params):
        """Get detailed information about a specific task including all progress entries"""
        try:
            task_id = params.get('task_id')
            from_date = params.get('from_date')
            to_date = params.get('to_date')
            
            if not task_id:
                return {'status': False, 'message': 'task_id is required'}
            
            task = request.env['kra.kpi'].sudo().browse(int(task_id))
            
            if not task.exists():
                return {'status': False, 'message': 'Task not found'}
            
            # Get all progress entries
            progress_domain = [('kpi_id', '=', task.id)]
            
            if from_date and to_date:
                progress_domain.extend([
                    ('create_date', '>=', from_date + ' 00:00:00'),
                    ('create_date', '<=', to_date + ' 23:59:59'),
                ])
            
            progress_records = request.env['kpi.progress'].search(
                progress_domain, order='create_date asc'
            )
            
            progress_list = []
            for prog in progress_records:
                progress_list.append({
                    'id': prog.id,
                    'date': prog.create_date.strftime('%Y-%m-%d'),
                    'date_display': prog.create_date.strftime('%d %b %Y'),
                    'time': prog.create_date.strftime('%H:%M:%S'),
                    'summary': prog.summary or '',
                    'has_attachment': bool(prog.uploaded_file),
                    'file_name': prog.file_name or '',
                })
            
            # Calculate times
            estimated_hours = (task.estimate_hours or 0) + ((task.estimate_minutes or 0) / 60.0)
            actual_hours = (task.timer_total_seconds or 0) / 3600.0
            
            return {
                'status': True,
                'task': {
                    'id': task.id,
                    'name': task.name,
                    'kra_name': task.kra_id.name if task.kra_id else '',
                    'employee_name': task.user_id.name if task.user_id else '',
                    'assigned_date': task.create_date.strftime('%d %b %Y %H:%M') if task.create_date else '',
                    'deadline': task.deadline.strftime('%d %b %Y') if task.deadline else '',
                    'priority': task.priority or 'regular',
                    'task_state': task.task_state or '',
                    'description': task.description or '',
                    'estimated_hours': estimated_hours,
                    'actual_hours': actual_hours,
                    'timer_total_seconds': task.timer_total_seconds or 0,
                    'paused_reason': task.paused_reason or '',
                    'completed_by': task.completed_by.name if task.completed_by else '',
                    'completion_date': task.completion_date.strftime('%d %b %Y %H:%M') if task.completion_date else '',
                    'approved_by': task.approved_by.name if task.approved_by else '',
                    'approval_date': task.approval_date.strftime('%d %b %Y %H:%M') if task.approval_date else '',
                },
                'progress_entries': progress_list,
                'total_entries': len(progress_list),
            }
        except Exception as e:
            import logging
            _logger = logging.getLogger(__name__)
            _logger.error(f"Error in get_employee_task_details: {str(e)}")
            return {
                'status': False,
                'message': str(e),
            }

    # ============================================================
    #                CLIENT INVOICE APIs
    # ============================================================

    @http.route('/kpi_client_invoice/get_currencies', type='json', auth='user', methods=['POST'], csrf=False)
    def get_currencies(self, **params):
        """Return the 3 invoice currencies (INR / OMR / USD).
        We hard-restrict the picker to the operator's mandated set + activate
        any of the 3 that's currently archived in the DB (so OMR works even
        on stock installs where it ships inactive).
        """
        try:
            wanted = ['INR', 'OMR', 'USD']
            Cur = request.env['res.currency'].sudo()
            currencies = Cur.with_context(active_test=False).search([('name', 'in', wanted)])
            # Activate any of the 3 that's currently flagged inactive.
            inactive = currencies.filtered(lambda c: not c.active)
            if inactive:
                inactive.write({'active': True})
            # Preserve INR -> OMR -> USD ordering.
            by_name = {c.name: c for c in currencies}
            ordered = [by_name[w] for w in wanted if w in by_name]
            return {
                'status': True,
                'currencies': [
                    {'id': c.id, 'name': c.name, 'symbol': c.symbol or '', 'position': c.position or 'after'}
                    for c in ordered
                ],
            }
        except Exception as e:
            import logging
            logging.getLogger(__name__).error(f"get_currencies: {str(e)}")
            return {'status': False, 'message': str(e), 'currencies': []}

    @http.route('/kpi_client_invoice/get_filters', type='json', auth='user', methods=['POST'], csrf=False)
    def get_client_invoice_filters(self, **params):
        """Return list of billable client KRAs for the filter dropdown.

        We list every level-2 KRA structurally (parent exists, grandparent
        does not) rather than relying on the stored ``is_client`` flag, which
        can drift out of sync if a record is created via raw SQL or if
        ``parent_id`` is set after creation.  This way, any newly added child
        of a root KRA shows up here automatically with no manual fix-up.
        """
        try:
            # Use the is_client flag (which is_client=True for root-level
            # clients under the new convention AND for legacy level-2 clients).
            clients = request.env['kra.master'].search(
                [('is_client', '=', True),
                 ('active', '=', True)],
                order='parent_id, sequence, name'
            )
            return {
                'status': True,
                'clients': [
                    {
                        'id': c.id,
                        'name': c.name,
                        'parent_name': c.parent_id.name or '',
                        'project_quoted_hours': round(c.client_quoted or 0.0, 2),
                    }
                    for c in clients
                ],
            }
        except Exception as e:
            import logging
            logging.getLogger(__name__).error(f"get_client_invoice_filters: {str(e)}")
            return {'status': False, 'message': str(e), 'clients': []}

    @http.route('/kpi_client_invoice/create', type='json', auth='user', methods=['POST'], csrf=False)
    def create_client_invoice(self, **params):
        """Create a new invoice and auto-populate its KPI lines from time logs."""
        try:
            vals = {
                'client_kra_id': int(params['client_kra_id']),
                'from_date': params['from_date'],
                'to_date': params['to_date'],
            }
            user_ids = params.get('filter_user_ids') or []
            kpi_ids = params.get('filter_kpi_ids') or []
            sub_kra_ids = params.get('filter_sub_kra_ids') or []
            if user_ids:
                vals['filter_user_ids'] = [(6, 0, [int(u) for u in user_ids])]
            if kpi_ids:
                vals['filter_kpi_ids'] = [(6, 0, [int(k) for k in kpi_ids])]
            if sub_kra_ids:
                vals['filter_sub_kra_ids'] = [(6, 0, [int(s) for s in sub_kra_ids])]
            # Optional task-type filter — list of prefixes like ['REQ', 'UPT'].
            task_types = params.get('task_types') or []
            if task_types:
                clean = [t.strip().upper() for t in task_types
                         if isinstance(t, str) and t.strip()]
                if clean:
                    vals['filter_task_types'] = ','.join(clean)
            # Optional one-shot fields the UI already passes for new invoices.
            if params.get('invoice_title'):
                vals['invoice_title'] = params['invoice_title']
            if params.get('hourly_rate') is not None:
                try:
                    vals['hourly_rate'] = float(params['hourly_rate'])
                except (TypeError, ValueError):
                    pass
            if params.get('currency_id'):
                vals['currency_id'] = int(params['currency_id'])
            # Per-task billing: set the method + the three type prices BEFORE populate,
            # so action_populate_from_logs seeds each line's unit_price by task type.
            if params.get('billing_method') in ('hourly', 'per_task'):
                vals['billing_method'] = params['billing_method']
            for pk in ('price_req', 'price_upt', 'price_bug'):
                if params.get(pk) is not None:
                    try:
                        vals[pk] = float(params[pk])
                    except (TypeError, ValueError):
                        pass
            inv = request.env['kpi.client.invoice'].create(vals)
            inv.action_populate_from_logs()
            return {'status': True, 'invoice_id': inv.id, 'data': inv._serialize_for_ui()}
        except Exception as e:
            import logging
            logging.getLogger(__name__).error(f"create_client_invoice: {str(e)}")
            return {'status': False, 'message': str(e)}

    @http.route('/kpi_client_invoice/get_sub_kras_for_client', type='json', auth='user', methods=['POST'], csrf=False)
    def list_sub_kras_for_invoice(self, **params):
        """For a chosen client KRA, return its direct child KRAs (project-level sub-KRAs).

        Distinct from get_sub_kras_for_client (which returns all descendants, used by
        the completion-cert picker). Here we want only the immediate level-2 projects
        because that is the granularity the Owner invoices at.
        """
        try:
            client = request.env['kra.master'].browse(int(params['client_kra_id']))
            if not client.exists():
                return {'status': False, 'message': 'Client KRA not found.', 'sub_kras': []}
            subs = request.env['kra.master'].search(
                [('parent_id', '=', client.id), ('active', '=', True)],
                order='sequence, name'
            )
            return {
                'status': True,
                'sub_kras': [{'id': s.id, 'name': s.name or ''} for s in subs],
            }
        except Exception as e:
            import logging
            logging.getLogger(__name__).error(f"list_sub_kras_for_invoice: {str(e)}")
            return {'status': False, 'message': str(e), 'sub_kras': []}

    @http.route('/kpi_client_invoice/get_client_users_and_kpis', type='json', auth='user', methods=['POST'], csrf=False)
    def get_client_users_and_kpis(self, **params):
        """For a chosen client KRA, return the developers + KPIs to populate filter dropdowns.

        Optional 'sub_kra_ids' narrows the search to KPIs under those sub-KRAs only.
        """
        try:
            client = request.env['kra.master'].browse(int(params['client_kra_id']))
            if not client.exists():
                return {'status': False, 'message': 'Client KRA not found.'}
            kra_ids = client._get_descendant_ids()
            sub_kra_ids = params.get('sub_kra_ids') or []
            if sub_kra_ids:
                sub_set = set()
                subs = request.env['kra.master'].browse([int(s) for s in sub_kra_ids])
                for sub in subs:
                    sub_set.update(sub._get_descendant_ids())
                kra_ids = list(sub_set & set(kra_ids))
            kpis = request.env['kra.kpi'].search([('kra_id', 'in', kra_ids)], order='name')
            user_set = set()
            for k in kpis:
                if k.user_id:
                    user_set.add((k.user_id.id, k.user_id.name or k.user_id.login or ''))
                for c in k.contributor_ids:
                    user_set.add((c.id, c.name or c.login or ''))
            # Also include anyone who has time-logs against these KPIs (might be ex-contributors)
            logs = request.env['kpi.time.log'].search([('kpi_id', 'in', kpis.ids)])
            for l in logs:
                if l.user_id:
                    user_set.add((l.user_id.id, l.user_id.name or l.user_id.login or ''))
            users = sorted([{'id': uid, 'name': name} for uid, name in user_set], key=lambda u: u['name'])
            kpi_list = [{'id': k.id, 'name': k.name, 'kra_name': k.kra_id.name or ''} for k in kpis]
            return {'status': True, 'users': users, 'kpis': kpi_list}
        except Exception as e:
            import logging
            logging.getLogger(__name__).error(f"get_client_users_and_kpis: {str(e)}")
            return {'status': False, 'message': str(e), 'users': [], 'kpis': []}

    @http.route('/kpi_client_invoice/get', type='json', auth='user', methods=['POST'], csrf=False)
    def get_client_invoice(self, **params):
        """Return a single invoice's full serialized form for the detail screen."""
        try:
            inv = request.env['kpi.client.invoice'].browse(int(params['invoice_id']))
            if not inv.exists():
                return {'status': False, 'message': 'Invoice not found.'}
            return {'status': True, 'data': inv._serialize_for_ui()}
        except Exception as e:
            import logging
            logging.getLogger(__name__).error(f"get_client_invoice: {str(e)}")
            return {'status': False, 'message': str(e)}

    @http.route('/kpi_client_invoice/list', type='json', auth='user', methods=['POST'], csrf=False)
    def list_client_invoices(self, **params):
        """List invoices. Optional filters:
          - client_kra_id: only invoices for one client
          - status_filter: 'all' (default) | 'draft' | 'finalized' | 'sent' | 'due' | 'paid'
              'due'  → state in (finalized, sent) AND payment_status = 'unpaid'
              'paid' → payment_status = 'paid'
        """
        try:
            domain = []
            if params.get('client_kra_id'):
                domain.append(('client_kra_id', '=', int(params['client_kra_id'])))

            status = (params.get('status_filter') or 'all').lower()
            if status == 'draft':
                domain.append(('state', '=', 'draft'))
            elif status == 'finalized':
                domain.append(('state', '=', 'finalized'))
            elif status == 'sent':
                domain.append(('state', '=', 'sent'))
            elif status == 'due':
                domain += [('state', 'in', ('finalized', 'sent')),
                           ('payment_status', '=', 'unpaid')]
            elif status == 'paid':
                domain.append(('payment_status', '=', 'paid'))
            # 'all' → no extra filter

            invs = request.env['kpi.client.invoice'].search(domain, order='invoice_date desc, id desc')
            # Compute summary counts (used by the filter-chip badges).
            counts = {
                'all':       request.env['kpi.client.invoice'].search_count([]),
                'draft':     request.env['kpi.client.invoice'].search_count([('state','=','draft')]),
                'finalized': request.env['kpi.client.invoice'].search_count([('state','=','finalized')]),
                'sent':      request.env['kpi.client.invoice'].search_count([('state','=','sent')]),
                'due':       request.env['kpi.client.invoice'].search_count(
                                  [('state','in',('finalized','sent')),('payment_status','=','unpaid')]),
                'paid':      request.env['kpi.client.invoice'].search_count([('payment_status','=','paid')]),
            }
            return {
                'status': True,
                'invoices': [i._serialize_for_list() for i in invs],
                'counts':   counts,
            }
        except Exception as e:
            import logging
            logging.getLogger(__name__).error(f"list_client_invoices: {str(e)}")
            return {'status': False, 'message': str(e), 'invoices': []}

    @http.route('/kpi_client_invoice/mark_paid', type='json', auth='user', methods=['POST'], csrf=False)
    def mark_invoice_paid(self, **params):
        """Mark an invoice paid (or revert paid → unpaid).  Authorized for
        owner / coordinator / system.

        params:
          invoice_id  (required)
          paid        (bool, default True)
          note        (optional)
        """
        try:
            user = request.env.user
            allowed = (
                user.has_group('base.group_system')
                or user.has_group('kra_kpi_module.group_kra_owner')
                or user.has_group('kra_kpi_module.group_kra_admin')
            )
            if not allowed:
                return {'status': False, 'message': 'Not authorized.'}

            inv = request.env['kpi.client.invoice'].browse(int(params['invoice_id']))
            if not inv.exists():
                return {'status': False, 'message': 'Invoice not found.'}
            note = (params.get('note') or '').strip()
            paid = params.get('paid', True)
            if paid:
                inv.action_mark_paid(note=note)
            else:
                inv.action_mark_unpaid(note=note)
            return {'status': True, 'data': inv._serialize_for_ui()}
        except Exception as e:
            import logging
            logging.getLogger(__name__).error(f"mark_invoice_paid: {str(e)}")
            return {'status': False, 'message': str(e)}

    @http.route('/kpi_client_invoice/save_line', type='json', auth='user', methods=['POST'], csrf=False)
    def save_invoice_line(self, **params):
        """Update editable fields on an invoice line. Allowed only when invoice is in draft."""
        try:
            line = request.env['kpi.client.invoice.line'].browse(int(params['line_id']))
            if not line.exists():
                return {'status': False, 'message': 'Line not found.'}
            line.invoice_id._guard_draft()
            # Already-billed (locked) lines are read-only — the UI disables them, and this
            # blocks a crafted request from editing them (which would take effect if the
            # owning invoice were ever reset to draft).
            if line.invoice_id._is_locked_line(line):
                return {'status': False, 'message': 'This task is already billed on another invoice and cannot be edited here.'}
            updates = {}
            for k in ('description', 'notes'):
                if k in params:
                    updates[k] = params[k]
            if 'quoted_hours' in params:
                updates['quoted_hours'] = float(params['quoted_hours'] or 0.0)
            if 'unit_price' in params:
                updates['unit_price'] = float(params['unit_price'] or 0.0)
            if updates:
                line.write(updates)
            return {'status': True, 'data': line._serialize()}
        except Exception as e:
            import logging
            logging.getLogger(__name__).error(f"save_invoice_line: {str(e)}")
            return {'status': False, 'message': str(e)}

    @http.route('/kpi_client_invoice/add_adjustment', type='json', auth='user', methods=['POST'], csrf=False)
    def add_invoice_adjustment(self, **params):
        """Append an adjustment line (positive or negative hours)."""
        try:
            inv = request.env['kpi.client.invoice'].browse(int(params['invoice_id']))
            if not inv.exists():
                return {'status': False, 'message': 'Invoice not found.'}
            inv._guard_draft()
            max_seq = max(inv.line_ids.mapped('sequence') or [0]) + 10
            line = request.env['kpi.client.invoice.line'].create({
                'invoice_id': inv.id,
                'sequence': max_seq,
                'line_type': 'adjustment',
                'description': params.get('description') or 'Adjustment',
                'quoted_hours': float(params.get('quoted_hours') or 0.0),
                'notes': params.get('notes') or '',
            })
            return {'status': True, 'line_id': line.id, 'data': line._serialize()}
        except Exception as e:
            import logging
            logging.getLogger(__name__).error(f"add_invoice_adjustment: {str(e)}")
            return {'status': False, 'message': str(e)}

    @http.route('/kpi_client_invoice/remove_line', type='json', auth='user', methods=['POST'], csrf=False)
    def remove_invoice_line(self, **params):
        """Delete an invoice line (admin only; only when invoice is in draft)."""
        try:
            line = request.env['kpi.client.invoice.line'].browse(int(params['line_id']))
            if not line.exists():
                return {'status': False, 'message': 'Line not found.'}
            line.invoice_id._guard_draft()
            line.unlink()
            return {'status': True}
        except Exception as e:
            import logging
            logging.getLogger(__name__).error(f"remove_invoice_line: {str(e)}")
            return {'status': False, 'message': str(e)}

    @http.route('/kpi_client_invoice/save_header', type='json', auth='user', methods=['POST'], csrf=False)
    def save_invoice_header(self, **params):
        """Update editable header fields (invoice_title, hourly_rate). Draft only."""
        try:
            inv = request.env['kpi.client.invoice'].browse(int(params['invoice_id']))
            if not inv.exists():
                return {'status': False, 'message': 'Invoice not found.'}
            inv._guard_draft()
            updates = {}
            if 'invoice_title' in params:
                updates['invoice_title'] = params['invoice_title'] or ''
            if 'hourly_rate' in params:
                updates['hourly_rate'] = float(params['hourly_rate'] or 0.0)
            if 'currency_id' in params:
                updates['currency_id'] = int(params['currency_id']) if params['currency_id'] else False
            if 'billing_method' in params and params['billing_method'] in ('hourly', 'per_task'):
                updates['billing_method'] = params['billing_method']
            price_touched = False
            for pk in ('price_req', 'price_upt', 'price_bug'):
                if pk in params:
                    updates[pk] = float(params[pk] or 0.0)
                    price_touched = True
            if updates:
                inv.write(updates)
            # Re-price every KPI line by its type when the method or any type price
            # changed — "change REQ price → all REQ tasks update".
            if price_touched or 'billing_method' in updates:
                inv._apply_type_prices()
            return {'status': True, 'data': inv._serialize_for_ui()}
        except Exception as e:
            import logging
            logging.getLogger(__name__).error(f"save_invoice_header: {str(e)}")
            return {'status': False, 'message': str(e)}

    @http.route('/kpi_client_invoice/pdf_data', type='json', auth='user', methods=['POST'], csrf=False)
    def get_invoice_pdf_data(self, **params):
        """Return data for client-side PDF generation."""
        try:
            inv = request.env['kpi.client.invoice'].browse(int(params['invoice_id']))
            if not inv.exists():
                return {'status': False, 'message': 'Invoice not found.'}
            return {'status': True, 'data': inv.get_pdf_data()}
        except Exception as e:
            import logging
            logging.getLogger(__name__).error(f"get_invoice_pdf_data: {str(e)}")
            return {'status': False, 'message': str(e)}

    @http.route('/kpi_client_invoice/pdf', type='json', auth='user', methods=['POST'], csrf=False)
    def get_invoice_pdf(self, **params):
        """Server-side branded invoice PDF (base64) for the mobile app, which can't
        run the web's jsPDF. Reuses get_pdf_data() (record-rule scoped, so a client
        can only render their own invoice) and renders it with reportlab in the same
        branded style as the daily/work reports."""
        try:
            inv = request.env['kpi.client.invoice'].browse(int(params.get('invoice_id') or 0))
            if not inv.exists():
                return {'status': False, 'message': 'Invoice not found.'}
            data = inv.get_pdf_data()
            pdf_bytes = self._build_invoice_pdf(data)
        except Exception as e:
            import logging, traceback
            logging.getLogger(__name__).error(f"get_invoice_pdf: {e}\n{traceback.format_exc()}")
            return {'status': False, 'message': str(e)}
        import base64
        name = (data.get('invoice', {}).get('name') or str(inv.id)).replace('/', '_')
        return {
            'status': True,
            'file_name': 'Invoice_%s.pdf' % name,
            'mimetype': 'application/pdf',
            'data_b64': base64.b64encode(pdf_bytes).decode(),
        }

    def _build_invoice_pdf(self, data):
        """Render a client invoice to PDF bytes with reportlab (Times, #38385C
        header, zebra rows, faint logo watermark) — matching the other PDFs."""
        import base64
        import io
        from xml.sax.saxutils import escape as _esc
        from reportlab.lib import colors
        from reportlab.lib.pagesizes import A4
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib.units import mm
        from reportlab.lib.utils import ImageReader
        from reportlab.platypus import (
            SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, Image, HRFlowable,
        )

        inv = data.get('invoice', {}) or {}
        cur = data.get('currency', {}) or {}
        cur_name = cur.get('name') or ''

        brand = colors.HexColor('#38385C')
        zebra = colors.HexColor('#F7F7FB')
        grid = colors.HexColor('#E1E1E1')
        rule = colors.HexColor('#BEBEBE')
        ink = colors.black
        muted = colors.HexColor('#64748B')

        def _money(amt):
            try:
                dp = int(cur.get('decimal_places', 2))
            except Exception:
                dp = 2
            s = f"{float(amt or 0):,.{dp}f}"
            return f"{cur_name} {s}".strip()

        logo_bytes = None
        if data.get('company_logo_b64'):
            try:
                logo_bytes = base64.b64decode(data['company_logo_b64'])
            except Exception:
                logo_bytes = None
        logo_reader = None
        logo_ratio = 1.0
        if logo_bytes:
            try:
                logo_reader = ImageReader(io.BytesIO(logo_bytes))
                iw, ih = logo_reader.getSize()
                if iw and ih:
                    logo_ratio = float(iw) / float(ih)
            except Exception:
                logo_reader = None

        styles = getSampleStyleSheet()
        s_company = ParagraphStyle('co', parent=styles['Normal'], fontName='Times-Bold', fontSize=22, alignment=2, textColor=ink, leading=24)
        s_doctype = ParagraphStyle('dt', parent=styles['Normal'], fontName='Times-Roman', fontSize=12, alignment=2, textColor=muted, leading=16)
        s_meta = ParagraphStyle('me', parent=styles['Normal'], fontName='Times-Roman', fontSize=10, textColor=ink, leading=15)
        s_lbl = ParagraphStyle('lb', parent=styles['Normal'], fontName='Times-Bold', fontSize=10, textColor=muted, leading=14)
        cell = ParagraphStyle('c', parent=styles['Normal'], fontName='Times-Roman', fontSize=9, textColor=ink, leading=11)
        s_tot = ParagraphStyle('t', parent=styles['Normal'], fontName='Times-Bold', fontSize=12, alignment=2, textColor=ink, leading=16)

        PW, PH = A4
        MARGIN = 48
        content_w = PW - 2 * MARGIN

        def _watermark(canvas, doc_):
            if not logo_reader:
                return
            canvas.saveState()
            wm_w = PW * 0.55
            wm_h = wm_w / (logo_ratio or 1.0)
            if wm_h > PH * 0.5:
                wm_h = PH * 0.5
                wm_w = wm_h * (logo_ratio or 1.0)
            try:
                canvas.setFillAlpha(0.08)
                canvas.setStrokeAlpha(0.08)
            except Exception:
                pass
            try:
                canvas.drawImage(logo_reader, (PW - wm_w) / 2.0, (PH - wm_h) / 2.0, wm_w, wm_h, mask='auto', preserveAspectRatio=True)
            except Exception:
                pass
            canvas.restoreState()

        buf = io.BytesIO()
        doc = SimpleDocTemplate(buf, pagesize=A4, leftMargin=MARGIN, rightMargin=MARGIN, topMargin=MARGIN, bottomMargin=MARGIN, title='Invoice')
        story = []

        right_cell = [Paragraph(_esc(data.get('company_name') or ''), s_company),
                      Paragraph('INVOICE', s_doctype)]
        if logo_reader:
            h = 56
            w = min(h * logo_ratio, 150)
            left_cell = Image(io.BytesIO(logo_bytes), width=w, height=h)
            left_cell.hAlign = 'LEFT'
        else:
            left_cell = Paragraph('', cell)
        header = Table([[left_cell, right_cell]], colWidths=[content_w * 0.45, content_w * 0.55])
        header.setStyle(TableStyle([
            ('VALIGN', (0, 0), (-1, -1), 'TOP'),
            ('LEFTPADDING', (0, 0), (-1, -1), 0), ('RIGHTPADDING', (0, 0), (-1, -1), 0),
            ('TOPPADDING', (0, 0), (-1, -1), 0), ('BOTTOMPADDING', (0, 0), (-1, -1), 0),
        ]))
        story.append(header)
        story.append(Spacer(1, 6))
        story.append(HRFlowable(width='100%', thickness=0.8, color=rule, spaceBefore=0, spaceAfter=8))

        pay = (inv.get('payment_status') or '').upper()
        story.append(Paragraph('Invoice #: <b>%s</b>%s' % (
            _esc(inv.get('name') or ''),
            ('   —   %s' % pay) if pay else ''), s_meta))
        if inv.get('invoice_title'):
            story.append(Paragraph('<b>%s</b>' % _esc(inv.get('invoice_title')), s_meta))
        story.append(Paragraph('Date: %s   •   Period: %s to %s' % (
            inv.get('invoice_date') or '', inv.get('from_date') or '', inv.get('to_date') or ''), s_meta))
        if cur_name:
            story.append(Paragraph('Currency: %s' % cur_name, s_meta))
        story.append(Spacer(1, 8))
        story.append(Paragraph('Bill To:', s_lbl))
        story.append(Paragraph('<b>%s</b>' % _esc(data.get('bill_to_name') or ''), s_meta))
        if data.get('project_name'):
            story.append(Paragraph('Project: %s' % _esc(data.get('project_name')), s_meta))
        story.append(Spacer(1, 12))

        billing = inv.get('billing_method') or 'hourly'
        has_amount = bool(data.get('has_amount'))
        rate = float(inv.get('hourly_rate') or 0)
        lines = [l for l in (inv.get('lines') or [])]

        if billing == 'per_task':
            head = ['Description', 'Type'] + (['Price'] if has_amount else [])
            rows = []
            for l in lines:
                row = [Paragraph(_esc(l.get('description')), cell), (l.get('task_kind') or l.get('line_type') or '').upper()]
                if has_amount:
                    row.append(_money(l.get('unit_price')))
                rows.append(row)
            colw = [content_w - (26 * mm) - (34 * mm if has_amount else 0), 26 * mm] + ([34 * mm] if has_amount else [])
        else:
            head = ['Description', 'Hours'] + (['Rate', 'Amount'] if has_amount else [])
            rows = []
            for l in lines:
                qh = float(l.get('quoted_hours') or 0)
                row = [Paragraph(_esc(l.get('description')), cell), f"{qh:g}"]
                if has_amount:
                    row.append(_money(rate))
                    row.append(_money(qh * rate))
                rows.append(row)
            colw = ([content_w - (22 * mm) - (34 * mm) - (34 * mm), 22 * mm, 34 * mm, 34 * mm]
                    if has_amount else [content_w - (26 * mm), 26 * mm])

        tbl = Table([head] + rows, colWidths=colw, repeatRows=1)
        tbl.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), brand),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('FONTNAME', (0, 0), (-1, -1), 'Times-Roman'),
            ('FONTNAME', (0, 0), (-1, 0), 'Times-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 9),
            ('ALIGN', (1, 0), (-1, -1), 'RIGHT'),
            ('VALIGN', (0, 0), (-1, -1), 'TOP'),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, zebra]),
            ('GRID', (0, 0), (-1, -1), 0.5, grid),
            ('TOPPADDING', (0, 0), (-1, -1), 5), ('BOTTOMPADDING', (0, 0), (-1, -1), 5),
            ('LEFTPADDING', (0, 0), (-1, -1), 6), ('RIGHTPADDING', (0, 0), (-1, -1), 6),
        ]))
        story.append(tbl)
        story.append(Spacer(1, 10))

        if billing == 'hourly':
            story.append(Paragraph('Total Hours: %g' % float(inv.get('grand_total_hours') or 0), s_tot))
        if has_amount:
            story.append(Paragraph('Total Amount: %s' % _money(inv.get('total_amount')), s_tot))

        doc.build(story, onFirstPage=_watermark, onLaterPages=_watermark)
        return buf.getvalue()

    @http.route('/kpi_client_invoice/save_notes', type='json', auth='user', methods=['POST'], csrf=False)
    def save_invoice_notes(self, **params):
        """Update the invoice's overall notes (Html). Draft only."""
        try:
            inv = request.env['kpi.client.invoice'].browse(int(params['invoice_id']))
            if not inv.exists():
                return {'status': False, 'message': 'Invoice not found.'}
            inv._guard_draft()
            inv.notes = params.get('notes') or ''
            return {'status': True}
        except Exception as e:
            import logging
            logging.getLogger(__name__).error(f"save_invoice_notes: {str(e)}")
            return {'status': False, 'message': str(e)}

    @http.route('/kpi_client_invoice/finalize', type='json', auth='user', methods=['POST'], csrf=False)
    def finalize_client_invoice(self, **params):
        try:
            inv = request.env['kpi.client.invoice'].browse(int(params['invoice_id']))
            if not inv.exists():
                return {'status': False, 'message': 'Invoice not found.'}
            inv.action_finalize()
            return {'status': True, 'data': inv._serialize_for_ui()}
        except Exception as e:
            import logging
            logging.getLogger(__name__).error(f"finalize_client_invoice: {str(e)}")
            return {'status': False, 'message': str(e)}

    @http.route('/kpi_client_invoice/send', type='json', auth='user', methods=['POST'], csrf=False)
    def send_client_invoice(self, **params):
        try:
            inv = request.env['kpi.client.invoice'].browse(int(params['invoice_id']))
            if not inv.exists():
                return {'status': False, 'message': 'Invoice not found.'}
            inv.action_send()
            return {'status': True, 'data': inv._serialize_for_ui()}
        except Exception as e:
            import logging
            logging.getLogger(__name__).error(f"send_client_invoice: {str(e)}")
            return {'status': False, 'message': str(e)}

    @http.route('/kpi_client_invoice/client_mark_paid', type='json', auth='user', methods=['POST'], csrf=False)
    def client_mark_invoice_paid(self, **params):
        """Client-only: flag that they've paid this invoice → notifies admins to
        confirm. Does NOT set payment_status (only an admin's Mark Paid does).
        Clients have READ-ONLY access (record rule blocks write), so we act via
        sudo() guarded by an explicit ownership check on client_user_ids."""
        try:
            inv = request.env['kpi.client.invoice'].sudo().browse(int(params.get('invoice_id') or 0))
            if not inv.exists():
                return {'status': False, 'message': 'Invoice not found.'}
            # Ownership: the caller must be one of this client's linked users.
            if request.env.user not in inv.client_kra_id.client_user_ids:
                return {'status': False, 'message': 'Not authorized for this invoice.'}
            if inv.state not in ('finalized', 'sent'):
                return {'status': False, 'message': 'This invoice is not ready yet.'}
            if inv.payment_status == 'paid':
                return {'status': False, 'message': 'This invoice is already marked paid.'}
            inv.write({
                'client_paid_claim': True,
                'client_paid_claim_date': fields.Datetime.now(),
            })
            # Ask the admins to confirm (recipients resolve from the role matrix).
            inv._fire_invoice_notification('invoice_client_paid')
            return {'status': True, 'data': inv._serialize_for_ui()}
        except Exception as e:
            import logging
            logging.getLogger(__name__).error(f"client_mark_invoice_paid: {str(e)}")
            return {'status': False, 'message': str(e)}

    # ============================================================
    #              COMPLETION CERTIFICATE APIs
    # ============================================================

    @http.route('/kpi_completion_cert/list', type='json', auth='user', methods=['POST'], csrf=False)
    def list_kpis_for_certificate(self, **params):
        """List KPIs eligible for a Completion Certificate.

        Filters (all optional):
          client_kra_id, project_kra_id, from_date, to_date, state
          By default returns only state='completed' KPIs.
        """
        try:
            domain = []
            state = params.get('state') or 'completed'
            if state and state != 'all':
                domain.append(('task_state', '=', state))
            if params.get('client_kra_id'):
                client = request.env['kra.master'].browse(int(params['client_kra_id']))
                if client.exists():
                    kra_ids = client._get_descendant_ids()
                    domain.append(('kra_id', 'in', kra_ids))
            if params.get('project_kra_id'):
                project = request.env['kra.master'].browse(int(params['project_kra_id']))
                if project.exists():
                    kra_ids = project._get_descendant_ids()
                    domain.append(('kra_id', 'in', kra_ids))
            from_date = params.get('from_date')
            to_date = params.get('to_date')
            if from_date:
                domain.append(('completion_date', '>=', from_date + ' 00:00:00'))
            if to_date:
                domain.append(('completion_date', '<=', to_date + ' 23:59:59'))

            kpis = request.env['kra.kpi'].search(domain, order='completion_date desc, id desc')
            result = []
            for k in kpis:
                # Root + project (immediate sub-KRA)
                root = k.kra_id
                while root and root.parent_id:
                    root = root.parent_id
                project = k.kra_id if k.kra_id != root else False
                result.append({
                    'id': k.id,
                    'name': k.name or '',
                    'client_name': root.name if root else '',
                    'project_name': (project.name if project else ''),
                    'task_state': k.task_state or '',
                    'priority': k.priority or '',
                    'delivery_version': k.delivery_version or '',
                    'requirement_version': k.requirement_version or '',
                    'requirement_document_name': k.requirement_document_name or '',
                    'has_requirement_document': bool(k.requirement_document),
                    'updates_document_name': k.updates_document_name or '',
                    'has_updates_document': bool(k.updates_document),
                    'has_signed_certificate': bool(k.signed_certificate),
                    'signed_certificate_name': k.signed_certificate_name or '',
                    'signed_certificate_date': str(k.signed_certificate_date) if k.signed_certificate_date else '',
                    'developer_ids': k.developer_ids.ids,
                    'tester_ids': k.tester_ids.ids,
                    'coordinator_ids': k.coordinator_ids.ids,
                    'lead_ids': k.lead_ids.ids,
                    'actual_hours': round((k.timer_total_seconds or 0) / 3600.0, 2),
                    'primary_assignee': k.user_id.name if k.user_id else '',
                    'completion_date': k.completion_date.strftime('%Y-%m-%d') if k.completion_date else '',
                })
            return {'status': True, 'kpis': result}
        except Exception as e:
            import logging
            logging.getLogger(__name__).error(f"list_kpis_for_certificate: {str(e)}")
            return {'status': False, 'message': str(e), 'kpis': []}

    @http.route('/kpi_completion_cert/pdf_data', type='json', auth='user', methods=['POST'], csrf=False)
    def get_certificate_pdf_data(self, **params):
        """Return data for the completion certificate PDF (one KPI)."""
        try:
            kpi = request.env['kra.kpi'].browse(int(params['kpi_id']))
            if not kpi.exists():
                return {'status': False, 'message': 'KPI not found.'}
            return {'status': True, 'data': kpi.get_completion_certificate_data()}
        except Exception as e:
            import logging
            logging.getLogger(__name__).error(f"get_certificate_pdf_data: {str(e)}")
            return {'status': False, 'message': str(e)}

    # ============================================================
    #            ADMIN — CLIENT TASK QUEUE (pre-publish review)
    # ============================================================

    @http.route('/kpi_pending_queue/list', type='json', auth='user', methods=['POST'], csrf=False)
    def list_pending_queue(self, **params):
        """Admin view: the Client Task Queue — CLIENT-submitted tasks awaiting a developer.

        Client-only by design: coordinator/owner submissions also carry published=False
        but are not part of this queue (see kra.kpi.client_submitted).
        """
        try:
            kpis = request.env['kra.kpi'].sudo().search(
                [('published', '=', False), ('active', '=', True),
                 ('client_submitted', '=', True)],
                order='create_date desc')
            result = []
            req_cache = {}

            def _req_options_for(root_kra, current_ref):
                """Existing REQ-xxx tasks in this client's subtree — for the Linked Req dropdown."""
                if not root_kra:
                    opts = []
                else:
                    if root_kra.id not in req_cache:
                        desc_ids = root_kra._get_descendant_ids()
                        reqs = request.env['kra.kpi'].sudo().search([
                            ('kra_id', 'in', desc_ids),
                            ('external_ref', 'ilike', 'REQ%'),
                            ('active', '=', True),
                        ], order='external_ref')
                        req_cache[root_kra.id] = [
                            {'ref': r.external_ref, 'name': r.name or ''}
                            for r in reqs if r.external_ref]
                    opts = list(req_cache[root_kra.id])
                cur = (current_ref or '').strip()
                if cur and not any(o['ref'] == cur for o in opts):
                    # keep an unknown existing value selectable so we never drop data
                    opts = [{'ref': cur, 'name': '(current — not a known REQ)'}] + opts
                return [{'ref': o['ref'],
                         'label': o['ref'] + (' — ' + o['name'][:34] if o['name'] else '')}
                        for o in opts]

            for k in kpis:
                root = k.kra_id
                while root and root.parent_id:
                    root = root.parent_id
                project = k.kra_id if k.kra_id != root else False
                nm = k.name or ''
                ref = (k.external_ref or '').upper()
                if nm.startswith('[Update]') or ref.startswith('UPD'):
                    kind = 'update'
                elif nm.startswith('[Bug]') or ref.startswith('BUG'):
                    kind = 'bug'
                else:
                    kind = 'requirement'
                result.append({
                    'id': k.id,
                    'name': k.name or '',
                    'kind': kind,
                    'client_name': root.name if root else '',
                    'project_name': project.name if project else '',
                    'project_id': project.id if project else (root.id if root else False),
                    'external_ref': k.external_ref or '',
                    'related_req_ref': k.related_req_ref or '',
                    'priority': k.priority or 'regular',
                    'description': k.description or '',
                    'estimate_hours': k.estimate_hours or 0,
                    'estimate_minutes': k.estimate_minutes or 0,
                    'user_id': k.user_id.id if k.user_id else False,
                    'user_name': k.user_id.name if k.user_id else '',
                    'submitted_by': k.submitted_by_uid.name if k.submitted_by_uid else (
                        k.create_uid.name if k.create_uid else 'Unknown'),
                    'submitted_login': k.submitted_by_uid.login if k.submitted_by_uid else (
                        k.create_uid.login if k.create_uid else ''),
                    'received_at': k.create_date.strftime('%Y-%m-%d %H:%M') if k.create_date else '',
                    'req_options': _req_options_for(root, k.related_req_ref),
                    # Came BACK to the queue because the client rejected the
                    # developer we assigned. Without this the card looks identical
                    # to a fresh submission and the admin can't see why it returned.
                    # 'rework' + decision 'reject' is exactly the state
                    # _return_to_queue_after_reject leaves behind.
                    # (The previous developer isn't shown: that flow clears user_id
                    # by design — the notification names them, the card can't.)
                    'rejected': (k.task_state == 'rework'
                                 and k.pre_approval_decision == 'reject'),
                    'reject_reason': k.pre_approval_feedback or '',
                    'rejected_by': k.pre_approval_decided_by_name or '',
                    'rejected_at': (k.pre_approval_decided_at.strftime('%Y-%m-%d %H:%M')
                                    if k.pre_approval_decided_at else ''),
                    # Photos / video the submitter attached, each with their reason
                    # (kpi.user.manual.description). Metadata only — never the file
                    # bytes, or a queue with videos in it would be unusable.
                    'attachments': [{
                        'id': m.id,
                        'file_name': m.file_name or 'file',
                        'reason': m.description or '',
                    } for m in k.user_manual_ids],
                })
            return {'status': True, 'queue': result}
        except Exception as e:
            import logging
            logging.getLogger(__name__).error(f"list_pending_queue: {str(e)}")
            return {'status': False, 'message': str(e), 'queue': []}

    @http.route('/kpi_pending_queue/update', type='json', auth='user', methods=['POST'], csrf=False)
    def update_pending_kpi(self, **params):
        """Admin updates editable fields on a pending KPI before publishing."""
        try:
            if not (request.env.user.has_group('kra_kpi_module.group_kra_admin')
                    or request.env.user.has_group('base.group_system')):
                return {'status': False, 'message': 'Admin only.'}
            kpi = request.env['kra.kpi'].sudo().browse(int(params['kpi_id']))
            if not kpi.exists() or kpi.published:
                return {'status': False, 'message': 'Pending task not found.'}
            updates = {}
            for fld in ('name', 'description', 'external_ref', 'related_req_ref', 'priority'):
                if fld in params:
                    updates[fld] = params[fld] or ''
            if 'estimate_hours' in params:
                updates['estimate_hours'] = int(params['estimate_hours'] or 0)
            if 'estimate_minutes' in params:
                updates['estimate_minutes'] = int(params['estimate_minutes'] or 0)
            if 'user_id' in params:
                updates['user_id'] = int(params['user_id']) if params['user_id'] else False
            if updates:
                kpi.write(updates)
            return {'status': True}
        except Exception as e:
            import logging
            logging.getLogger(__name__).error(f"update_pending_kpi: {str(e)}")
            return {'status': False, 'message': str(e)}

    @http.route('/kpi_pending_queue/publish', type='json', auth='user', methods=['POST'], csrf=False)
    def publish_pending_kpi(self, **params):
        """Admin publishes a pending KPI — sets published=True so it becomes visible to devs.

        Idempotent: if the row was already auto-published by a write override
        (e.g. because the coordinator just assigned a developer, which fires the
        pre-approval pipeline and flips published=True), we still return success.
        """
        try:
            if not (request.env.user.has_group('kra_kpi_module.group_kra_admin')
                    or request.env.user.has_group('base.group_system')):
                return {'status': False, 'message': 'Admin only.'}
            kpi = request.env['kra.kpi'].sudo().browse(int(params['kpi_id']))
            if not kpi.exists():
                return {'status': False, 'message': 'Task not found.'}
            # A developer is MANDATORY before publishing: the client's pre-approval
            # asks them to sign off on WHO the task went to, so there has to be a
            # developer to name.
            if not kpi.user_id:
                return {'status': False,
                        'message': 'Assign a developer before publishing this task.'}
            if kpi.published:
                # Already published — either by an earlier publish OR by the
                # write() override after a developer assignment.  Treat as a
                # no-op success so the UI doesn't show a confusing error.
                msg = (f'Already published: {kpi.name}'
                       + (' — client pre-approval triggered.'
                          if kpi.task_state == 'pre_approval_pending' else ''))
                return {'status': True, 'message': msg}
            kpi.write({'published': True})
            # Ask the client to approve the assignment.  The client decides in
            # their own time — silence is never treated as consent.  The write()
            # override only fires this when a developer is NEWLY assigned, so a
            # task that ALREADY had a developer would otherwise be published
            # without ever asking the client.
            #
            # admin_accepted is required: publishing must not leapfrog the admin
            # gate, which is the first step of the workflow.
            msg = f'Published: {kpi.name}'
            if kpi.task_state in ('assigned', 'queue_waiting') and not kpi.admin_accepted:
                # Say WHEN, not just that it is waiting — acceptance now happens
                # by itself, so "waiting for admin acceptance" alone would read
                # as "go find an admin" when in fact nobody needs to do anything.
                msg += ' — the client is asked once the task is accepted'
                if kpi.admin_accept_deadline_at:
                    msg += (' (automatically by %s if no admin acts)'
                            % kpi.admin_accept_deadline_at.strftime('%H:%M'))
                msg += '.'
            elif kpi.task_state in ('assigned', 'queue_waiting'):
                try:
                    kpi.request_client_pre_approval()
                    msg += ' — client pre-approval requested.'
                except Exception as exc:
                    import logging
                    logging.getLogger(__name__).warning(
                        "publish_pending_kpi: pre-approval request failed for %s: %s",
                        kpi.id, exc)
            return {'status': True, 'message': msg}
        except Exception as e:
            import logging
            logging.getLogger(__name__).error(f"publish_pending_kpi: {str(e)}")
            return {'status': False, 'message': str(e)}

    @http.route('/kpi_task_approval/list', type='json', auth='user', methods=['POST'], csrf=False)
    def task_approval_list(self, **params):
        """Return active tasks split into two groups:
          - approved_started   : client approved (token_used=True) AND task_state
                                 is in_progress / paused / partially_completed / awaiting_client / completed
          - unapproved_started : client hasn't approved yet (token_used=False)
                                 AND task_state IS a "started" state (in_progress / paused / partially_completed / awaiting_client)

        The second list is what the admin acts on — each row gets a manual
        "Send Approval Request" button that re-fires pre_approval_request.
        """
        try:
            if not (request.env.user.has_group('kra_kpi_module.group_kra_admin')
                    or request.env.user.has_group('kra_kpi_module.group_kra_owner')
                    or request.env.user.has_group('base.group_system')):
                return {'status': False, 'message': 'Admin / Owner only.'}
            started_states = ('in_progress', 'paused', 'partially_completed', 'awaiting_client')
            Kpi = request.env['kra.kpi'].sudo()

            approved = Kpi.search([
                ('active', '=', True),
                ('task_state', 'in', started_states + ('completed',)),
                ('approval_token_used', '=', True),
            ], order='write_date desc', limit=200)

            unapproved = Kpi.search([
                ('active', '=', True),
                ('task_state', 'in', started_states),
                ('approval_token_used', '=', False),
            ], order='write_date desc', limit=200)

            def _serialize(kpi):
                client_kra = getattr(kpi, 'client_kra_id', False) or kpi.kra_id
                clients = client_kra.client_user_ids if client_kra else False
                notified = kpi.pre_approval_notified_user_ids
                pending = clients.filtered(lambda u: u.id not in notified.ids) if clients else False
                current_client = notified[-1:] if notified else (clients[:1] if clients else False)
                next_client = pending[:1] if pending else False
                return {
                    'id':                kpi.id,
                    'external_ref':      kpi.external_ref or '',
                    'name':              kpi.name or '',
                    'task_state':        kpi.task_state,
                    'kra_name':          kpi.kra_id.name if kpi.kra_id else '',
                    'client_name':       client_kra.name if client_kra else '',
                    'developer':         kpi.user_id.name if kpi.user_id else '(unassigned)',
                    'current_client':    (current_client[0].name if current_client else ''),
                    'current_client_login': (current_client[0].login if current_client else ''),
                    'next_client':       (next_client[0].name if next_client else ''),
                    'notified_count':    len(notified),
                    'total_clients':     len(clients) if clients else 0,
                    'next_check_at':     kpi.pre_approval_next_check_at and kpi.pre_approval_next_check_at.isoformat() or '',
                    'approval_token_used': kpi.approval_token_used,
                    'approval_date':     kpi.approval_date and kpi.approval_date.isoformat() or '',
                    'decided_by':        kpi.pre_approval_decided_by_name or '',
                }

            return {
                'status': True,
                'approved_started':   [_serialize(k) for k in approved],
                'unapproved_started': [_serialize(k) for k in unapproved],
            }
        except Exception as e:
            import logging
            logging.getLogger(__name__).error(f"task_approval_list: {str(e)}")
            return {'status': False, 'message': str(e)}

    @http.route('/kpi_task_approval/get_available_clients', type='json', auth='user', methods=['POST'], csrf=False)
    def task_approval_get_available_clients(self, **params):
        """List every user in group_kra_client so admin can pick who to
        assign to a KRA that currently has no client set."""
        try:
            if not (request.env.user.has_group('kra_kpi_module.group_kra_admin')
                    or request.env.user.has_group('kra_kpi_module.group_kra_owner')
                    or request.env.user.has_group('base.group_system')):
                return {'status': False, 'message': 'Admin / Owner only.'}
            grp = request.env.ref('kra_kpi_module.group_kra_client', raise_if_not_found=False)
            users = grp.user_ids if grp else request.env['res.users']
            out = []
            for u in users.sorted('name'):
                p = u.partner_id
                out.append({
                    'id':    u.id,
                    'name':  u.name,
                    'login': u.login,
                    'phone': (p and (getattr(p, 'mobile', None) or p.phone) or ''),
                })
            return {'status': True, 'clients': out}
        except Exception as e:
            import logging
            logging.getLogger(__name__).error(f"task_approval_get_available_clients: {str(e)}")
            return {'status': False, 'message': str(e)}

    @http.route('/kpi_task_approval/set_kra_clients', type='json', auth='user', methods=['POST'], csrf=False)
    def task_approval_set_kra_clients(self, **params):
        """Assign a set of client users to the KRA that owns this KPI so the
        pre_approval_request can resolve to real recipients.  Also clears the
        task's notified state so the next send starts from client #1."""
        try:
            if not (request.env.user.has_group('kra_kpi_module.group_kra_admin')
                    or request.env.user.has_group('kra_kpi_module.group_kra_owner')
                    or request.env.user.has_group('base.group_system')):
                return {'status': False, 'message': 'Admin / Owner only.'}
            kpi = request.env['kra.kpi'].sudo().browse(int(params['kpi_id']))
            if not kpi.exists():
                return {'status': False, 'message': 'Task not found.'}
            client_kra = getattr(kpi, 'client_kra_id', False) or kpi.kra_id
            if not client_kra:
                return {'status': False, 'message': 'Task has no KRA.'}
            user_ids = [int(x) for x in (params.get('client_user_ids') or [])]
            if not user_ids:
                return {'status': False, 'message': 'Pick at least one client.'}
            client_kra.sudo().write({
                'client_user_ids': [(6, 0, user_ids)],
            })
            # Reset the task's escalation state so next send starts from client #1.
            kpi.sudo().write({
                'pre_approval_notified_user_ids': [(5, 0, 0)],
                'pre_approval_next_check_at': False,
            })
            kpi._log_action('kra_clients_assigned', source='web',
                            actor_user_id=request.env.user.id,
                            payload={'kra_id': client_kra.id,
                                     'client_user_ids': user_ids})
            return {'status': True,
                    'message': f'Assigned {len(user_ids)} client(s) to {client_kra.name}.'}
        except Exception as e:
            import logging
            logging.getLogger(__name__).error(f"task_approval_set_kra_clients: {str(e)}")
            return {'status': False, 'message': str(e)}

    @http.route('/kpi_task_approval/resend', type='json', auth='user', methods=['POST'], csrf=False)
    def task_approval_resend(self, **params):
        """Admin manually fires pre_approval_request on a task where the dev
        has already started but the client hasn't decided yet.  Advances the
        sequential escalation to the next non-notified client (or wraps back
        to the first if all clients have been tried).
        """
        try:
            if not (request.env.user.has_group('kra_kpi_module.group_kra_admin')
                    or request.env.user.has_group('kra_kpi_module.group_kra_owner')
                    or request.env.user.has_group('base.group_system')):
                return {'status': False, 'message': 'Admin / Owner only.'}
            kpi = request.env['kra.kpi'].sudo().browse(int(params['kpi_id']))
            if not kpi.exists():
                return {'status': False, 'message': 'Task not found.'}
            if kpi.approval_token_used:
                return {'status': False, 'message': 'Token already used (client has decided).'}
            client_kra = getattr(kpi, 'client_kra_id', False) or kpi.kra_id
            clients = client_kra.client_user_ids if client_kra else False
            if not clients:
                return {'status': False,
                        'message': 'No client users on this KRA — cannot send.'}
            wrap_flag = params.get('wrap_around', True)
            notified_ids = set(kpi.pre_approval_notified_user_ids.ids)
            pending = clients.filtered(lambda u: u.id not in notified_ids)
            if not pending and wrap_flag:
                # All clients exhausted — wrap: clear notified so next _notify picks
                # client #1 again.  Admin-initiated resend restarts the cycle.
                kpi.write({
                    'pre_approval_notified_user_ids': [(5, 0, 0)],
                    'pre_approval_next_check_at': False,
                })
            target_before = kpi.pre_approval_notified_user_ids
            try:
                kpi._notify('pre_approval_request')
            except Exception as exc:
                kpi._log_action(
                    'notification_failed', success=False,
                    payload={'event': 'pre_approval_request_manual_resend', 'error': str(exc)})
                return {'status': False, 'message': f'Notify raised: {exc}'}
            target_after = kpi.pre_approval_notified_user_ids
            picked = (target_after - target_before)
            picked_name = picked[0].name if picked else (target_after and target_after[-1].name or '?')
            kpi._log_action('pre_approval_manual_resend', source='web',
                            actor_user_id=request.env.user.id,
                            payload={'to': picked_name,
                                     'task_state': kpi.task_state,
                                     'notified_count': len(target_after)})
            return {'status': True,
                    'message': f'Sent to {picked_name}.',
                    'picked': picked_name}
        except Exception as e:
            import logging
            logging.getLogger(__name__).error(f"task_approval_resend: {str(e)}")
            return {'status': False, 'message': str(e)}

    @http.route('/kpi_pending_queue/resend_pre_approval', type='json', auth='user', methods=['POST'], csrf=False)
    def resend_pre_approval(self, **params):
        """Re-fire the pre-approval notification for a task that's already in
        `pre_approval_pending` state.  Useful when the original WhatsApp send
        failed (e.g. session was reconnecting) and we want to retry without
        having to re-assign the developer.
        """
        try:
            if not (request.env.user.has_group('kra_kpi_module.group_kra_admin')
                    or request.env.user.has_group('base.group_system')):
                return {'status': False, 'message': 'Admin only.'}
            kpi = request.env['kra.kpi'].sudo().browse(int(params['kpi_id']))
            if not kpi.exists():
                return {'status': False, 'message': 'Task not found.'}
            # Allow resend on either pending OR partial state.  In partial state
            # we fire BOTH events so the client gets another chance to approve
            # AND the developer/owner get the prep-mode reminder.
            if kpi.task_state == 'pre_approval_pending':
                events = ['pre_approval_request']
            elif kpi.task_state == 'pre_approval_partial':
                events = ['pre_approval_request', 'partial_timeout']
            else:
                return {'status': False,
                        'message': f'Task is in state {kpi.task_state} — '
                                   'resend only available for pre_approval_pending / pre_approval_partial.'}
            if kpi.approval_token_used:
                return {'status': False, 'message': 'Token already used (client has decided).'}
            # Admin-triggered resend restarts sequential client escalation from
            # the first client.  Clear the notified set + next-check so the
            # next pre_approval_request picks client #1 again.
            kpi.sudo().write({
                'pre_approval_notified_user_ids': [(5, 0, 0)],
                'pre_approval_next_check_at': False,
            })
            kpi._log_action('pre_approval_resent', source='web',
                            actor_user_id=request.env.user.id,
                            payload={'events': events, 'state': kpi.task_state})
            for event in events:
                try:
                    kpi._notify(event)
                except Exception as exc:
                    kpi._log_action(
                        'notification_failed', success=False,
                        payload={'event': f'{event}_resend', 'error': str(exc)})
            return {'status': True,
                    'message': f'Re-sent ({", ".join(events)}) for {kpi.name}.'}
        except Exception as e:
            import logging
            logging.getLogger(__name__).error(f"resend_pre_approval: {str(e)}")
            return {'status': False, 'message': str(e)}

    @http.route('/kpi_pending_queue/delete', type='json', auth='user', methods=['POST'], csrf=False)
    def delete_pending_kpi(self, **params):
        """Admin discards a pending KPI (rejected submission)."""
        try:
            if not (request.env.user.has_group('kra_kpi_module.group_kra_admin')
                    or request.env.user.has_group('base.group_system')):
                return {'status': False, 'message': 'Admin only.'}
            kpi = request.env['kra.kpi'].sudo().browse(int(params['kpi_id']))
            if not kpi.exists() or kpi.published:
                return {'status': False, 'message': 'Pending task not found (or already published).'}
            # Tell the client their submission was dropped — otherwise it just
            # silently disappears from their list with no explanation.
            #
            # The row is created with kpi_id=False ON PURPOSE: kpi.notification.kpi_id
            # is ondelete='cascade', so a notification pointing at this task would be
            # deleted along with it by the unlink() below and the client would never
            # see it. Detached also means it isn't tappable — correct, since the task
            # is gone. Never blocks the discard.
            try:
                if kpi.client_submitted and kpi.submitted_by_uid:
                    reason = (params.get('reason') or '').strip()
                    body = ("🗑 Your submitted task was not taken forward.\n"
                            "📌 %s\n" % (kpi.name or ''))
                    if reason:
                        body += "💬 Reason: %s\n" % reason
                    body += "Contact your coordinator if you think this is a mistake."
                    request.env['kpi.notification'].sudo().create({
                        'user_id': kpi.submitted_by_uid.id,
                        'kpi_id': False,
                        'event': 'client_task_discarded',
                        'role': 'client',
                        'title': 'Task discarded',
                        'body': body,
                    })
                    # Push is an HTTP call, not a DB row, so it survives the unlink.
                    kpi._send_push('client_task_discarded', 'Task discarded', body,
                                   [kpi.submitted_by_uid.id])
            except Exception as exc:
                import logging
                logging.getLogger(__name__).warning(
                    "discard notify failed for kpi %s: %s", kpi.id, exc)
            kpi.unlink()
            return {'status': True}
        except Exception as e:
            import logging
            logging.getLogger(__name__).error(f"delete_pending_kpi: {str(e)}")
            return {'status': False, 'message': str(e)}

    # ============================================================
    #            CLIENT PORTAL — DASHBOARD + ADD TASK
    # ============================================================

    def _resolve_client_kras_for_user(self):
        """Return the recordset of root KRAs the current user has client-portal access to."""
        return request.env['kra.master'].sudo().search([
            ('client_user_ids', 'in', [request.env.user.id]),
        ])

    def _is_current_user_client_only(self):
        """True if the current user is a Client (no admin/system rights)."""
        user = request.env.user
        if user.has_group('kra_kpi_module.group_kra_admin') or user.has_group('base.group_system'):
            return False
        return user.has_group('kra_kpi_module.group_kra_client')

    @http.route('/kpi_client_portal/dashboard', type='json', auth='user', methods=['POST'], csrf=False)
    def client_portal_dashboard(self, **params):
        """Return a dashboard payload for the logged-in client user:
          - their client KRAs + projects
          - all KPIs (with status filter)
          - pending approvals (partially_completed) — call-out
          - completed KPIs with signed cert (downloadable cert)
          - finalized invoices count
        """
        try:
            user = request.env.user
            client_kras = self._resolve_client_kras_for_user()
            if not client_kras:
                return {
                    'status': True,
                    'authorized': False,
                    'message': 'Your user is not linked to any client. Please contact your admin.',
                    'clients': [], 'kpis': [], 'pending_approvals': [], 'totals': {},
                }
            kra_ids = set()
            for k in client_kras:
                kra_ids.update(k._get_descendant_ids())
            kra_ids = list(kra_ids)

            state_filter = params.get('state') or ''
            # Only show the client tasks an admin has ACCEPTED into the official flow.
            # A freshly-created / self-created task (admin_accepted=False) must not
            # reach the client until an admin accepts it — matches the workflow rule.
            domain = [('kra_id', 'in', kra_ids), ('active', '=', True), ('admin_accepted', '=', True)]
            if state_filter and state_filter != 'all':
                domain.append(('task_state', '=', state_filter))

            kpis = request.env['kra.kpi'].sudo().search(domain, order='create_date desc, id desc')

            def _kpi_summary(k):
                root = k.kra_id
                while root and root.parent_id:
                    root = root.parent_id
                project = k.kra_id if k.kra_id != root else False
                kind = 'requirement'
                nm = k.name or ''
                ref = (k.external_ref or '').upper()
                if nm.startswith('[Update]') or ref.startswith('UPD'):
                    kind = 'update'
                elif nm.startswith('[Bug]') or ref.startswith('BUG'):
                    kind = 'bug'
                return {
                    'id': k.id,
                    'name': k.name or '',
                    'kind': kind,
                    'external_ref': k.external_ref or '',
                    'related_req_ref': k.related_req_ref or '',
                    'client_name': root.name if root else '',
                    'project_id': project.id if project else (k.kra_id.id if k.kra_id else 0),
                    'project_name': project.name if project else '',
                    'task_state': k.task_state or '',
                    'priority': k.priority or '',
                    'delivery_version': k.delivery_version or '',
                    'estimate_hours': round(((k.estimate_hours or 0) + (k.estimate_minutes or 0) / 60.0), 2),
                    'actual_hours': round((k.timer_total_seconds or 0) / 3600.0, 2),
                    'deadline': str(k.deadline) if k.deadline else '',
                    'completion_date': k.completion_date.strftime('%Y-%m-%d') if k.completion_date else '',
                    'has_signed_certificate': bool(k.signed_certificate),
                    'client_signed': bool(k.client_signature_text),
                    'primary_assignee_id': k.user_id.id if k.user_id else 0,
                    'primary_assignee': k.user_id.name if k.user_id else '',
                }

            all_summaries = [_kpi_summary(k) for k in kpis]

            # Pending approvals: tasks that admin has QA-approved (state='awaiting_client').
            # Also include legacy 'partially_completed' rows where admin_approved=True (back-compat).
            pending_kpis = request.env['kra.kpi'].sudo().search([
                ('kra_id', 'in', kra_ids),
                ('active', '=', True),
                '|',
                ('task_state', '=', 'awaiting_client'),
                '&', ('task_state', '=', 'partially_completed'), ('admin_approved', '=', True),
            ], order='completion_date desc')
            pending = [_kpi_summary(k) for k in pending_kpis]

            # Totals (across all admin-accepted tasks — un-accepted ones are hidden
            # from the client everywhere, same as the list above).
            all_active = request.env['kra.kpi'].sudo().search(
                [('kra_id', 'in', kra_ids), ('active', '=', True), ('admin_accepted', '=', True)])

            def _kind_of(k):
                nm = k.name or ''
                ref = (k.external_ref or '').upper()
                if nm.startswith('[Update]') or ref.startswith('UPD'):
                    return 'update'
                if nm.startswith('[Bug]') or ref.startswith('BUG'):
                    return 'bug'
                return 'requirement'

            totals = {
                'all': len(all_active),
                'assigned': len(all_active.filtered(lambda k: k.task_state == 'assigned')),
                'in_progress': len(all_active.filtered(lambda k: k.task_state == 'in_progress')),
                'paused': len(all_active.filtered(lambda k: k.task_state == 'paused')),
                'partially_completed': len(all_active.filtered(lambda k: k.task_state == 'partially_completed')),
                'completed': len(all_active.filtered(lambda k: k.task_state == 'completed')),
            }

            # Per-type breakdown (req/update/bug × open/done) + time totals
            by_type = {
                'requirement': {'total': 0, 'done': 0, 'open': 0, 'hours': 0.0},
                'update':      {'total': 0, 'done': 0, 'open': 0, 'hours': 0.0},
                'bug':         {'total': 0, 'done': 0, 'open': 0, 'hours': 0.0},
            }
            for k in all_active:
                kind = _kind_of(k)
                bucket = by_type[kind]
                bucket['total'] += 1
                if k.task_state == 'completed':
                    bucket['done'] += 1
                else:
                    bucket['open'] += 1
                bucket['hours'] += (k.timer_total_seconds or 0) / 3600.0
            for kind in by_type:
                by_type[kind]['hours'] = round(by_type[kind]['hours'], 2)

            # Invoices
            invs = request.env['kpi.client.invoice'].sudo().search([
                ('client_kra_id', 'in', client_kras.ids),
                ('state', 'in', ['finalized', 'sent']),
            ])
            invoice_count = len(invs)

            return {
                'status': True,
                'authorized': True,
                'clients': [{'id': c.id, 'name': c.name} for c in client_kras],
                'kpis': all_summaries,
                'pending_approvals': pending,
                'totals': totals,
                'by_type': by_type,
                'invoice_count': invoice_count,
                'state_filter': state_filter or 'all',
            }
        except Exception as e:
            import logging
            logging.getLogger(__name__).error(f"client_portal_dashboard: {str(e)}")
            return {'status': False, 'message': str(e)}

    @http.route('/kpi_client_portal/get_projects', type='json', auth='user', methods=['POST'], csrf=False)
    def client_portal_get_projects(self, **params):
        """List projects (sub-KRAs) under the client user's KRAs — used by 'New Task' form."""
        try:
            user_kras = self._resolve_client_kras_for_user()
            kra_ids = set()
            for k in user_kras:
                kra_ids.update(k._get_descendant_ids())
            kras = request.env['kra.master'].sudo().browse(list(kra_ids))
            result = []
            for k in kras:
                path = []
                cur = k
                while cur:
                    path.insert(0, cur.name)
                    cur = cur.parent_id
                result.append({'id': k.id, 'display': ' > '.join(path)})
            result.sort(key=lambda d: d['display'])
            return {'status': True, 'projects': result}
        except Exception as e:
            return {'status': False, 'message': str(e), 'projects': []}

    @http.route('/kpi_client_portal/add_task', type='json', auth='user', methods=['POST'], csrf=False)
    def client_portal_add_task(self, **params):
        """Allow a client portal user to self-submit a new task under one of their projects.
        Created in 'assigned' state without a user_id — admin will pick that up and reassign."""
        try:
            sub_kra_id = int(params.get('sub_kra_id') or 0)
            if not sub_kra_id:
                return {'status': False, 'message': 'Please pick a project.'}
            sub_kra = request.env['kra.master'].sudo().browse(sub_kra_id)
            if not sub_kra.exists():
                return {'status': False, 'message': 'Project not found.'}
            # Verify the user is linked to a KRA that contains this sub_kra
            user_kras = self._resolve_client_kras_for_user()
            allowed_ids = set()
            for k in user_kras:
                allowed_ids.update(k._get_descendant_ids())
            if sub_kra_id not in allowed_ids:
                return {'status': False, 'message': 'You are not authorized for this project.'}
            kind = params.get('kind') or 'requirement'
            name_prefix = {'requirement': '', 'update': '[Update] ', 'bug': '[Bug] '}.get(kind, '')
            raw_name = (params.get('name') or '').strip()
            if not raw_name:
                return {'status': False, 'message': 'Task name required.'}
            vals = {
                'name': name_prefix + raw_name,
                'kra_id': sub_kra_id,
                'task_state': 'assigned',
                'priority': params.get('priority') or 'regular',
                'description': params.get('description') or '',
                'external_ref': params.get('external_ref') or '',
                'related_req_ref': params.get('related_req_ref') or '',
                'active': True,
                # Client-submitted tasks land in admin's pending queue, unpublished.
                'published': False,
                'submitted_by_uid': request.env.user.id,
                # This route is the client portal — always a client submission.
                'client_submitted': True,
            }
            kpi = request.env['kra.kpi'].sudo().create(vals)
            # Tell the admins a client task is queued (re-nudged while unread).
            try:
                kpi._notify('client_task_queued')
                kpi._schedule_queue_nudge()
            except Exception as exc:
                import logging
                logging.getLogger(__name__).warning(
                    "client_task_queued notify failed for kpi %s: %s", kpi.id, exc)
            return {'status': True, 'kpi_id': kpi.id, 'name': kpi.name}
        except Exception as e:
            import logging
            logging.getLogger(__name__).error(f"client_portal_add_task: {str(e)}")
            return {'status': False, 'message': str(e)}

    @http.route('/kpi_client_portal/download_cert', type='http', auth='user')
    def client_portal_download_cert(self, kpi_id=None, **kw):
        """Stream the signed certificate file (if any) for a KPI in the client's scope."""
        import base64
        try:
            if not kpi_id:
                return request.not_found()
            kpi = request.env['kra.kpi'].sudo().browse(int(kpi_id))
            if not kpi.exists() or not kpi.signed_certificate:
                return request.not_found()
            # Check the user has portal access to this KPI's client KRA
            user_kras = self._resolve_client_kras_for_user()
            kra_ids = set()
            for k in user_kras:
                kra_ids.update(k._get_descendant_ids())
            if kpi.kra_id.id not in kra_ids:
                return request.not_found()
            data = base64.b64decode(kpi.signed_certificate)
            fname = kpi.signed_certificate_name or f'signed_cert_{kpi.id}.txt'
            return request.make_response(data, headers=[
                ('Content-Type', 'application/octet-stream'),
                ('Content-Disposition', f'attachment; filename="{fname}"'),
            ])
        except Exception:
            return request.not_found()

    # ============================================================
    #            CURRENT USER — ROLE FLAGS
    # ============================================================

    @http.route('/kpi_user/info', type='json', auth='user', methods=['POST'], csrf=False)
    def kpi_user_info(self, **params):
        """Return the current user's role flags so the OWL UI can adapt
        (e.g. hide developer/time fields from client-only users)."""
        try:
            user = request.env.user
            is_system = user.has_group('base.group_system')
            is_owner = user.has_group('kra_kpi_module.group_kra_owner')
            is_admin = user.has_group('kra_kpi_module.group_kra_admin')
            is_developer = user.has_group('kra_kpi_module.group_kra_developer')
            is_client = user.has_group('kra_kpi_module.group_kra_client')
            # "client-only" = client and NOT admin (since admin implies dev, owner implies admin)
            is_client_only = is_client and not (is_admin or is_system or is_owner)
            return {
                'status': True,
                'user_id': user.id,
                'name': user.name or '',
                'login': user.login or '',
                'is_system': is_system,
                'is_owner': is_owner,
                'is_admin': is_admin,
                'is_developer': is_developer,
                'allow_multitask': user.kpi_allow_multitask,
                'is_client': is_client,
                'is_client_only': is_client_only,
                # THE KRA/KPI role — 'admin' | 'client' | 'developer' — the single
                # value shown in Login Management's role dropdown
                # (res.users.kpi_role / _compute_kpi_role). Prefer this over the
                # boolean flags above: one field, same answer everywhere, and it
                # can't drift the way a hand-rolled combination of flags can.
                # NB: the selection's third value is 'developer' (shown as "User"
                # in the UI) — not 'user'.
                'kpi_role': user.kpi_role or 'developer',
                # True while the app password is still the default 1111 (admin reset
                # / freshly created) — drives the red "change your password" nag on
                # Profile until they set their own via Forgot Password.
                'must_change': bool(user.kpi_app_must_change_password),
            }
        except Exception as e:
            import logging
            logging.getLogger(__name__).error(f"kpi_user_info: {str(e)}")
            return {'status': False, 'message': str(e)}

    # ============================================================
    #            OWNER DASHBOARD — CLIENT-WISE BILLABLE
    # ============================================================

    @http.route('/kpi_owner/certified_by_client', type='json', auth='user', methods=['POST'], csrf=False)
    def owner_certified_by_client(self, **params):
        """Return per-client summary of signed-off (completed) tasks: count + hours + last
        invoice. Used by the Owner Dashboard to decide what to invoice next."""
        try:
            Kra = request.env['kra.master'].sudo()
            Kpi = request.env['kra.kpi'].sudo()
            Inv = request.env['kpi.client.invoice'].sudo()

            # Use the is_client flag (root-level clients under the new
            # convention + legacy level-2 clients both set is_client=True).
            client_kras = Kra.search(
                [('is_client', '=', True),
                 ('active', '=', True)],
                order='parent_id, name')

            in_progress_states = ['assigned', 'in_progress', 'paused',
                                  'partially_completed', 'urgent', 'important', 'regular']

            def _descendants(root_id):
                ids = [root_id]
                frontier = [root_id]
                while frontier:
                    children = Kra.search([('parent_id', 'in', frontier)]).ids
                    if not children:
                        break
                    ids.extend(children)
                    frontier = children
                return ids

            def _bucket_counts(kra_ids):
                completed = Kpi.search([
                    ('kra_id', 'in', kra_ids), ('active', '=', True),
                    ('task_state', '=', 'completed'),
                ], order='completion_date desc')
                awaiting = Kpi.search([
                    ('kra_id', 'in', kra_ids), ('active', '=', True),
                    ('task_state', '=', 'awaiting_client'),
                ])
                in_progress = Kpi.search([
                    ('kra_id', 'in', kra_ids), ('active', '=', True),
                    ('task_state', 'in', in_progress_states),
                ])
                total_hours = sum((completed.mapped('timer_total_seconds')) or [0]) / 3600.0
                quoted_hours = sum((completed.mapped('client_quoted')) or [0])
                return {
                    'completed_count': len(completed),
                    'awaiting_client_count': len(awaiting),
                    'in_progress_count': len(in_progress),
                    'total_actual_hours': round(total_hours, 2),
                    'total_quoted_hours': round(quoted_hours, 2),
                }

            result = []
            for client in client_kras:
                kra_ids = _descendants(client.id)
                counts = _bucket_counts(kra_ids)

                # Per-project breakdown: direct children of the client KRA (the "projects").
                projects = []
                direct_children = Kra.search(
                    [('parent_id', '=', client.id), ('active', '=', True)],
                    order='sequence, name',
                )
                for proj in direct_children:
                    proj_ids = _descendants(proj.id)
                    proj_counts = _bucket_counts(proj_ids)
                    projects.append({
                        'project_id': proj.id,
                        'project_name': proj.name or '',
                        **proj_counts,
                    })

                # Most recent invoice for this client
                last_inv = Inv.search([('client_kra_id', '=', client.id)],
                                       order='invoice_date desc, id desc', limit=1)
                # Open invoice (draft) for this client
                open_inv = Inv.search([('client_kra_id', '=', client.id),
                                       ('state', '=', 'draft')], limit=1)

                result.append({
                    'client_id': client.id,
                    'client_name': client.name or '',
                    'parent_name': client.parent_id.name or '',
                    'currency_id': client.currency_id.id if client.currency_id else False,
                    'currency_name': client.currency_id.name if client.currency_id else '',
                    **counts,
                    'projects': projects,
                    'last_invoice_name': last_inv.name if last_inv else '',
                    'last_invoice_date': str(last_inv.invoice_date) if last_inv and last_inv.invoice_date else '',
                    'last_invoice_state': last_inv.state if last_inv else '',
                    'open_invoice_id': open_inv.id if open_inv else False,
                })
            return {'status': True, 'clients': result}
        except Exception as e:
            import logging
            logging.getLogger(__name__).error(f"owner_certified_by_client: {str(e)}")
            return {'status': False, 'message': str(e), 'clients': []}

    @http.route('/kpi_owner/client_certified_tasks', type='json', auth='user', methods=['POST'], csrf=False)
    def owner_client_certified_tasks(self, **params):
        """List the actual signed-off tasks for one client (drill-down from the dashboard)."""
        try:
            client_id = int(params['client_kra_id'])
            Kra = request.env['kra.master'].sudo()
            client = Kra.browse(client_id)
            if not client.exists():
                return {'status': False, 'message': 'Client not found'}
            kra_ids = [client.id]
            frontier = [client.id]
            while frontier:
                children = Kra.search([('parent_id', 'in', frontier)]).ids
                if not children:
                    break
                kra_ids.extend(children)
                frontier = children
            kpis = request.env['kra.kpi'].sudo().search([
                ('kra_id', 'in', kra_ids), ('active', '=', True),
                ('task_state', '=', 'completed'),
            ], order='completion_date desc')
            rows = [{
                'id': k.id,
                'name': k.name,
                'external_ref': k.external_ref or '',
                'priority': k.priority or '',
                'actual_hours': round((k.timer_total_seconds or 0) / 3600.0, 2),
                'quoted_hours': round(k.client_quoted or 0.0, 2),
                'primary_assignee': k.user_id.name if k.user_id else '',
                'client_signed_by': k.client_signature_text or '',
                'completion_date': k.completion_date.strftime('%Y-%m-%d') if k.completion_date else '',
                'signed_date': str(k.signed_certificate_date) if k.signed_certificate_date else '',
            } for k in kpis]
            return {'status': True, 'client_name': client.name, 'tasks': rows}
        except Exception as e:
            import logging
            logging.getLogger(__name__).error(f"owner_client_certified_tasks: {str(e)}")
            return {'status': False, 'message': str(e), 'tasks': []}

    # ============================================================
    #         CLIENT PORTAL — SAVED SIGNATURE IMAGE
    # ============================================================

    @http.route('/kpi_client_portal/upload_signature', type='json', auth='user', methods=['POST'], csrf=False)
    def upload_user_signature(self, **params):
        """Upload the current user's saved signature image (used for sign-off auto-fill)."""
        try:
            user = request.env.user
            updates = {}
            if 'file_data' in params:
                updates['kpi_signature_image'] = params['file_data'] or False
                updates['kpi_signature_image_name'] = params.get('file_name') or ''
            if updates:
                user.sudo().write(updates)
            return {
                'status': True,
                'has_signature': bool(user.kpi_signature_image),
                'file_name': user.kpi_signature_image_name or '',
            }
        except Exception as e:
            import logging
            logging.getLogger(__name__).error(f"upload_user_signature: {str(e)}")
            return {'status': False, 'message': str(e)}

    @http.route('/kpi_client_portal/get_signature_info', type='json', auth='user', methods=['POST'], csrf=False)
    def get_user_signature_info(self, **params):
        """Return info + base64 of the current user's saved signature image."""
        try:
            user = request.env.user
            raw = user.kpi_signature_image
            b64 = raw.decode() if isinstance(raw, bytes) else (raw or '')
            return {
                'status': True,
                'has_signature': bool(raw),
                'file_name': user.kpi_signature_image_name or '',
                'image_b64': b64,
                'user_name': user.name or user.login or '',
            }
        except Exception as e:
            import logging
            logging.getLogger(__name__).error(f"get_user_signature_info: {str(e)}")
            return {'status': False, 'message': str(e), 'has_signature': False, 'image_b64': ''}

    # ============================================================
    #            DEVELOPER WORKDAY SESSIONS
    # ============================================================

    @http.route('/kpi_workday/status', type='json', auth='user', methods=['POST'], csrf=False)
    def workday_status(self, **params):
        """Read-only: does the current user have an OPEN session for today?
        Used by OWL onMounted to decide whether to show the green
        'Start Workday' button or the red 'End Workday' one.
        Returns: {status, is_open: bool, session_id, login_at, today_closed: bool}
        """
        try:
            user = request.env.user
            Session = request.env['kpi.work.session'].sudo()
            today = fields.Date.context_today(request.env.user)
            # Prefer the OPEN session so is_open stays correct even when a closed
            # (e.g. auto-closed, then restarted) session from earlier today exists.
            open_row = Session.search([
                ('user_id', '=', user.id),
                ('session_date', '=', today),
                ('state', '=', 'open'),
            ], limit=1)
            row = open_row or Session.search([
                ('user_id', '=', user.id),
                ('session_date', '=', today),
            ], limit=1)
            # day_done = developer ENDED today themselves → block restart (one
            # start + one end per day). today_closed kept for older clients.
            day_done = Session._day_done(user)
            return {
                'status':       True,
                'is_open':      bool(open_row),
                'today_closed': bool(row and row.state == 'closed' and not open_row),
                'day_done':     day_done,
                'session_id':   row.id if row else False,
                'login_at':     fields.Datetime.to_string(row.login_at) if row and row.login_at else '',
            }
        except Exception as e:
            import logging
            logging.getLogger(__name__).error(f"workday_status: {e}")
            return {'status': False, 'message': str(e), 'is_open': False}

    @http.route('/kpi_workday/ping', type='json', auth='user', methods=['POST'], csrf=False)
    def workday_ping(self, **params):
        """Open today's workday session for the current user if absent.
        Idempotent. Called from the green 'Start Workday' button when the
        developer explicitly clicks to start their day.
        Returns: {status, session_id, login_at, opened: bool}
        """
        try:
            user = request.env.user
            Session = request.env['kpi.work.session'].sudo()
            # One start + one end per day: don't reopen a workday already ended
            # today by the developer. (Auto-closed days can still restart.)
            if Session._day_done(user):
                return {'status': False, 'day_done': True,
                        'message': 'Workday already ended for today'}
            existed = Session.search([
                ('user_id', '=', user.id),
                ('session_date', '=', fields.Date.context_today(request.env.user)),
                ('state', '=', 'open'),
            ], limit=1)
            sess = Session._get_or_open_today(user=user)
            if sess.state == 'closed':
                # Defensive: the choke-point guard handed back a closed session
                # (day done) — never report it as an opened workday.
                return {'status': False, 'day_done': True,
                        'message': 'Workday already ended for today'}
            return {
                'status':     True,
                'session_id': sess.id,
                'login_at':   fields.Datetime.to_string(sess.login_at),
                'opened':     not bool(existed),
            }
        except Exception as e:
            import logging
            logging.getLogger(__name__).error(f"workday_ping: {e}")
            return {'status': False, 'message': str(e)}

    @http.route('/kpi_workday/today_summary', type='json', auth='user', methods=['POST'], csrf=False)
    def workday_today_summary(self, **params):
        """Return today's payload (login time, totals, per-task array, Workday Map)
        for the current user. Used to render the End Workday modal.
        Auto-opens the session if missing so the modal works first time too.

        full_day=True to match action_end_day, which is what actually gets sent and
        saved. Without it this preview took login_at from the CURRENT session only
        while still pulling the whole day's breaks — so after a restart the Map
        showed earlier breaks BEFORE "Started workday", and the preview disagreed
        with the record. Same call, same day, same story.
        """
        try:
            user = request.env.user
            Session = request.env['kpi.work.session'].sudo()
            sess = Session._get_or_open_today(user=user)
            # Force re-compute so live presence reflects "now".
            sess.invalidate_recordset(['productive_seconds', 'presence_seconds',
                                       'task_count'])
            payload = sess._daily_summary_payload(full_day=True)
            payload['status'] = True
            return payload
        except Exception as e:
            import logging
            logging.getLogger(__name__).error(f"workday_today_summary: {e}")
            return {'status': False, 'message': str(e)}

    @http.route('/kpi_workday/idle_reason', type='json', auth='user', methods=['POST'], csrf=False)
    def workday_idle_reason(self, **params):
        """The developer says why nothing is running: meeting | break | no_tasks.

        Reached two ways — the 10-min prompt, or the board's status button (so a
        09:31 meeting can be declared at 09:31 rather than waiting to be asked).

        ALL THREE open a non-task block, so the time lands on the Workday Map
        instead of silently becoming "unexplained idle". They differ only in who
        gets told:
          meeting/break → admins hear NOTHING. A meeting is not a problem, and
                          telling them would be the false alarm that makes the
                          whole feed worth ignoring.
          no_tasks      → admins told straight away; that one IS theirs to fix.

        `no_tasks` records time TOO (it used to record nothing): a developer with
        nothing assigned would otherwise look identical to one who is slacking —
        untrue, and unfair to them.

        `note` is optional free text, clamped to NOTE_MAX here because a client
        maxLength is a courtesy, not a guarantee, and this text has to fit in a
        Workday Map chip.
        """
        try:
            reason = (params.get('reason') or '').strip().lower()
            if reason not in NONTASK_REASON_CODES:
                return {'status': False, 'message': 'Invalid reason.'}
            note = (params.get('note') or '').strip()[:NOTE_MAX]
            user = request.env.user
            sess = request.env['kpi.work.session'].sudo()._get_or_open_today(user=user)

            # All the switching/idempotency rules live on the model so they can be
            # probed from an Odoo shell — a controller-side copy could only be tested
            # by re-implementing it, which tests the copy and not the code.
            res = sess._set_nontask_reason(reason, note=note)

            return {'status': True, 'reason': reason, 'note': note,
                    'admins_notified': res['admins_notified'],
                    # What this tap ENDED, so the client can say "Meeting saved".
                    'switched_from': res['switched_from'],
                    'started_at': (sess.nontask_started_at.strftime('%Y-%m-%d %H:%M:%S')
                                   if sess.nontask_started_at else '')}
        except Exception as e:
            import logging
            logging.getLogger(__name__).error(f"workday_idle_reason: {e}")
            return {'status': False, 'message': str(e)}

    @http.route('/kpi_workday/end_nontask', type='json', auth='user', methods=['POST'], csrf=False)
    def workday_end_nontask(self, **params):
        """Close the open Meeting/Break block ([End] on the board pill).

        Starting a task closes it automatically (see kra.kpi.start_task); this is
        for the case a task ISN'T started next — e.g. the meeting ends and they
        still have nothing to work on. `ask_again` tells the app to re-open the
        same popup, so 'I have no tasks' is one tap rather than a 10-min wait.
        """
        try:
            user = request.env.user
            sess = request.env['kpi.work.session'].sudo()._get_or_open_today(user=user)
            closed = sess._close_nontask_block()
            return {'status': True, 'closed': closed,
                    'ask_again': not sess._has_running_task()}
        except Exception as e:
            import logging
            logging.getLogger(__name__).error(f"workday_end_nontask: {e}")
            return {'status': False, 'message': str(e)}

    @http.route('/kpi_workday/end', type='json', auth='user', methods=['POST'], csrf=False)
    def workday_end(self, **params):
        """Close today's session, fire the WhatsApp summary, return the payload.
        Idempotent: if already closed, returns the existing payload.
        """
        try:
            user = request.env.user
            Session = request.env['kpi.work.session'].sudo()
            sess = Session._get_or_open_today(user=user)
            note = (params.get('note') or '').strip()
            payload = sess.action_end_day(note=note)
            payload['status'] = True
            return payload
        except Exception as e:
            import logging
            logging.getLogger(__name__).error(f"workday_end: {e}")
            return {'status': False, 'message': str(e)}

    # ------------------------------------------------------------------ #
    # Workday Snapshots — the frozen End-Workday images, for admins       #
    # ------------------------------------------------------------------ #
    def _snapshot_admin_ok(self):
        """Admin/coordinator/owner/system only. The same triple the other admin
        screens use. Snapshots are a record of someone's day — a developer must
        not be able to read a colleague's."""
        user = request.env.user
        return (user.has_group('kra_kpi_module.group_kra_admin')
                or user.has_group('kra_kpi_module.group_kra_owner')
                or user.has_group('base.group_system'))

    @http.route('/kpi_workday_snapshot/list', type='json', auth='user', methods=['POST'], csrf=False)
    def workday_snapshot_list(self, **params):
        """Snapshots grouped date -> developers, for the admin screen's
        date ▸ developer ▸ image tree.

        Deliberately carries NO image bytes: the tree can hold weeks of days and a
        base64 PNG per row would be megabytes of JSON for pictures nobody has
        opened yet. The client pulls one image only when a row is expanded.

        params: limit_days (default 30) — how many recent days to return.
        """
        try:
            if not self._snapshot_admin_ok():
                return {'status': False, 'message': 'Not authorized.'}
            limit_days = max(1, min(365, int(params.get('limit_days') or 30)))
            Snap = request.env['kpi.workday.snapshot'].sudo()
            rows = Snap.search([], order='session_date desc, id desc')

            days, seen = [], {}
            for r in rows:
                key = str(r.session_date)
                if key not in seen:
                    if len(days) >= limit_days:
                        continue
                    seen[key] = {'date': key, 'label': r.session_date.strftime('%d %b %Y'),
                                 'devs': []}
                    days.append(seen[key])
                seen[key]['devs'].append({
                    'snapshot_id': r.id,
                    'user_id': r.user_id.id,
                    'name': r.user_id.name or r.user_id.login or '—',
                    'presence_display': _fmt_hms(r.presence_seconds or 0),
                    'productive_display': _fmt_hms(r.productive_seconds or 0),
                    'standard_hours': r.standard_hours or 0.0,
                    # Frozen on the row, NOT recomputed: the standard may change
                    # later, and this day was judged against the rule of its time.
                    'met_standard': bool(r.met_standard),
                    'task_count': r.task_count or 0,
                    'image_name': r.image_name or 'workday.png',
                    'generated_at': (fields.Datetime.to_string(r.generated_at)
                                     if r.generated_at else ''),
                })
            return {'status': True, 'days': days}
        except Exception as e:
            import logging
            logging.getLogger(__name__).error(f"workday_snapshot_list: {e}")
            return {'status': False, 'message': str(e)}

    @http.route('/kpi_workday_snapshot/image_b64', type='json', auth='user', methods=['POST'], csrf=False)
    def workday_snapshot_image_b64(self, **params):
        """One snapshot's PNG as base64 — the APP's path.

        The app CANNOT use /web/image or any type='http' URL: React Native's image
        loader sends no session cookie, so it silently lands on the login page and
        renders nothing (the same trap documented on /kpi/progress/file_b64). The
        web board uses /web/image directly and doesn't need this.
        """
        try:
            if not self._snapshot_admin_ok():
                return {'status': False, 'message': 'Not authorized.'}
            sid = int(params.get('snapshot_id') or 0)
            rec = request.env['kpi.workday.snapshot'].sudo().browse(sid)
            if not rec.exists() or not rec.image:
                return {'status': False, 'message': 'Snapshot not found.'}
            data = rec.image
            if isinstance(data, bytes):
                data = data.decode()
            return {'status': True, 'data_b64': data, 'mimetype': 'image/png',
                    'file_name': rec.image_name or 'workday.png',
                    'dev': rec.user_id.name or '', 'date': str(rec.session_date),
                    'presence_display': _fmt_hms(rec.presence_seconds or 0),
                    'met_standard': bool(rec.met_standard)}
        except Exception as e:
            import logging
            logging.getLogger(__name__).error(f"workday_snapshot_image_b64: {e}")
            return {'status': False, 'message': str(e)}

    # ------------------------------------------------------------------ #
    # My Snapshots — the CURRENT user's OWN End-Workday images (self only) #
    # ------------------------------------------------------------------ #
    @http.route('/kpi_workday_snapshot/my_list', type='json', auth='user', methods=['POST'], csrf=False)
    def my_workday_snapshot_list(self, **params):
        """The logged-in user's OWN workday snapshots — no admin gate.

        HARD-SCOPED to request.env.user: the domain is fixed to this user's id and
        NO user_id is ever read from params, so a developer can only ever see their
        own days. Same shape as /kpi_workday_snapshot/list (date ▸ [details] ▸ image)
        so the app screen is reused unchanged.
        """
        try:
            uid = request.env.user.id
            limit_days = max(1, min(365, int(params.get('limit_days') or 30)))
            Snap = request.env['kpi.workday.snapshot'].sudo()
            rows = Snap.search([('user_id', '=', uid)], order='session_date desc, id desc')
            name = request.env.user.name or request.env.user.login or 'You'
            days, seen = [], {}
            for r in rows:
                key = str(r.session_date)
                if key not in seen:
                    if len(days) >= limit_days:
                        continue
                    seen[key] = {'date': key, 'label': r.session_date.strftime('%d %b %Y'),
                                 'devs': []}
                    days.append(seen[key])
                seen[key]['devs'].append({
                    'snapshot_id': r.id,
                    'user_id': r.user_id.id,
                    'name': name,
                    'presence_display': _fmt_hms(r.presence_seconds or 0),
                    'productive_display': _fmt_hms(r.productive_seconds or 0),
                    'standard_hours': r.standard_hours or 0.0,
                    'met_standard': bool(r.met_standard),
                    'task_count': r.task_count or 0,
                    'image_name': r.image_name or 'workday.png',
                    'generated_at': (fields.Datetime.to_string(r.generated_at)
                                     if r.generated_at else ''),
                })
            return {'status': True, 'days': days}
        except Exception as e:
            import logging
            logging.getLogger(__name__).error(f"my_workday_snapshot_list: {e}")
            return {'status': False, 'message': str(e)}

    @http.route('/kpi_workday_snapshot/my_image_b64', type='json', auth='user', methods=['POST'], csrf=False)
    def my_workday_snapshot_image_b64(self, **params):
        """One of the CURRENT user's OWN snapshot images. Ownership is verified
        server-side (rec.user_id must equal the caller) so a developer can never
        pull a colleague's image by guessing an id."""
        try:
            sid = int(params.get('snapshot_id') or 0)
            rec = request.env['kpi.workday.snapshot'].sudo().browse(sid)
            if not rec.exists() or rec.user_id.id != request.env.user.id or not rec.image:
                return {'status': False, 'message': 'Snapshot not found.'}
            data = rec.image
            if isinstance(data, bytes):
                data = data.decode()
            return {'status': True, 'data_b64': data, 'mimetype': 'image/png',
                    'file_name': rec.image_name or 'workday.png',
                    'dev': rec.user_id.name or '', 'date': str(rec.session_date),
                    'presence_display': _fmt_hms(rec.presence_seconds or 0),
                    'met_standard': bool(rec.met_standard)}
        except Exception as e:
            import logging
            logging.getLogger(__name__).error(f"my_workday_snapshot_image_b64: {e}")
            return {'status': False, 'message': str(e)}

    @http.route('/kpi_reports/developer_summary', type='json', auth='user', methods=['POST'], csrf=False)
    def developer_summary_report(self, **params):
        """Daily / weekly / monthly per-developer report (coord+owner only).

        params:
          time_frame: 'daily' | 'weekly' | 'monthly'  (default 'daily')
          user_id:    optional — filter to one developer
          from_date / to_date: optional ISO dates (defaults derived from time_frame)
        Returns:
          {status, time_frame, periods:[{period_key, label, productive_seconds,
                                          presence_seconds, task_count, sessions:[...],
                                          tasks:[{ref,name,duration_seconds,duration_display}]}]}
        """
        try:
            user = request.env.user
            allowed = (
                user.has_group('kra_kpi_module.group_kra_admin')
                or user.has_group('kra_kpi_module.group_kra_owner')
                or user.has_group('base.group_system')
            )
            if not allowed:
                return {'status': False, 'message': 'Not authorized.'}

            import datetime as _dt
            time_frame = (params.get('time_frame') or 'daily').lower()
            if time_frame not in ('daily', 'weekly', 'monthly'):
                time_frame = 'daily'
            today = fields.Date.context_today(user)
            defaults = {
                'daily':   today - _dt.timedelta(days=30),
                'weekly':  today - _dt.timedelta(weeks=12),
                'monthly': today - _dt.timedelta(days=180),
            }
            from_date = params.get('from_date')
            to_date   = params.get('to_date')
            from_date = fields.Date.from_string(from_date) if from_date else defaults[time_frame]
            to_date   = fields.Date.from_string(to_date) if to_date else today
            uid_filter = int(params.get('user_id')) if params.get('user_id') else None

            Session = request.env['kpi.work.session'].sudo()
            Log     = request.env['kpi.time.log'].sudo()
            dom_sess = [
                ('session_date', '>=', from_date),
                ('session_date', '<=', to_date),
            ]
            if uid_filter:
                dom_sess.append(('user_id', '=', uid_filter))
            sessions = Session.search(dom_sess, order='session_date asc, login_at asc')

            # Bucket sessions by (user, period_key).
            def _period_key(d):
                if time_frame == 'daily':
                    return (d.isoformat(),    d.strftime('%a %d %b %Y'))
                if time_frame == 'weekly':
                    iso = d.isocalendar()
                    return (f"{iso[0]}-W{iso[1]:02d}", f"Week {iso[1]} of {iso[0]}")
                # monthly
                return (d.strftime('%Y-%m'), d.strftime('%B %Y'))

            from collections import defaultdict
            bucket = defaultdict(lambda: {
                'productive_seconds': 0.0, 'presence_seconds': 0.0,
                'task_count': 0, 'sessions': [], 'tasks': {}, 'user_name': '',
            })
            for s in sessions:
                key, label = _period_key(s.session_date)
                slot_id = (s.user_id.id, key)
                slot = bucket[slot_id]
                slot['productive_seconds'] += float(s.productive_seconds or 0.0)
                slot['presence_seconds']   += float(s.presence_seconds or 0.0)
                slot['user_name']           = s.user_id.name or s.user_id.login or '—'
                slot['label']               = label
                slot['period_key']          = key
                slot['user_id']             = s.user_id.id
                slot['sessions'].append({
                    'session_id':         s.id,
                    'session_date':       str(s.session_date),
                    'login_at':           fields.Datetime.to_string(s.login_at),
                    'logout_at':          fields.Datetime.to_string(s.logout_at) if s.logout_at else '',
                    'productive_seconds': float(s.productive_seconds or 0.0),
                    'presence_seconds':   float(s.presence_seconds or 0.0),
                    'task_count':         int(s.task_count or 0),
                    'auto_closed':        bool(s.auto_closed),
                })

            # Per-period per-task drill-down (uses kpi.time.log directly so we
            # don't depend on session row existence — e.g. for legacy logs).
            dom_log = [
                ('work_date', '>=', from_date),
                ('work_date', '<=', to_date),
                ('is_active', '=', False),
            ]
            if uid_filter:
                dom_log.append(('user_id', '=', uid_filter))
            for log in Log.search(dom_log):
                key, _ = _period_key(log.work_date)
                slot_id = (log.user_id.id, key)
                slot = bucket[slot_id]
                if not slot.get('label'):
                    # Log without a session row (legacy) — synthesise.
                    slot['user_name']  = log.user_id.name or log.user_id.login or '—'
                    slot['label']      = _period_key(log.work_date)[1]
                    slot['period_key'] = key
                    slot['user_id']    = log.user_id.id
                k = log.kpi_id
                if not k:
                    continue
                tkey = k.id
                if tkey not in slot['tasks']:
                    slot['tasks'][tkey] = {
                        'kpi_id':           k.id,
                        'ref':              k.external_ref or f"#{k.id}",
                        'name':             k.name or '',
                        'duration_seconds': 0.0,
                    }
                slot['tasks'][tkey]['duration_seconds'] += float(log.duration_seconds or 0.0)

            # Running (is_active=True) logs carry a STORED duration of 0 — end_time
            # is unset — so their time is invisible to both the session
            # productive_seconds and the finished-log drill-down above. Add it live
            # (now - start_time) so the rolling total ticks WITH the running task
            # instead of freezing until the developer pauses, and the task shows
            # its real elapsed time rather than 0s. Active logs are at most ~one per
            # logged-in developer, so the unbounded search is cheap.
            now = fields.Datetime.now()
            dom_active = [('is_active', '=', True)]
            if uid_filter:
                dom_active.append(('user_id', '=', uid_filter))
            for log in Log.search(dom_active):
                wd = log.work_date
                if not log.start_time or not wd or wd < from_date or wd > to_date:
                    continue
                live_secs = max(0.0, (now - log.start_time).total_seconds())
                key, _ = _period_key(wd)
                slot_id = (log.user_id.id, key)
                slot = bucket[slot_id]
                if not slot.get('label'):
                    slot['user_name']  = log.user_id.name or log.user_id.login or '—'
                    slot['label']      = _period_key(wd)[1]
                    slot['period_key'] = key
                    slot['user_id']    = log.user_id.id
                slot['productive_seconds'] += live_secs
                k = log.kpi_id
                if not k:
                    continue
                tkey = k.id
                if tkey not in slot['tasks']:
                    slot['tasks'][tkey] = {
                        'kpi_id':           k.id,
                        'ref':              k.external_ref or f"#{k.id}",
                        'name':             k.name or '',
                        'duration_seconds': 0.0,
                        'running':          False,
                    }
                slot['tasks'][tkey]['duration_seconds'] += live_secs
                slot['tasks'][tkey]['running'] = True

            # Flatten + sort.
            def _fmt(secs):
                s = int(secs or 0)
                h, rem = divmod(s, 3600)
                m = rem // 60
                if h and m:
                    return f"{h}h {m}m"
                if h:
                    return f"{h}h"
                return f"{m}m"

            periods = []
            for (uid, key), slot in bucket.items():
                tasks = []
                for t in sorted(slot['tasks'].values(),
                                key=lambda x: -x['duration_seconds']):
                    tasks.append({**t,
                                  'duration_display': _fmt(t['duration_seconds'])})
                periods.append({
                    'period_key':          slot.get('period_key', key),
                    'label':               slot.get('label', key),
                    'user_id':             uid,
                    'user_name':           slot.get('user_name', ''),
                    'productive_seconds':  slot['productive_seconds'],
                    'presence_seconds':    slot['presence_seconds'],
                    'productive_display':  _fmt(slot['productive_seconds']),
                    'presence_display':    _fmt(slot['presence_seconds']),
                    'task_count':          len(tasks),
                    'sessions':            slot['sessions'],
                    'tasks':               tasks,
                })
            # Sort by user then period_key desc (most recent first).
            periods.sort(key=lambda p: (p['user_name'], p['period_key']), reverse=False)
            return {
                'status':     True,
                'time_frame': time_frame,
                'from_date':  str(from_date),
                'to_date':    str(to_date),
                'periods':    periods,
            }
        except Exception as e:
            import logging, traceback
            logging.getLogger(__name__).error(f"developer_summary_report: {e}\n{traceback.format_exc()}")
            return {'status': False, 'message': str(e)}

    # ============================================================
    #     EMPLOYEE TRACKER — live board + granular action log
    # ============================================================

    @http.route('/kpi_owner/live_board', type='json', auth='user', methods=['POST'], csrf=False)
    def kpi_owner_live_board(self, **params):
        """Live 'who is working on what right now' board for the Admin Dashboard.

        Iterates today's OPEN work sessions (people who started their workday),
        joins the active time-log (is_active=True) to get each developer's current
        task + live elapsed, and marks the row ACTIVE vs present-but-idle by
        last_heartbeat freshness (company.away_after_minutes — the same signal the
        auto-away cron uses). Owner / Coordinator / System only, matching the
        Employee Tracker report gate.
        """
        user = request.env.user
        allowed = (
            user.has_group('kra_kpi_module.group_kra_admin')
            or user.has_group('kra_kpi_module.group_kra_owner')
            or user.has_group('base.group_system')
        )
        if not allowed:
            return {'status': False, 'message': 'Not authorized.'}
        try:
            Session = request.env['kpi.work.session'].sudo()
            Log = request.env['kpi.time.log'].sudo()
            now = fields.Datetime.now()
            today = fields.Date.context_today(user)
            try:
                away_min = max(1, int(request.env.company.away_after_minutes or 5))
            except Exception:
                away_min = 5
            stale_cutoff = now - timedelta(minutes=away_min)

            open_sessions = Session.search([
                ('session_date', '=', today),
                ('state', '=', 'open'),
            ], order='login_at asc')

            rows = []
            for s in open_sessions:
                u = s.user_id
                active_log = Log.search([
                    ('user_id', '=', u.id),
                    ('is_active', '=', True),
                ], limit=1)
                task = active_log.kpi_id if active_log else False
                hb = s.last_heartbeat
                is_active_now = bool(hb and hb >= stale_cutoff)  # heartbeat fresh = at the desk
                elapsed = 0
                if active_log and active_log.start_time:
                    elapsed = max(0, int((now - active_log.start_time).total_seconds()))
                rows.append({
                    'user_id':         u.id,
                    'user_name':       u.name or u.login or '—',
                    'login_raw':       (s.login_at.isoformat() + 'Z') if s.login_at else False,
                    'active':          is_active_now,
                    'on_task':         bool(task),
                    'task_id':         task.id if task else False,
                    'task_ref':        (task.external_ref or f"#{task.id}") if task else '',
                    'task_name':       task.name if task else '',
                    'task_state':      task.task_state if task else '',
                    # ISO+Z so the client ticks it with the same helper it uses for a
                    # running task (see live_status) and stays in sync with the server clock.
                    'task_start_raw':  (active_log.start_time.isoformat() + 'Z') if (active_log and active_log.start_time) else False,
                    'elapsed_seconds': elapsed,
                    'elapsed_display': _fmt_hms(elapsed),
                })
            return {
                'status':             True,
                'as_of':              fields.Datetime.to_string(now),
                'away_after_minutes': away_min,
                'present_count':      len(rows),
                'active_count':       sum(1 for r in rows if r['active']),
                'on_task_count':      sum(1 for r in rows if r['on_task']),
                'rows':               rows,
            }
        except Exception as e:
            import logging, traceback
            logging.getLogger(__name__).error(f"kpi_owner_live_board: {e}\n{traceback.format_exc()}")
            return {'status': False, 'message': str(e)}

    @http.route('/kpi_owner/live_tracking', type='json', auth='user', methods=['POST'], csrf=False)
    def kpi_owner_live_tracking(self, **params):
        """Admin LIVE TRACKING: every developer's current activity + attendance.

        For the full developer roster (owners/coordinators/system excluded — they
        imply the developer group), returns per person what they are doing RIGHT
        NOW — running task + live elapsed / active break / lunch / meeting / idle /
        offline — and today's attendance (present / late-with-arrival / absent /
        pending), judged against a configurable local cutoff
        (company.attendance_cutoff_hour, default 10:40). Owner / Coordinator /
        System only.
        """
        import pytz
        from datetime import datetime as _dt, time as _time, timedelta as _td
        user = request.env.user
        allowed = (
            user.has_group('kra_kpi_module.group_kra_admin')
            or user.has_group('kra_kpi_module.group_kra_owner')
            or user.has_group('base.group_system'))
        if not allowed:
            return {'status': False, 'message': 'Not authorized.'}
        try:
            Session = request.env['kpi.work.session'].sudo()
            Log = request.env['kpi.time.log'].sudo()
            Break = request.env['kpi.break.log'].sudo()
            now = fields.Datetime.now()
            today = fields.Date.context_today(user)
            try:
                away_min = max(1, int(request.env.company.away_after_minutes or 5))
            except Exception:
                away_min = 5
            stale_cutoff = now - _td(minutes=away_min)

            # Local attendance cutoff (default 10:40 IST) -> naive-UTC for a direct
            # compare against the naive-UTC login_at.
            tz = pytz.timezone(user.tz or 'Asia/Kolkata')
            try:
                cut = float(request.env.company.attendance_cutoff_hour or 10.667)
            except Exception:
                cut = 10.667
            ch = int(cut)
            cm = int(round((cut - ch) * 60))
            if cm >= 60:
                ch, cm = ch + 1, 0
            ch = min(max(ch, 0), 23)
            cm = min(max(cm, 0), 59)
            cutoff_local = tz.localize(_dt.combine(today, _time(ch, cm)))
            cutoff_utc = cutoff_local.astimezone(pytz.UTC).replace(tzinfo=None)
            past_cutoff = now > cutoff_utc

            break_sel = dict(Break._fields['break_type'].selection)
            nontask_sel = dict(Session._fields['nontask_reason'].selection)

            # Roster = pure developers only. owner->admin->developer implies down, so
            # exclude the admin tiers + system to keep admins out of the roster.
            devs = request.env.ref('kra_kpi_module.group_kra_developer').sudo().user_ids.filtered(
                lambda u: u.active and not u.share and u.kpi_app_login_enabled
                and not u.has_group('kra_kpi_module.group_kra_admin')
                and not u.has_group('base.group_system'))

            rows = []
            for u in devs.sorted(lambda x: (x.name or x.login or '').lower()):
                sessions = Session.search(
                    [('user_id', '=', u.id), ('session_date', '=', today)], order='login_at asc')
                logins = [s.login_at for s in sessions if s.login_at]
                first_login = min(logins) if logins else False
                open_sess = sessions.filtered(lambda s: s.state == 'open')[:1]

                # ---- attendance ----
                if first_login:
                    attendance = 'late' if first_login > cutoff_utc else 'present'
                    arrival = pytz.UTC.localize(first_login).astimezone(tz).strftime('%H:%M')
                else:
                    attendance = 'absent' if past_cutoff else 'pending'
                    arrival = ''

                # ---- current activity (only meaningful with an open session) ----
                activity_kind, activity_label, elapsed, active_flag = 'offline', '', 0, False
                if open_sess:
                    s = open_sess
                    active_flag = bool(s.last_heartbeat and s.last_heartbeat >= stale_cutoff)
                    active_log = Log.search([('user_id', '=', u.id), ('is_active', '=', True)], limit=1)
                    brk = Break.search([('user_id', '=', u.id), ('is_active', '=', True)], limit=1)
                    if active_log and active_log.kpi_id:
                        task = active_log.kpi_id
                        activity_kind = 'task'
                        activity_label = task.name or (task.external_ref or ("#%s" % task.id))
                        if active_log.start_time:
                            elapsed = max(0, int((now - active_log.start_time).total_seconds()))
                    elif brk:
                        bt = brk.break_type
                        activity_kind = bt if bt in ('break', 'lunch', 'meeting') else 'break'
                        activity_label = break_sel.get(bt, bt)
                        if brk.start_time:
                            elapsed = max(0, int((now - brk.start_time).total_seconds()))
                    elif s.nontask_started_at:
                        r = s.nontask_reason or 'no_tasks'
                        activity_kind = r if r in ('break', 'lunch', 'meeting', 'no_tasks') else 'no_tasks'
                        activity_label = nontask_sel.get(r, r)
                        elapsed = max(0, int((now - s.nontask_started_at).total_seconds()))
                    else:
                        activity_kind, activity_label = 'idle', 'Idle'

                rows.append({
                    'user_id': u.id,
                    'name': u.name or u.login or '—',
                    'attendance': attendance,          # present | late | absent | pending
                    'arrival': arrival,                # local "12:00"
                    'started': bool(open_sess),        # workday currently open
                    'active': active_flag,             # heartbeat fresh (at desk)
                    'activity_kind': activity_kind,    # task|break|lunch|meeting|no_tasks|idle|offline
                    'activity_label': activity_label,  # e.g. "abc"
                    'elapsed_seconds': elapsed,
                    'elapsed_display': _fmt_hms(elapsed) if elapsed else '',
                })

            return {
                'status': True,
                'as_of': fields.Datetime.to_string(now),
                'cutoff_display': cutoff_local.strftime('%H:%M'),
                'total': len(rows),
                'present_count': sum(1 for r in rows if r['attendance'] in ('present', 'late')),
                'active_count': sum(1 for r in rows if r['active']),
                'absent_count': sum(1 for r in rows if r['attendance'] == 'absent'),
                'late_count': sum(1 for r in rows if r['attendance'] == 'late'),
                'rows': rows,
            }
        except Exception as e:
            import logging, traceback
            logging.getLogger(__name__).error(f"kpi_owner_live_tracking: {e}\n{traceback.format_exc()}")
            return {'status': False, 'message': str(e)}

    @http.route('/kpi_owner/action_log', type='json', auth='user', methods=['POST'], csrf=False)
    def kpi_owner_action_log(self, **params):
        """Granular lifecycle log for the Employee Tracker's 'expand arrow'.

        Reads kpi.action.log — the per-task audit trail (start / pause / resume /
        complete / approve / pre-approval / notifications) written everywhere by
        _log_action but, until now, read by NO route. Filter by task (kpi_id),
        actor (user_id), and/or a day / date-range on the event time. Owner /
        Coordinator / System only (matches the model ACL).
        """
        user = request.env.user
        allowed = (
            user.has_group('kra_kpi_module.group_kra_admin')
            or user.has_group('kra_kpi_module.group_kra_owner')
            or user.has_group('base.group_system')
        )
        if not allowed:
            return {'status': False, 'message': 'Not authorized.'}
        try:
            Log = request.env['kpi.action.log'].sudo()
            domain = []
            if params.get('kpi_id'):
                domain.append(('kpi_id', '=', int(params['kpi_id'])))
            if params.get('user_id'):
                domain.append(('actor_user_id', '=', int(params['user_id'])))
            # Event-time window. 'date' = a single day; else optional from/to.
            day = params.get('date')
            if day:
                domain += [('create_date', '>=', day + ' 00:00:00'),
                           ('create_date', '<=', day + ' 23:59:59')]
            else:
                if params.get('from_date'):
                    domain.append(('create_date', '>=', params['from_date'] + ' 00:00:00'))
                if params.get('to_date'):
                    domain.append(('create_date', '<=', params['to_date'] + ' 23:59:59'))
            try:
                limit = min(max(1, int(params.get('limit') or 500)), 2000)
            except (TypeError, ValueError):
                limit = 500
            src_labels = dict(Log._fields['source'].selection)
            entries = Log.search(domain, order='create_date desc, id desc', limit=limit)
            rows = []
            for e in entries:
                payload = {}
                if e.payload_json:
                    try:
                        payload = json.loads(e.payload_json)
                    except (ValueError, TypeError):
                        payload = {'raw': e.payload_json}
                rows.append({
                    'id':            e.id,
                    'kpi_id':        e.kpi_id.id if e.kpi_id else False,
                    'task_ref':      (e.kpi_id.external_ref or f"#{e.kpi_id.id}") if e.kpi_id else '',
                    'task_name':     e.kpi_id.name if e.kpi_id else '',
                    'event':         e.event or '',
                    'source':        e.source or '',
                    'source_label':  src_labels.get(e.source, e.source or ''),
                    'actor':         (e.actor_user_id.name if e.actor_user_id else (e.actor_label or '')),
                    'actor_user_id': e.actor_user_id.id if e.actor_user_id else False,
                    # Raw UTC + ISO+Z; the client localises with the same helper it uses elsewhere.
                    'when':          fields.Datetime.to_string(e.create_date),
                    'when_raw':      (e.create_date.isoformat() + 'Z') if e.create_date else False,
                    'success':       bool(e.success),
                    'device_ip':     e.device_ip or '',
                    'payload':       payload,
                })
            return {'status': True, 'count': len(rows), 'entries': rows}
        except Exception as e:
            import logging, traceback
            logging.getLogger(__name__).error(f"kpi_owner_action_log: {e}\n{traceback.format_exc()}")
            return {'status': False, 'message': str(e)}

    # ============================================================
    #     WORK REPORT — completed tasks over a period (task/client)
    # ============================================================

    @http.route('/kpi_reports/period_work/generate', type='json', auth='user', methods=['POST'], csrf=False)
    def period_work_report(self, **params):
        """Completed-tasks report over a period, viewable task-wise or client-wise.

        params:
          from_date / to_date : ISO dates (default = current month → today)
          group_by            : 'task' (flat list) | 'client' (grouped)
          client_kra_id       : optional — restrict to one client (to send them their report)
        Returns each COMPLETED task (task_state='completed', completion_date in range) with
        its client/project/assignee, completion date, and hours from kpi.time.log in range.
        """
        user = request.env.user
        allowed = (
            user.has_group('kra_kpi_module.group_kra_manager')
            or user.has_group('kra_kpi_module.group_kra_admin')
            or user.has_group('kra_kpi_module.group_kra_owner')
            or user.has_group('base.group_system')
        )
        if not allowed:
            return {'status': False, 'message': 'Not authorized.'}
        try:
            today = fields.Date.context_today(user)
            from_date = fields.Date.from_string(params.get('from_date')) if params.get('from_date') else today.replace(day=1)
            to_date = fields.Date.from_string(params.get('to_date')) if params.get('to_date') else today
            group_by = (params.get('group_by') or 'task').lower()
            if group_by not in ('task', 'client'):
                group_by = 'task'

            # The picked dates are the user's LOCAL calendar days. completion_date and
            # time-log start_time are stored naive-UTC, so convert the local
            # midnight..midnight window to UTC bounds — otherwise tasks completed near
            # local midnight land in the wrong day/period.
            import pytz
            from datetime import datetime as _dtcls, time as _tcls
            tz = pytz.timezone(user.tz or 'Asia/Kolkata')
            start_s = fields.Datetime.to_string(
                tz.localize(_dtcls.combine(from_date, _tcls.min)).astimezone(pytz.UTC).replace(tzinfo=None))
            end_s = fields.Datetime.to_string(
                tz.localize(_dtcls.combine(to_date, _tcls.max)).astimezone(pytz.UTC).replace(tzinfo=None))

            Kpi = request.env['kra.kpi'].sudo()
            Log = request.env['kpi.time.log'].sudo()

            domain = [
                ('task_state', '=', 'completed'),
                ('completion_date', '>=', start_s),
                ('completion_date', '<=', end_s),
            ]
            if params.get('client_kra_id'):
                domain.append(('client_kra_id', '=', int(params['client_kra_id'])))
            tasks = Kpi.search(domain, order='completion_date desc')

            def _fmt_h(secs):
                s = int(secs or 0)
                h, rem = divmod(s, 3600)
                m = rem // 60
                return (f"{h}h {m}m" if h else f"{m}m") if (h or m) else "0h"

            rows = []
            for t in tasks:
                logs = Log.search([
                    ('kpi_id', '=', t.id),
                    ('is_active', '=', False),
                    ('start_time', '>=', start_s),
                    ('start_time', '<=', end_s),
                ])
                secs = sum(logs.mapped('duration_seconds'))
                rows.append({
                    'task_id':         t.id,
                    'ref':             t.external_ref or f"#{t.id}",
                    'name':            t.name or '',
                    'client_id':       t.client_kra_id.id if t.client_kra_id else False,
                    'client_name':     t.client_kra_id.name if t.client_kra_id else '—',
                    'project_name':    t.kra_id.name if t.kra_id else '',
                    'assignee':        t.user_id.name if t.user_id else '',
                    'completion_date': str(t.completion_date.date()) if t.completion_date else '',
                    'priority':        t.priority or '',
                    'hours_seconds':   secs,
                    'hours_display':   _fmt_h(secs),
                })

            # Brand header/watermark: the Company Branding logo (skip Odoo's default
            # placeholder), else the bundled brand logo — same source as the invoice PDF.
            company = request.env.company
            logo_b64 = ''
            if company.logo and not company.uses_default_logo:
                raw = company.logo
                logo_b64 = raw.decode() if isinstance(raw, bytes) else raw
            if not logo_b64:
                try:
                    import base64
                    from odoo.tools import file_open
                    with file_open('kra_kpi_module/static/src/img/alphalize_logo.png', 'rb') as f:
                        logo_b64 = base64.b64encode(f.read()).decode()
                except Exception:
                    logo_b64 = ''

            total_secs = sum(r['hours_seconds'] for r in rows)
            base = {
                'status':             True,
                'group_by':           group_by,
                'from_date':          str(from_date),
                'to_date':            str(to_date),
                'total_tasks':        len(rows),
                'total_hours_display': _fmt_h(total_secs),
                'company_name':       request.env.company.name or '',
                'company_logo_b64':   logo_b64,
            }

            if group_by == 'client':
                from collections import defaultdict
                buckets = defaultdict(lambda: {'client_name': '', 'client_id': False, 'tasks': [], 'total_seconds': 0.0})
                for r in rows:
                    b = buckets[r['client_id']]
                    b['client_name'] = r['client_name']
                    b['client_id'] = r['client_id']
                    b['tasks'].append(r)
                    b['total_seconds'] += r['hours_seconds']
                clients = []
                for b in buckets.values():
                    clients.append({
                        'client_id':     b['client_id'],
                        'client_name':   b['client_name'],
                        'task_count':    len(b['tasks']),
                        'total_seconds': b['total_seconds'],
                        'total_display': _fmt_h(b['total_seconds']),
                        'tasks':         b['tasks'],
                    })
                clients.sort(key=lambda c: (c['client_name'] or '').lower())
                base['clients'] = clients
                return base

            base['tasks'] = rows
            return base
        except Exception as e:
            import logging, traceback
            logging.getLogger(__name__).error(f"period_work_report: {e}\n{traceback.format_exc()}")
            return {'status': False, 'message': str(e)}

    @http.route('/kpi_reports/period_work/clients', type='json', auth='user', methods=['POST'], csrf=False)
    def period_work_clients(self, **params):
        """Client dropdown for the Work Report (is_client KRAs)."""
        user = request.env.user
        allowed = (
            user.has_group('kra_kpi_module.group_kra_manager')
            or user.has_group('kra_kpi_module.group_kra_admin')
            or user.has_group('kra_kpi_module.group_kra_owner')
            or user.has_group('base.group_system')
        )
        if not allowed:
            return {'status': False, 'message': 'Not authorized.'}
        try:
            clients = request.env['kra.master'].sudo().search(
                [('is_client', '=', True), ('active', '=', True)], order='name')
            return {'status': True, 'clients': [
                {'id': c.id, 'name': (c.parent_id.name + ' > ' + c.name) if c.parent_id else c.name}
                for c in clients
            ]}
        except Exception as e:
            return {'status': False, 'message': str(e)}

    @http.route('/kpi_reports/period_work/pdf', type='json', auth='user', methods=['POST'], csrf=False)
    def period_work_pdf(self, **params):
        """Server-side branded PDF of the Work Report (for the mobile app, which
        can't run the web's jsPDF). Reuses period_work_report for the data + access
        gate, then renders it with reportlab in the same style as the daily report."""
        data = self.period_work_report(**params)
        if not data.get('status'):
            return data   # not authorized / error — pass straight through
        try:
            pdf_bytes = self._build_work_report_pdf(data)
        except Exception as e:
            import logging, traceback
            logging.getLogger(__name__).error(f"period_work_pdf: {e}\n{traceback.format_exc()}")
            return {'status': False, 'message': str(e)}
        import base64
        return {
            'status': True,
            'file_name': 'Work_Report_%s_to_%s.pdf' % (data.get('from_date'), data.get('to_date')),
            'mimetype': 'application/pdf',
            'data_b64': base64.b64encode(pdf_bytes).decode(),
        }

    def _build_work_report_pdf(self, data):
        """Render the Work Report payload to PDF bytes with reportlab — Times fonts,
        #38385C table headers, zebra rows, a faint centred logo watermark, mirroring
        the daily task report / invoice PDFs. Branding logo comes in the payload."""
        import base64
        import io
        from xml.sax.saxutils import escape as _esc
        from reportlab.lib import colors
        from reportlab.lib.pagesizes import A4
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib.units import mm
        from reportlab.lib.utils import ImageReader
        from reportlab.platypus import (
            SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, Image, HRFlowable,
        )

        brand = colors.HexColor('#38385C')
        zebra = colors.HexColor('#F7F7FB')
        grid = colors.HexColor('#E1E1E1')
        rule = colors.HexColor('#BEBEBE')
        ink = colors.black
        muted = colors.HexColor('#64748B')

        logo_bytes = None
        if data.get('company_logo_b64'):
            try:
                logo_bytes = base64.b64decode(data['company_logo_b64'])
            except Exception:
                logo_bytes = None
        logo_reader = None
        logo_ratio = 1.0
        if logo_bytes:
            try:
                logo_reader = ImageReader(io.BytesIO(logo_bytes))
                iw, ih = logo_reader.getSize()
                if iw and ih:
                    logo_ratio = float(iw) / float(ih)
            except Exception:
                logo_reader = None

        styles = getSampleStyleSheet()
        s_company = ParagraphStyle('co', parent=styles['Normal'], fontName='Times-Bold', fontSize=22, alignment=2, textColor=ink, leading=24)
        s_doctype = ParagraphStyle('dt', parent=styles['Normal'], fontName='Times-Roman', fontSize=12, alignment=2, textColor=muted, leading=16)
        s_meta = ParagraphStyle('me', parent=styles['Normal'], fontName='Times-Roman', fontSize=10, textColor=muted, leading=14)
        s_sect = ParagraphStyle('se', parent=styles['Normal'], fontName='Times-Bold', fontSize=13, textColor=brand, leading=16, spaceBefore=8, spaceAfter=4)
        cell = ParagraphStyle('c', parent=styles['Normal'], fontName='Times-Roman', fontSize=9, textColor=ink, leading=11)

        PW, PH = A4
        MARGIN = 48
        content_w = PW - 2 * MARGIN

        def _watermark(canvas, doc_):
            if not logo_reader:
                return
            canvas.saveState()
            wm_w = PW * 0.55
            wm_h = wm_w / (logo_ratio or 1.0)
            if wm_h > PH * 0.5:
                wm_h = PH * 0.5
                wm_w = wm_h * (logo_ratio or 1.0)
            x = (PW - wm_w) / 2.0
            y = (PH - wm_h) / 2.0
            try:
                canvas.setFillAlpha(0.08)
                canvas.setStrokeAlpha(0.08)
            except Exception:
                pass
            try:
                canvas.drawImage(logo_reader, x, y, wm_w, wm_h, mask='auto', preserveAspectRatio=True)
            except Exception:
                pass
            canvas.restoreState()

        buf = io.BytesIO()
        doc = SimpleDocTemplate(
            buf, pagesize=A4, leftMargin=MARGIN, rightMargin=MARGIN,
            topMargin=MARGIN, bottomMargin=MARGIN, title='Work Report',
        )
        story = []

        right_cell = [Paragraph(_esc(data.get('company_name') or ''), s_company),
                      Paragraph('Work Report', s_doctype)]
        if logo_reader:
            h = 56
            w = min(h * logo_ratio, 150)
            left_cell = Image(io.BytesIO(logo_bytes), width=w, height=h)
            left_cell.hAlign = 'LEFT'
        else:
            left_cell = Paragraph('', cell)
        header = Table([[left_cell, right_cell]], colWidths=[content_w * 0.45, content_w * 0.55])
        header.setStyle(TableStyle([
            ('VALIGN', (0, 0), (-1, -1), 'TOP'),
            ('LEFTPADDING', (0, 0), (-1, -1), 0), ('RIGHTPADDING', (0, 0), (-1, -1), 0),
            ('TOPPADDING', (0, 0), (-1, -1), 0), ('BOTTOMPADDING', (0, 0), (-1, -1), 0),
        ]))
        story.append(header)
        story.append(Spacer(1, 6))
        story.append(HRFlowable(width='100%', thickness=0.8, color=rule, spaceBefore=0, spaceAfter=8))
        story.append(Paragraph('Period: %s to %s' % (data.get('from_date'), data.get('to_date')), s_meta))
        story.append(Paragraph('Completed tasks: %s • Total hours: %s' % (
            data.get('total_tasks', 0), data.get('total_hours_display', '0h')), s_meta))
        story.append(Spacer(1, 10))

        def _hcell(txt):
            return txt

        def _tbl(head, rows, colw):
            t = Table([head] + rows, colWidths=colw, repeatRows=1)
            t.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), brand),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
                ('FONTNAME', (0, 0), (-1, -1), 'Times-Roman'),
                ('FONTNAME', (0, 0), (-1, 0), 'Times-Bold'),
                ('FONTSIZE', (0, 0), (-1, -1), 9),
                ('VALIGN', (0, 0), (-1, -1), 'TOP'),
                ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, zebra]),
                ('GRID', (0, 0), (-1, -1), 0.5, grid),
                ('TOPPADDING', (0, 0), (-1, -1), 4), ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
                ('LEFTPADDING', (0, 0), (-1, -1), 5), ('RIGHTPADDING', (0, 0), (-1, -1), 5),
            ]))
            return t

        if data.get('group_by') == 'client':
            clients = data.get('clients') or []
            if not clients:
                story.append(Paragraph('No completed tasks in this period.', s_meta))
            for c in clients:
                story.append(Paragraph('%s — %s task(s), %s' % (
                    _esc(c.get('client_name') or '—'), c.get('task_count', 0),
                    c.get('total_display', '0h')), s_sect))
                head = ['Ref', 'Task', 'Project', 'Assignee', 'Completed', 'Hours']
                rows = [[
                    Paragraph(_esc(t.get('ref')), cell), Paragraph(_esc(t.get('name')), cell),
                    Paragraph(_esc(t.get('project_name')), cell), Paragraph(_esc(t.get('assignee')), cell),
                    t.get('completion_date', ''), t.get('hours_display', ''),
                ] for t in c.get('tasks', [])]
                story.append(_tbl(head, rows, [24 * mm, 44 * mm, 34 * mm, 30 * mm, 22 * mm, 18 * mm]))
                story.append(Spacer(1, 6))
        else:
            rows_data = data.get('tasks') or []
            if not rows_data:
                story.append(Paragraph('No completed tasks in this period.', s_meta))
            else:
                head = ['Ref', 'Task', 'Client', 'Project', 'Assignee', 'Completed', 'Hours']
                rows = [[
                    Paragraph(_esc(t.get('ref')), cell), Paragraph(_esc(t.get('name')), cell),
                    Paragraph(_esc(t.get('client_name')), cell), Paragraph(_esc(t.get('project_name')), cell),
                    Paragraph(_esc(t.get('assignee')), cell),
                    t.get('completion_date', ''), t.get('hours_display', ''),
                ] for t in rows_data]
                story.append(_tbl(head, rows, [20 * mm, 38 * mm, 26 * mm, 26 * mm, 26 * mm, 20 * mm, 16 * mm]))

        doc.build(story, onFirstPage=_watermark, onLaterPages=_watermark)
        return buf.getvalue()

    @http.route('/kpi_reports/my_workday', type='json', auth='user', methods=['POST'], csrf=False)
    def my_workday_report(self, **params):
        """A developer's OWN workday report for the mobile app. HARD-SCOPED to the
        logged-in user (request.env.user) — it NEVER reads a user_id from the client,
        so a developer can only ever see their own days, never a colleague's. No admin
        gate. Returns per-day presence, productive hours, and the tasks they worked on.
        Params: from_date / to_date (default = last 30 days)."""
        user = request.env.user
        try:
            import datetime as _dt
            from collections import defaultdict
            today = fields.Date.context_today(user)
            from_date = fields.Date.from_string(params.get('from_date')) if params.get('from_date') else (today - _dt.timedelta(days=30))
            to_date = fields.Date.from_string(params.get('to_date')) if params.get('to_date') else today

            uid = user.id  # HARD-SCOPED — deliberately NOT from params
            Session = request.env['kpi.work.session'].sudo()
            Log = request.env['kpi.time.log'].sudo()

            days = defaultdict(lambda: {'presence': 0.0, 'productive': 0.0, 'tasks': {}, 'sessions': 0})
            for s in Session.search([('user_id', '=', uid),
                                     ('session_date', '>=', from_date),
                                     ('session_date', '<=', to_date)]):
                d = str(s.session_date)
                days[d]['presence'] += float(s.presence_seconds or 0.0)
                days[d]['productive'] += float(s.productive_seconds or 0.0)
                days[d]['sessions'] += 1

            for log in Log.search([('user_id', '=', uid), ('is_active', '=', False),
                                   ('work_date', '>=', from_date), ('work_date', '<=', to_date)]):
                k = log.kpi_id
                if not k:
                    continue
                slot = days[str(log.work_date)]['tasks'].setdefault(
                    k.id, {'ref': k.external_ref or f"#{k.id}", 'name': k.name or '', 'secs': 0.0})
                slot['secs'] += float(log.duration_seconds or 0.0)

            def _fmt(secs):
                s = int(secs or 0)
                h, rem = divmod(s, 3600)
                m = rem // 60
                return (f"{h}h {m}m" if h else f"{m}m") if (h or m) else "0h"

            result = []
            for d in sorted(days.keys(), reverse=True):
                info = days[d]
                tasks = sorted(info['tasks'].values(), key=lambda x: -x['secs'])
                result.append({
                    'date': d,
                    'presence_display': _fmt(info['presence']),
                    'productive_display': _fmt(info['productive']),
                    'session_count': info['sessions'],
                    'task_count': len(tasks),
                    'tasks': [{'ref': t['ref'], 'name': t['name'], 'hours_display': _fmt(t['secs'])} for t in tasks],
                })
            return {'status': True, 'user_name': user.name or user.login or '',
                    'from_date': str(from_date), 'to_date': str(to_date), 'days': result}
        except Exception as e:
            import logging, traceback
            logging.getLogger(__name__).error(f"my_workday_report: {e}\n{traceback.format_exc()}")
            return {'status': False, 'message': str(e)}

    # ============================================================
    #            COORDINATOR REVIEW — first-open timestamp
    # ============================================================

    @http.route('/kpi_review/mark_started', type='json', auth='user', methods=['POST'], csrf=False)
    def mark_review_started(self, **params):
        """Stamp coordinator_review_started_at the first time a coordinator
        opens a `partially_completed` task card. Idempotent — repeat calls on
        the same record are no-ops. Caller must be in group_kra_admin or
        group_kra_owner (Coordinator/Owner). Returns:
            {status, already_started: bool, started_at: str|None, by: str|None}
        """
        try:
            kpi_id = int(params.get('kpi_id') or 0)
            if not kpi_id:
                return {'status': False, 'message': 'kpi_id is required.'}
            user = request.env.user
            allowed = (
                user.has_group('kra_kpi_module.group_kra_admin')
                or user.has_group('kra_kpi_module.group_kra_owner')
                or user.has_group('base.group_system')
            )
            if not allowed:
                return {'status': False, 'message': 'Not authorized to mark review started.'}

            kpi = request.env['kra.kpi'].sudo().browse(kpi_id)
            if not kpi.exists():
                return {'status': False, 'message': 'Task not found.'}
            # Lock the row so two concurrent clicks can't double-stamp.
            request.env.cr.execute(
                'SELECT id FROM kra_kpi WHERE id = %s FOR UPDATE',
                (kpi.id,),
            )
            kpi.invalidate_recordset(['coordinator_review_started_at',
                                      'coordinator_review_started_by'])

            if kpi.coordinator_review_started_at:
                return {
                    'status': True,
                    'already_started': True,
                    'started_at': fields.Datetime.to_string(kpi.coordinator_review_started_at),
                    'by': kpi.coordinator_review_started_by.name or '',
                }
            # Only stamp while the task is in the right state.
            if kpi.task_state != 'partially_completed':
                return {
                    'status': True,
                    'already_started': False,
                    'started_at': None,
                    'by': None,
                    'message': f"Task state is '{kpi.task_state}' — review timestamp not applicable.",
                }
            now = fields.Datetime.now()
            kpi.write({
                'coordinator_review_started_at': now,
                'coordinator_review_started_by': user.id,
            })
            try:
                kpi._log_action(
                    'review_started',
                    source='web',
                    actor_user_id=user.id,
                    payload={'task_state': kpi.task_state},
                )
            except Exception as exc:
                import logging
                logging.getLogger(__name__).warning(f"review_started audit log failed: {exc}")
            return {
                'status': True,
                'already_started': False,
                'started_at': fields.Datetime.to_string(now),
                'by': user.name or '',
            }
        except Exception as e:
            import logging
            logging.getLogger(__name__).error(f"mark_review_started: {str(e)}")
            return {'status': False, 'message': str(e)}

    # ============================================================
    #            CLIENT PORTAL — DIGITAL APPROVAL
    # ============================================================

    @http.route('/kpi_client_approval/list', type='json', auth='user', methods=['POST'], csrf=False)
    def list_client_pending_approvals(self, **params):
        """List KPIs awaiting the current client user's approval.

        Filters: only tasks in 'partially_completed' state under client-KRAs the
        current user is linked to (via client_user_ids on the KRA).
        """
        try:
            user = request.env.user
            client_kras = request.env['kra.master'].sudo().search([
                ('client_user_ids', 'in', [user.id]),
            ])
            if not client_kras:
                return {'status': True, 'kpis': [], 'message': 'You are not linked to any client KRA yet.'}
            # Walk all descendants of those client KRAs
            kra_ids = set()
            for k in client_kras:
                kra_ids.update(k._get_descendant_ids())
            kpis = request.env['kra.kpi'].sudo().search([
                ('kra_id', 'in', list(kra_ids)),
                '|',
                ('task_state', '=', 'awaiting_client'),
                '&', ('task_state', '=', 'partially_completed'), ('admin_approved', '=', True),
            ], order='completion_date desc')
            result = []
            for k in kpis:
                root = k.kra_id
                while root and root.parent_id:
                    root = root.parent_id
                project = k.kra_id if k.kra_id != root else False
                result.append({
                    'id': k.id,
                    'name': k.name or '',
                    'external_ref': k.external_ref or '',
                    'client_name': root.name if root else '',
                    'project_name': project.name if project else '',
                    'priority': k.priority or '',
                    'delivery_version': k.delivery_version or '',
                    'actual_hours': round((k.timer_total_seconds or 0) / 3600.0, 2),
                    'primary_assignee': k.user_id.name if k.user_id else '',
                    'completion_date': k.completion_date.strftime('%Y-%m-%d %H:%M') if k.completion_date else '',
                })
            return {'status': True, 'kpis': result}
        except Exception as e:
            import logging
            logging.getLogger(__name__).error(f"list_client_pending_approvals: {str(e)}")
            return {'status': False, 'message': str(e), 'kpis': []}

    def _is_client_authorized_for_kpi(self, kpi):
        """True if the current user is linked (via client_user_ids) to ANY ancestor KRA of this KPI."""
        user = request.env.user
        cur = kpi.kra_id
        while cur:
            if user.id in cur.client_user_ids.ids:
                return True
            cur = cur.parent_id
        return False

    # ================================================================== #
    # STAGE-1 pre-approval, IN-APP.  The client approves the developer    #
    # the admin assigned, from the app — no WhatsApp link, no token.      #
    # The decision engine (record_client_pre_decision) is reused as-is;   #
    # these routes are just an authenticated JSON door into it, next to   #
    # the existing token/WhatsApp door (controllers/kpi_wa_approval.py).  #
    # ================================================================== #
    def _pre_approval_authorized(self, kpi):
        """Client may decide if they submitted it, or are a client on its KRA tree."""
        if kpi.client_submitted and kpi.submitted_by_uid.id == request.env.user.id:
            return True
        return self._is_client_authorized_for_kpi(kpi)

    @http.route('/kpi_pre_approval/list', type='json', auth='user', methods=['POST'], csrf=False)
    def list_pre_approvals(self, **params):
        """Tasks awaiting THIS client's approval (app 'Approval needed' section)."""
        try:
            kpis = request.env['kra.kpi'].sudo().search([
                ('active', '=', True),
                # A developer is mandatory — approving "Unassigned" is meaningless.
                # Also hides rows created before request_client_pre_approval()
                # started skipping dev-less tasks, with no data migration.
                ('user_id', '!=', False),
                '|',
                ('task_state', 'in', ('pre_approval_pending', 'pre_approval_partial')),
                # AUTO-RELEASED but still undecided. Without this arm the card
                # disappears the instant the window closes and the client has no
                # way left to object — which would make "you can still object"
                # a lie. The work is unfinished, so their answer still matters.
                '&', '&',
                ('pre_approval_auto_released', '=', True),
                ('pre_approval_decision', '=', False),
                ('task_state', 'not in', ('completed', 'awaiting_client')),
            ], order='pre_approval_deadline_at asc, id desc')
            out = []
            for k in kpis:
                if not self._pre_approval_authorized(k):
                    continue
                root = k.kra_id
                while root and root.parent_id:
                    root = root.parent_id
                project = k.kra_id if k.kra_id != root else False
                nm = k.name or ''
                ref = (k.external_ref or '').upper()
                if nm.startswith('[Update]') or ref.startswith('UPD'):
                    kind = 'update'
                elif nm.startswith('[Bug]') or ref.startswith('BUG'):
                    kind = 'bug'
                else:
                    kind = 'requirement'
                out.append({
                    'id': k.id,
                    'name': nm,
                    'kind': kind,
                    'external_ref': k.external_ref or '',
                    'description': k.description or '',
                    'client_name': root.name if root else '',
                    'project_name': project.name if project else '',
                    'priority': k.priority or '',
                    'task_state': k.task_state or '',
                    # Who the admin assigned — the thing being approved.
                    'developer': k.user_id.name if k.user_id else '',
                    # Drives the app's countdown. The RELEASE clock, not the
                    # reminder clock: the reminder deadline is cleared once the
                    # nudges run out, which would freeze the countdown at 0
                    # while the task was in fact still waiting.
                    'deadline_at': (k.pre_approval_release_at.strftime('%Y-%m-%d %H:%M:%S')
                                    if k.pre_approval_release_at
                                    else (k.pre_approval_deadline_at.strftime('%Y-%m-%d %H:%M:%S')
                                          if k.pre_approval_deadline_at else '')),
                    'server_now': fields.Datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
                    # Whether this task will ACTUALLY auto-release, decided here so
                    # the app never has to re-derive the rule (and never promises
                    # something that can't happen). Mirrors the cron domain in
                    # _sweep_pre_approval_release exactly. A held task has no
                    # release clock → False.
                    'auto_releases': bool(k.task_state == 'pre_approval_pending'
                                          and not k.pre_approval_held
                                          and k.pre_approval_release_at),
                    # Legacy key, kept so an older mobile build keeps rendering
                    # its countdown. Same value; the name is the lie we are
                    # retiring — releasing is not approving.
                    'auto_approves': bool(k.task_state == 'pre_approval_pending'
                                          and not k.pre_approval_held
                                          and k.pre_approval_release_at),
                    # Already released: the window closed, work may have started,
                    # and this client still hasn't answered. The app shows these
                    # as "started without your reply — you can still object".
                    'auto_released': bool(k.pre_approval_auto_released),
                    'released_at': (k.pre_approval_auto_released_at.strftime('%Y-%m-%d %H:%M:%S')
                                    if k.pre_approval_auto_released_at else ''),
                    'held': bool(k.pre_approval_held),
                })
            return {'status': True, 'kpis': out}
        except Exception as e:
            import logging
            logging.getLogger(__name__).error(f"list_pre_approvals: {str(e)}")
            return {'status': False, 'message': str(e), 'kpis': []}

    @http.route('/kpi_pre_approval/decide', type='json', auth='user', methods=['POST'], csrf=False)
    def decide_pre_approval(self, **params):
        """Client approves/rejects the assigned developer, in-app.

        Delegates to the existing engine so the app decision gets the same
        row-locking, idempotency, late-decision handling and audit trail as the
        WhatsApp path. `source` must be 'web' — the field only allows web|whatsapp.
        """
        try:
            action = (params.get('action') or '').strip().lower()
            if action not in ('approve', 'reject', 'hold', 'resume'):
                return {'status': False, 'message': 'Invalid action.'}
            kpi = request.env['kra.kpi'].sudo().browse(int(params.get('kpi_id') or 0))
            if not kpi.exists():
                return {'status': False, 'message': 'Task not found.'}
            if not self._pre_approval_authorized(kpi):
                return {'status': False, 'message': 'Not authorized for this task.'}
            feedback = (params.get('feedback') or '').strip()

            # Hold / Resume are NOT decisions — they only stop/start the
            # auto-release clock, so they bypass record_client_pre_decision
            # (whose 'hold' action would mark the token used and lock the client
            # out of ever approving). See _hold_pre_approval.
            if action in ('hold', 'resume'):
                if action == 'hold':
                    kpi._hold_pre_approval()
                else:
                    kpi._resume_pre_approval()
                auto = bool(kpi.task_state == 'pre_approval_pending'
                            and not kpi.pre_approval_held
                            and kpi.pre_approval_release_at)
                return {'status': True,
                        'state': kpi.task_state,
                        'held': bool(kpi.pre_approval_held),
                        'auto_releases': auto,
                        'auto_approves': auto,     # legacy key — see the list route
                        'deadline_at': (kpi.pre_approval_release_at.strftime('%Y-%m-%d %H:%M:%S')
                                        if kpi.pre_approval_release_at else ''),
                        'server_now': fields.Datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
                        'message': ('Held — the team will not start.' if action == 'hold'
                                    else 'Resumed.')}

            # A reject must say why — the admin needs it to act. Enforced here
            # and not only in the UI: never trust the client.
            if action == 'reject' and not feedback:
                return {'status': False, 'message': 'Please tell us what to change.'}

            user = request.env.user
            res = kpi.record_client_pre_decision(
                action,
                partner_name=user.name or '',
                feedback=feedback,
                source='web',
                partner_id=user.partner_id.id if user.partner_id else None,
            )
            # Reject → hand the task back to the queue so an admin can assign a
            # different developer (which re-fires a fresh client approval).
            # AFTER the engine call: it's what stamps pre_approval_feedback, which
            # the notification body reads. Only for client-submitted tasks — the
            # WhatsApp/legacy flow keeps its existing park-in-rework behaviour.
            if (action == 'reject' and kpi.client_submitted
                    and res.get('status') != 'already_decided' and not res.get('late')):
                kpi._return_to_queue_after_reject()
            return {'status': bool(res.get('status', True)), **res}
        except Exception as e:
            import logging
            logging.getLogger(__name__).error(f"decide_pre_approval: {str(e)}")
            return {'status': False, 'message': str(e)}

    @http.route('/kpi_client_approval/get', type='json', auth='user', methods=['POST'], csrf=False)
    def get_client_approval_detail(self, **params):
        """Return full detail for one KPI the client is reviewing."""
        try:
            kpi = request.env['kra.kpi'].sudo().browse(int(params['kpi_id']))
            if not kpi.exists():
                return {'status': False, 'message': 'Task not found.'}
            if not self._is_client_authorized_for_kpi(kpi):
                return {'status': False, 'message': 'You are not authorized to review this task.'}
            data = kpi.get_completion_certificate_data()
            data['kpi']['client_chk_delivered'] = kpi.client_chk_delivered
            data['kpi']['client_chk_works'] = kpi.client_chk_works
            data['kpi']['client_chk_authorized'] = kpi.client_chk_authorized
            data['kpi']['client_approval_notes'] = kpi.client_approval_notes or ''
            data['kpi']['client_signature_text'] = kpi.client_signature_text or ''
            return {'status': True, 'data': data}
        except Exception as e:
            import logging
            logging.getLogger(__name__).error(f"get_client_approval_detail: {str(e)}")
            return {'status': False, 'message': str(e)}

    @http.route('/kpi_client_approval/approve', type='json', auth='user', methods=['POST'], csrf=False)
    def submit_client_approval(self, **params):
        """Client submits digital approval. Auto-generates a signed cert record and approves the task."""
        try:
            kpi = request.env['kra.kpi'].sudo().browse(int(params['kpi_id']))
            if not kpi.exists():
                return {'status': False, 'message': 'Task not found.'}
            if not self._is_client_authorized_for_kpi(kpi):
                return {'status': False, 'message': 'You are not authorized to approve this task.'}
            user = request.env.user
            # Signature: prefer the user's saved signature image; allow optional typed name on top.
            signature = (params.get('signature_text') or '').strip()
            has_saved_image = bool(user.kpi_signature_image)
            if not signature and not has_saved_image:
                return {'status': False,
                        'message': 'Please upload a signature image (Completions menu) or type your name.'}
            if not signature:
                # Fall back to user's display name as the typed signature when only the image is used
                signature = user.name or user.login or 'Client'
            chk_delivered = bool(params.get('chk_delivered'))
            chk_works = bool(params.get('chk_works'))
            chk_authorized = bool(params.get('chk_authorized'))
            if not (chk_delivered and chk_works and chk_authorized):
                return {'status': False, 'message': 'Please tick all three confirmation checkboxes before signing.'}
            notes = params.get('notes') or ''

            now = fields.Datetime.now()
            today = fields.Date.today()

            # Minimal plaintext digital-approval receipt — stored as the signed_certificate
            # so the approve_task() gate is satisfied. Admin can still view/download this.
            import base64 as _b64
            receipt = (
                f"DIGITAL APPROVAL RECEIPT\n"
                f"========================\n"
                f"Task:      {kpi.name}\n"
                f"Ref:       {kpi.external_ref or ''}\n"
                f"Signed by: {signature}  (user login: {user.login}, id={user.id})\n"
                f"Signed at: {now}\n"
                f"Checklist: delivered={chk_delivered}, works={chk_works}, authorized={chk_authorized}\n"
                f"Notes:     {notes}\n"
            ).encode('utf-8')
            # Stage 2: client digital sign-off. Records signature, generates a receipt,
            # and moves the task to 'completed'.
            kpi.write({
                'client_signature_text': signature,
                'client_signed_datetime': now,
                'client_approval_notes': notes,
                'client_chk_delivered': chk_delivered,
                'client_chk_works': chk_works,
                'client_chk_authorized': chk_authorized,
                'signed_certificate': _b64.b64encode(receipt).decode(),
                'signed_certificate_name': f'digital_approval_{kpi.id}.txt',
                'signed_certificate_date': today,
                # Final state transition — client sign-off closes the task
                'task_state': 'completed',
                'approved_by': user.id,
                'approval_date': now,
                # If admin hadn't run QA-approve yet, mark it now (client signed = effective approval)
                'admin_approved': True,
                # The client's signature IS the billable event — without this the task
                # looks fully approved but can never be invoiced, because the invoice
                # domain filters on client_final_approved (kpi_client_invoice.py:228).
                # Matches the app path (kpi_workflow.py) and approve_task() (kpi.py).
                'client_final_approved': True,
            })
            # Audit + notify owner / coordinator / developer that the client signed off.
            try:
                kpi._log_action(
                    'final_approved',
                    source='web',
                    actor_user_id=user.id,
                    actor_label=signature,
                    payload={'kpi_id': kpi.id, 'notes': notes},
                )
                kpi._notify('final_approved', decided_by=signature, feedback=notes, source='web')
            except Exception as nx:
                try:
                    kpi._log_action(
                        'notification_failed',
                        payload={'event': 'final_approved', 'error': str(nx)},
                        success=False,
                    )
                except Exception:
                    pass
            return {'status': True, 'message': 'Approved successfully.', 'kpi_id': kpi.id}
        except Exception as e:
            import logging
            logging.getLogger(__name__).error(f"submit_client_approval: {str(e)}")
            return {'status': False, 'message': str(e)}

    @http.route('/kpi_client_approval/reject', type='json', auth='user', methods=['POST'], csrf=False)
    def submit_client_rejection(self, **params):
        """Client rejects with a reason — task returns to paused state for the dev."""
        try:
            kpi = request.env['kra.kpi'].sudo().browse(int(params['kpi_id']))
            if not kpi.exists():
                return {'status': False, 'message': 'Task not found.'}
            if not self._is_client_authorized_for_kpi(kpi):
                return {'status': False, 'message': 'You are not authorized.'}
            reason = (params.get('reason') or '').strip() or 'Client requested changes.'
            kpi.reject_task(reason)
            return {'status': True}
        except Exception as e:
            import logging
            logging.getLogger(__name__).error(f"submit_client_rejection: {str(e)}")
            return {'status': False, 'message': str(e)}

    @http.route('/kpi_completion_cert/set_version', type='json', auth='user', methods=['POST'], csrf=False)
    def set_kpi_delivery_version(self, **params):
        """Quick endpoint to set delivery_version on a KPI (e.g. before generating cert)."""
        try:
            kpi = request.env['kra.kpi'].browse(int(params['kpi_id']))
            if not kpi.exists():
                return {'status': False, 'message': 'KPI not found.'}
            kpi.delivery_version = params.get('delivery_version') or ''
            return {'status': True}
        except Exception as e:
            import logging
            logging.getLogger(__name__).error(f"set_kpi_delivery_version: {str(e)}")
            return {'status': False, 'message': str(e)}

    @http.route('/kpi_completion_cert/upload_doc', type='json', auth='user', methods=['POST'], csrf=False)
    def upload_kpi_doc(self, **params):
        """Upload (or clear) a doc on the KPI. doc_type = 'requirement' or 'updates'."""
        try:
            kpi = request.env['kra.kpi'].browse(int(params['kpi_id']))
            if not kpi.exists():
                return {'status': False, 'message': 'KPI not found.'}
            doc_type = params.get('doc_type') or 'requirement'
            updates = {}
            if doc_type == 'requirement':
                if 'file_data' in params:
                    updates['requirement_document'] = params['file_data'] or False
                    updates['requirement_document_name'] = params.get('file_name') or ''
                if 'requirement_version' in params:
                    updates['requirement_version'] = params['requirement_version'] or ''
            elif doc_type == 'updates':
                if 'file_data' in params:
                    updates['updates_document'] = params['file_data'] or False
                    updates['updates_document_name'] = params.get('file_name') or ''
            else:
                return {'status': False, 'message': f'Unknown doc_type: {doc_type}'}
            if updates:
                kpi.write(updates)
            return {
                'status': True,
                'requirement_document_name': kpi.requirement_document_name or '',
                'requirement_version': kpi.requirement_version or '',
                'has_requirement_document': bool(kpi.requirement_document),
                'updates_document_name': kpi.updates_document_name or '',
                'has_updates_document': bool(kpi.updates_document),
            }
        except Exception as e:
            import logging
            logging.getLogger(__name__).error(f"upload_kpi_doc: {str(e)}")
            return {'status': False, 'message': str(e)}

    @http.route('/kpi_completion_cert/upload_signed', type='json', auth='user', methods=['POST'], csrf=False)
    def upload_signed_cert(self, **params):
        """Upload the client-signed completion certificate PDF onto a KPI."""
        try:
            kpi = request.env['kra.kpi'].browse(int(params['kpi_id']))
            if not kpi.exists():
                return {'status': False, 'message': 'KPI not found.'}
            updates = {}
            if 'file_data' in params:
                updates['signed_certificate'] = params['file_data'] or False
                updates['signed_certificate_name'] = params.get('file_name') or ''
            if 'signed_date' in params:
                updates['signed_certificate_date'] = params['signed_date'] or False
            if updates:
                kpi.write(updates)
            return {
                'status': True,
                'has_signed_certificate': bool(kpi.signed_certificate),
                'signed_certificate_name': kpi.signed_certificate_name or '',
                'signed_certificate_date': str(kpi.signed_certificate_date) if kpi.signed_certificate_date else '',
            }
        except Exception as e:
            import logging
            logging.getLogger(__name__).error(f"upload_signed_cert: {str(e)}")
            return {'status': False, 'message': str(e)}

    @http.route('/kpi_completion_cert/set_roles', type='json', auth='user', methods=['POST'], csrf=False)
    def set_kpi_roles(self, **params):
        """Set the 4 role lists on a KPI. Each role param is a list of user ids (or []).
        Pass only the role lists you want to change; omitted roles are left untouched."""
        try:
            kpi = request.env['kra.kpi'].browse(int(params['kpi_id']))
            if not kpi.exists():
                return {'status': False, 'message': 'KPI not found.'}
            updates = {}
            for role_field in ('developer_ids', 'tester_ids', 'coordinator_ids', 'lead_ids'):
                if role_field in params:
                    ids = [int(u) for u in (params[role_field] or [])]
                    updates[role_field] = [(6, 0, ids)]
            if updates:
                kpi.write(updates)
            return {
                'status': True,
                'developer_ids': kpi.developer_ids.ids,
                'tester_ids': kpi.tester_ids.ids,
                'coordinator_ids': kpi.coordinator_ids.ids,
                'lead_ids': kpi.lead_ids.ids,
            }
        except Exception as e:
            import logging
            logging.getLogger(__name__).error(f"set_kpi_roles: {str(e)}")
            return {'status': False, 'message': str(e)}

    # ============================================================
    #     REQUIREMENTS / UPDATES / BUG-REPORT BULK UPLOAD APIs
    # ============================================================

    # Mapping from doc_type to the external_ref prefix.  Used by both
    # /kpi_requirements/peek_next_ref (UI preview) and the bulk-create endpoint
    # so the client / admin / manager all share the same global sequence.
    _DOC_TYPE_REF_PREFIX = {
        'requirement': 'REQ',
        'update':      'UPT',
        'bug':         'BUG',
    }

    def _next_ref_number_for_prefix(self, prefix):
        """Return the next integer N such that '<PREFIX>-<NNN>' doesn't yet
        exist in kra_kpi.external_ref.  Case-insensitive match on the prefix,
        any zero-padding allowed (REQ-1, REQ-001, REQ-0001 are all considered).
        """
        if not prefix:
            return 1
        refs = request.env['kra.kpi'].sudo().search([
            ('external_ref', '=ilike', '%s-%%' % prefix),
        ]).mapped('external_ref')
        pat = re.compile(r'^%s-(\d+)$' % re.escape(prefix), re.IGNORECASE)
        max_n = 0
        for r in refs:
            m = pat.match((r or '').strip())
            if m:
                try:
                    n = int(m.group(1))
                    if n > max_n:
                        max_n = n
                except ValueError:
                    continue
        return max_n + 1

    @http.route('/kpi_requirements/peek_next_ref', type='json', auth='user', methods=['POST'], csrf=False)
    def peek_next_ref(self, **params):
        """Return the next external_ref number that bulk-create would assign for
        the given doc_type.  The UI uses this so it can preview row refs as the
        user adds rows.  Final refs are still assigned at create-time so two
        users opening the page at once don't collide.
        """
        try:
            doc_type = params.get('doc_type') or 'requirement'
            prefix = self._DOC_TYPE_REF_PREFIX.get(doc_type, 'REQ')
            n = self._next_ref_number_for_prefix(prefix)
            return {'status': True, 'prefix': prefix, 'next_number': n}
        except Exception as e:
            import logging
            logging.getLogger(__name__).error('peek_next_ref: %s' % str(e))
            return {'status': False, 'message': str(e), 'prefix': 'REQ', 'next_number': 1}

    @http.route('/kpi_requirements/create_bulk_tasks', type='json', auth='user', methods=['POST'], csrf=False)
    def create_bulk_tasks_from_doc(self, **params):
        """Create multiple KPIs at once from a single uploaded document.

        params:
          sub_kra_id (required) — the project KRA under which tasks are created
          doc_type ('requirement' | 'update' | 'bug') — affects doc storage slot + name prefix
          file_data, file_name (optional) — shared source document attached to every created KPI
          requirement_version (optional) — applied as requirement_version on all created KPIs

          tasks (preferred) — list of dicts, each with per-task config:
            { name, estimate_hours, estimate_minutes, primary_user_id, priority }

          Legacy fallback if `tasks` is missing:
            task_names + flat estimate_hours/estimate_minutes/primary_user_id/priority
        """
        try:
            sub_kra_id = int(params['sub_kra_id'])
            sub_kra = request.env['kra.master'].sudo().browse(sub_kra_id)
            if not sub_kra.exists():
                return {'status': False, 'message': 'Sub-KRA / project not found.'}
            is_client = self._is_current_user_client_only()
            if is_client:
                # Verify the sub-KRA is one this client user is linked to (or a descendant)
                allowed_ids = set()
                for k in self._resolve_client_kras_for_user():
                    allowed_ids.update(k._get_descendant_ids())
                if sub_kra_id not in allowed_ids:
                    return {'status': False, 'message': 'You are not authorized for this project.'}
            doc_type = params.get('doc_type') or 'requirement'
            name_prefix_map = {'requirement': '', 'update': '[Update] ', 'bug': '[Bug] '}
            name_prefix = name_prefix_map.get(doc_type, '')
            file_data = params.get('file_data') or ''
            file_name = params.get('file_name') or ''
            requirement_version = params.get('requirement_version') or ''

            # Build a unified list of task dicts whether the caller passed `tasks` or legacy `task_names`.
            tasks_in = params.get('tasks') or []
            if not tasks_in:
                names = [n.strip() for n in (params.get('task_names') or []) if n and n.strip()]
                flat_hours = int(params.get('estimate_hours') or 0)
                flat_min = int(params.get('estimate_minutes') or 0)
                flat_user = int(params['primary_user_id']) if params.get('primary_user_id') else False
                flat_prio = params.get('priority') or 'regular'
                tasks_in = [{
                    'name': n, 'estimate_hours': flat_hours, 'estimate_minutes': flat_min,
                    'primary_user_id': flat_user, 'priority': flat_prio,
                } for n in names]

            if not tasks_in:
                return {'status': False, 'message': 'No tasks provided.'}

            # is_client was already computed above (used for authorization).
            # Clients use sudo() to bypass ACL on kra.kpi.create (mirroring
            # /kpi_client_portal/add_task) since their portal group lacks CRUD on kra.kpi.
            kpi_model = request.env['kra.kpi'].sudo() if is_client else request.env['kra.kpi']

            # Server-controlled external_ref assignment.  The UI shows previewed
            # refs (REQ-076 etc.) but the actual value at insert time always
            # comes from this counter so concurrent uploads don't collide.
            ref_prefix = self._DOC_TYPE_REF_PREFIX.get(doc_type, 'REQ')
            ref_counter = self._next_ref_number_for_prefix(ref_prefix)

            created_ids = []
            for t in tasks_in:
                raw_name = (t.get('name') or '').strip()
                if not raw_name:
                    continue
                vals = {
                    'name': name_prefix + raw_name,
                    'kra_id': sub_kra_id,
                    'estimate_hours': int(t.get('estimate_hours') or 0),
                    'estimate_minutes': int(t.get('estimate_minutes') or 0),
                    'priority': t.get('priority') or 'regular',
                    'task_state': 'assigned',
                    'active': True,
                    # Every submission (client OR coordinator OR owner) lands in
                    # the Client Task Queue first — coordinator assigns developer
                    # + estimate, which auto-fires the pre-approval pipeline.
                    'published': False,
                    'submitted_by_uid': request.env.user.id,
                    # Marks the client-only queue + auto-approve path.
                    'client_submitted': bool(is_client),
                }
                # Clients can't pre-assign developers — only admin/manager can.
                if not is_client:
                    pu = t.get('primary_user_id')
                    if pu:
                        vals['user_id'] = int(pu)
                # Always overwrite external_ref with the server-side auto value
                # so client / admin / manager share one global sequence and
                # nobody can fork the numbering by typing their own.
                vals['external_ref'] = '%s-%03d' % (ref_prefix, ref_counter)
                ref_counter += 1
                if t.get('related_req_ref'):
                    vals['related_req_ref'] = t['related_req_ref']
                # The client's reason / detail for this task. This was silently
                # dropped before, which is why the Client Task Queue's Description
                # box was always empty for uploaded tasks even though it renders it.
                # (Mirrors /kpi_completion_cert/create_task_from_doc.)
                if t.get('description'):
                    vals['description'] = t['description']
                if file_data:
                    if doc_type == 'update':
                        vals['updates_document'] = file_data
                        vals['updates_document_name'] = file_name
                    else:
                        vals['requirement_document'] = file_data
                        vals['requirement_document_name'] = file_name
                if requirement_version:
                    vals['requirement_version'] = requirement_version
                kpi = kpi_model.create(vals)
                created_ids.append(kpi.id)
                # Client submission → tell the admins it's sitting in the queue
                # needing a developer. Re-nudged while unread by
                # kra.kpi._cron_nudge_queued_tasks. Never raises.
                if is_client:
                    try:
                        kpi._notify('client_task_queued')
                        kpi._schedule_queue_nudge()
                    except Exception as exc:
                        import logging
                        logging.getLogger(__name__).warning(
                            "client_task_queued notify failed for kpi %s: %s", kpi.id, exc)

            return {'status': True, 'created_count': len(created_ids), 'kpi_ids': created_ids}
        except Exception as e:
            import logging
            logging.getLogger(__name__).error(f"create_bulk_tasks_from_doc: {str(e)}")
            return {'status': False, 'message': str(e)}

    @http.route('/kpi_project/completion_status', type='json', auth='user', methods=['POST'], csrf=False)
    def get_project_completion_status(self, **params):
        """Return a hierarchical breakdown of all KPIs under a project (sub-KRA),
        grouped into Requirements / Updates / Bugs, showing completion progress.

        params: project_kra_id (required)
        """
        try:
            project_kra_id = int(params.get('project_kra_id') or 0)
            if not project_kra_id:
                return {'status': False, 'message': 'project_kra_id is required.'}
            project = request.env['kra.master'].browse(project_kra_id)
            if not project.exists():
                return {'status': False, 'message': 'Project KRA not found.'}

            kra_ids = project._get_descendant_ids()
            kpis = request.env['kra.kpi'].search([('kra_id', 'in', kra_ids), ('active', '=', True)])

            def _classify(kpi):
                """Return 'requirement' | 'update' | 'bug' based on name prefix or external_ref prefix."""
                nm = (kpi.name or '')
                if nm.startswith('[Update]'):
                    return 'update'
                if nm.startswith('[Bug]'):
                    return 'bug'
                ref = (kpi.external_ref or '').upper()
                if ref.startswith('UPD'):
                    return 'update'
                if ref.startswith('BUG'):
                    return 'bug'
                return 'requirement'

            def _serialize(k):
                est_hours = (k.estimate_hours or 0) + (k.estimate_minutes or 0) / 60.0
                actual_hours = (k.timer_total_seconds or 0) / 3600.0
                return {
                    'id': k.id,
                    'name': k.name,
                    'external_ref': k.external_ref or '',
                    'related_req_ref': k.related_req_ref or '',
                    'task_state': k.task_state or '',
                    'priority': k.priority or '',
                    'primary_assignee': k.user_id.name if k.user_id else '',
                    'estimate_hours': round(est_hours, 2),
                    'actual_hours': round(actual_hours, 2),
                    'has_signed_cert': bool(k.signed_certificate),
                    'deadline': str(k.deadline) if k.deadline else '',
                    'completion_date': k.completion_date.strftime('%Y-%m-%d') if k.completion_date else '',
                }

            # Bucketize
            requirements = {}   # external_ref -> {req: dict, updates: [], bugs: []}
            orphan_updates = []
            orphan_bugs = []
            for k in kpis:
                kind = _classify(k)
                ser = _serialize(k)
                if kind == 'requirement':
                    key = ser['external_ref'] or f'KPI-{k.id}'
                    bucket = requirements.setdefault(key, {'req': ser, 'updates': [], 'bugs': []})
                    bucket['req'] = ser  # in case multiple matches, last wins
                elif kind == 'update':
                    parent = (ser['related_req_ref'] or '').strip()
                    if parent and parent in requirements:
                        requirements[parent]['updates'].append(ser)
                    elif parent:
                        # Parent listed but not yet in dict — record and post-process
                        bucket = requirements.setdefault(parent, {'req': None, 'updates': [], 'bugs': []})
                        bucket['updates'].append(ser)
                    else:
                        orphan_updates.append(ser)
                elif kind == 'bug':
                    parent = (ser['related_req_ref'] or '').strip()
                    if parent and parent in requirements:
                        requirements[parent]['bugs'].append(ser)
                    elif parent:
                        bucket = requirements.setdefault(parent, {'req': None, 'updates': [], 'bugs': []})
                        bucket['bugs'].append(ser)
                    else:
                        orphan_bugs.append(ser)

            # Totals
            def _is_done(s):
                return s.get('task_state') == 'completed'

            total_req = len(requirements)
            done_req = sum(1 for v in requirements.values() if v['req'] and _is_done(v['req']))
            all_updates = []
            all_bugs = []
            for v in requirements.values():
                all_updates.extend(v['updates'])
                all_bugs.extend(v['bugs'])
            all_updates.extend(orphan_updates)
            all_bugs.extend(orphan_bugs)
            total_upd = len(all_updates)
            done_upd = sum(1 for u in all_updates if _is_done(u))
            total_bug = len(all_bugs)
            done_bug = sum(1 for b in all_bugs if _is_done(b))

            project_complete = (
                total_req + total_upd + total_bug > 0
                and done_req == total_req
                and done_upd == total_upd
                and done_bug == total_bug
            )

            return {
                'status': True,
                'project': {'id': project.id, 'name': project.name or ''},
                'requirements': [
                    {'ref': k, 'req': v['req'], 'updates': v['updates'], 'bugs': v['bugs']}
                    for k, v in sorted(requirements.items())
                ],
                'orphan_updates': orphan_updates,
                'orphan_bugs': orphan_bugs,
                'totals': {
                    'requirements_total': total_req,
                    'requirements_done': done_req,
                    'updates_total': total_upd,
                    'updates_done': done_upd,
                    'bugs_total': total_bug,
                    'bugs_done': done_bug,
                    'project_complete': project_complete,
                },
            }
        except Exception as e:
            import logging
            logging.getLogger(__name__).error(f"get_project_completion_status: {str(e)}")
            return {'status': False, 'message': str(e)}

    @http.route('/kpi_requirements/parse_file', type='json', auth='user', methods=['POST'], csrf=False)
    def parse_file_for_tasks(self, **params):
        """Parse an uploaded XLSX or CSV file and return a list of task names.

        params:
          file_data (base64 string, required)
          file_name (str, required) — used to detect extension
          doc_type ('requirement' | 'update' | 'bug') — hints which column heuristics to prefer
        Returns: {status, tasks: [...], source: 'xlsx' | 'csv', columns: [...]}
        """
        import base64
        try:
            file_data = params.get('file_data')
            file_name = (params.get('file_name') or '').lower()
            doc_type = params.get('doc_type') or 'requirement'
            if not file_data or not file_name:
                return {'status': False, 'message': 'Both file_data and file_name are required.'}
            raw = base64.b64decode(file_data)

            # Preferred column-name patterns by doc type.
            # NOTE: we ALSO honor an explicit "(task name)" suffix on the column header — this is
            # our convention for marking which column the import should treat as the task title.
            preferred_by_type = {
                'requirement': ['description', 'requirement', 'feature', 'title', 'summary', 'task', 'name'],
                'update': ['description of change', 'change', 'description', 'amendment', 'task', 'title', 'name', 'update'],
                'bug': ['description', 'bug', 'issue', 'title', 'summary', 'name'],
            }
            preferred = preferred_by_type.get(doc_type, preferred_by_type['requirement'])

            def _pick_column(headers):
                """Return index of best-match column. First check for an explicit
                '(task name)' suffix in any header, then fall back to keyword preferences."""
                norm = [(h or '').strip().lower() for h in headers]
                for i, h in enumerate(norm):
                    if '(task name)' in h or '(taskname)' in h:
                        return i, headers[i]
                for pat in preferred:
                    for i, h in enumerate(norm):
                        if pat in h:
                            return i, headers[i]
                return 0, (headers[0] if headers else '')

            def _find_marker_column(headers, marker):
                """Find the first column whose header contains the marker (e.g. '(id)', '(linked req)')."""
                norm = [(h or '').strip().lower() for h in headers]
                for i, h in enumerate(norm):
                    if marker in h:
                        return i
                return None

            tasks = []
            columns_seen = []

            def _row_to_task(row, col_idx, id_idx, linked_idx, module_idx):
                """Extract (name, external_ref, related_req_ref) from a row.
                Returns None if the row is empty/placeholder."""
                if col_idx >= len(row):
                    return None
                val = row[col_idx]
                if val is None:
                    return None
                val_str = str(val).strip()
                if (not val_str
                        or val_str.lower().startswith('(describe')
                        or len(val_str) > 250 or len(val_str) < 3):
                    return None
                if doc_type == 'bug' and module_idx is not None and module_idx < len(row):
                    mod = row[module_idx]
                    if mod:
                        val_str = f"{str(mod).strip()}: {val_str}"
                ext_ref = ''
                if id_idx is not None and id_idx < len(row) and row[id_idx]:
                    ext_ref = str(row[id_idx]).strip()
                linked = ''
                if linked_idx is not None and linked_idx < len(row) and row[linked_idx]:
                    linked = str(row[linked_idx]).strip()
                return {'name': val_str, 'external_ref': ext_ref, 'related_req_ref': linked}

            if file_name.endswith('.csv') or file_name.endswith('.txt'):
                import csv
                from io import StringIO
                try:
                    text = raw.decode('utf-8-sig')
                except UnicodeDecodeError:
                    text = raw.decode('latin-1')
                reader = csv.reader(StringIO(text))
                rows = list(reader)
                if not rows:
                    return {'status': True, 'tasks': [], 'source': 'csv', 'columns': []}
                headers = rows[0]
                columns_seen = list(headers)
                col_idx, col_name = _pick_column(headers)
                id_idx = _find_marker_column(headers, '(id)')
                linked_idx = _find_marker_column(headers, '(linked req)')
                module_idx = None
                if doc_type == 'bug':
                    for i, h in enumerate(headers):
                        if 'module' in (h or '').lower() or 'screen' in (h or '').lower():
                            module_idx = i; break
                header_is_data = not any(pat in (h or '').lower() for h in headers for pat in preferred)
                data_rows = rows if header_is_data else rows[1:]
                for r in data_rows:
                    t = _row_to_task(r, col_idx, id_idx, linked_idx, module_idx)
                    if t:
                        tasks.append(t)
                return {'status': True, 'tasks': tasks[:200], 'source': 'csv',
                        'columns': columns_seen, 'matched_column': col_name}

            if file_name.endswith('.xlsx') or file_name.endswith('.xls'):
                try:
                    import openpyxl
                    from io import BytesIO
                except Exception:
                    return {'status': False, 'message': 'openpyxl not available on this Odoo install.'}
                wb = openpyxl.load_workbook(BytesIO(raw), read_only=True, data_only=True)
                ws = wb.active
                rows_iter = list(ws.iter_rows(values_only=True))
                if not rows_iter:
                    return {'status': True, 'tasks': [], 'source': 'xlsx', 'columns': []}
                header_idx = 0
                for i, r in enumerate(rows_iter):
                    non_empty = [c for c in r if c not in (None, '')]
                    if len(non_empty) >= 2:
                        if all(isinstance(c, str) and len(c) <= 60 for c in non_empty):
                            header_idx = i
                            break
                headers = [str(c) if c is not None else '' for c in rows_iter[header_idx]]
                columns_seen = headers
                col_idx, col_name = _pick_column(headers)
                id_idx = _find_marker_column(headers, '(id)')
                linked_idx = _find_marker_column(headers, '(linked req)')
                module_idx = None
                if doc_type == 'bug':
                    for i, h in enumerate(headers):
                        if 'module' in h.lower() or 'screen' in h.lower():
                            module_idx = i; break
                for r in rows_iter[header_idx + 1:]:
                    t = _row_to_task(r, col_idx, id_idx, linked_idx, module_idx)
                    if t:
                        tasks.append(t)
                return {'status': True, 'tasks': tasks[:200], 'source': 'xlsx',
                        'columns': columns_seen, 'matched_column': col_name}

            return {'status': False,
                    'message': f'Unsupported file type: {file_name}. Use .csv, .xlsx, or .xls.',
                    'tasks': []}
        except Exception as e:
            import logging
            logging.getLogger(__name__).error(f"parse_file_for_tasks: {str(e)}")
            return {'status': False, 'message': str(e), 'tasks': []}

    @http.route('/kpi_requirements/suggest_tasks', type='json', auth='user', methods=['POST'], csrf=False)
    def suggest_tasks_from_text(self, **params):
        """Use Anthropic Claude (or a fallback heuristic) to suggest task names from pasted text.

        params:
          text (required) — pasted document text
          doc_type ('requirement' | 'update' | 'bug') — affects the system prompt
        Returns: {status, tasks: [str, ...], source: 'ai' | 'heuristic', message: str}
        """
        try:
            text = (params.get('text') or '').strip()
            doc_type = params.get('doc_type') or 'requirement'
            if not text:
                return {'status': False, 'message': 'No text provided.', 'tasks': []}

            # Cap input length to keep API costs bounded
            if len(text) > 16000:
                text = text[:16000]

            ICP = request.env['ir.config_parameter'].sudo()
            api_key = ICP.get_param('kra_kpi.anthropic_api_key')
            # FREE alternative: Google Gemini Flash (free tier — 15 req/min, 1500/day).
            # Grab a key from https://aistudio.google.com/apikey and save it via:
            #   ir.config_parameter `kra_kpi.gemini_api_key` = <your-key>
            gemini_key = ICP.get_param('kra_kpi.gemini_api_key')
            # FREE alternative #2: Groq (free tier, very fast Llama 3 inference).
            # Get a key from https://console.groq.com/keys and save as:
            #   ir.config_parameter `kra_kpi.groq_api_key` = <your-key>
            groq_key = ICP.get_param('kra_kpi.groq_api_key')

            prompts = {
                'requirement': (
                    "You are helping a software project manager extract atomic, implementable task names "
                    "from a client's REQUIREMENT document. Each task should be 3-12 words. "
                    "Return ONLY a JSON array of strings, no prose."
                ),
                'update': (
                    "You are extracting AMENDMENT/UPDATE task names from a client's update document. "
                    "Focus on the new/changed items. Each task 3-12 words. "
                    "Return ONLY a JSON array of strings, no prose."
                ),
                'bug': (
                    "You are extracting BUG-FIX task names from a bug report. One task per bug. "
                    "Each task 3-12 words, start with 'Fix: '. "
                    "Return ONLY a JSON array of strings, no prose."
                ),
            }
            system_prompt = prompts.get(doc_type, prompts['requirement'])

            import json as _json
            import urllib.request
            import logging as _logging
            _log = _logging.getLogger(__name__)
            # Groq (and some Gemini endpoints) sit behind Cloudflare, which
            # 403s the default `Python-urllib/x.y` user-agent as a suspected
            # bot.  A plain Mozilla UA gets through cleanly.
            _UA = 'Mozilla/5.0 (Odoo KRA-KPI Module)'

            # ---------- FREE: Google Gemini Flash ----------
            if gemini_key:
                try:
                    body = _json.dumps({
                        'system_instruction': {'parts': [{'text': system_prompt}]},
                        'contents': [{'role': 'user', 'parts': [{'text': text}]}],
                        'generationConfig': {
                            'temperature': 0.2,
                            'maxOutputTokens': 1024,
                            'responseMimeType': 'application/json',
                        },
                    }).encode('utf-8')
                    url = ('https://generativelanguage.googleapis.com/v1beta/models/'
                           'gemini-2.0-flash:generateContent?key=' + gemini_key)
                    req = urllib.request.Request(
                        url, data=body, method='POST',
                        headers={'content-type': 'application/json', 'User-Agent': _UA},
                    )
                    with urllib.request.urlopen(req, timeout=30) as resp:
                        raw = resp.read().decode('utf-8')
                        rb = _json.loads(raw)
                        candidates = rb.get('candidates') or []
                        text_out = ''
                        for c in candidates:
                            parts = (c.get('content') or {}).get('parts') or []
                            for p in parts:
                                text_out += p.get('text', '')
                        tasks = _parse_task_array(text_out)
                        if tasks:
                            return {'status': True, 'source': 'ai-gemini', 'tasks': tasks}
                except Exception as ai_err:
                    _log.warning(f"suggest_tasks Gemini error: {ai_err}")

            # ---------- FREE: Groq (Llama 3.3 70B) ----------
            if groq_key:
                try:
                    body = _json.dumps({
                        'model': 'llama-3.3-70b-versatile',
                        'messages': [
                            {'role': 'system', 'content': system_prompt},
                            {'role': 'user',   'content': text},
                        ],
                        'temperature': 0.2,
                        'max_tokens': 1024,
                        'response_format': {'type': 'json_object'},
                    }).encode('utf-8')
                    req = urllib.request.Request(
                        'https://api.groq.com/openai/v1/chat/completions',
                        data=body, method='POST',
                        headers={
                            'Authorization': f'Bearer {groq_key}',
                            'content-type': 'application/json',
                            'User-Agent': _UA,
                        },
                    )
                    with urllib.request.urlopen(req, timeout=30) as resp:
                        raw = resp.read().decode('utf-8')
                        rb = _json.loads(raw)
                        choices = rb.get('choices') or []
                        text_out = ''
                        for c in choices:
                            msg = c.get('message') or {}
                            text_out += msg.get('content', '')
                        tasks = _parse_task_array(text_out)
                        if tasks:
                            return {'status': True, 'source': 'ai-groq', 'tasks': tasks}
                except Exception as ai_err:
                    _log.warning(f"suggest_tasks Groq error: {ai_err}")

            # ---------- PAID: Anthropic Claude (only if key set) ----------
            if api_key:
                req_body = _json.dumps({
                    'model': 'claude-haiku-4-5-20251001',
                    'max_tokens': 1024,
                    'system': system_prompt,
                    'messages': [{'role': 'user', 'content': text}],
                }).encode('utf-8')
                req = urllib.request.Request(
                    'https://api.anthropic.com/v1/messages',
                    data=req_body,
                    method='POST',
                    headers={
                        'x-api-key': api_key,
                        'anthropic-version': '2023-06-01',
                        'content-type': 'application/json',
                        'User-Agent': _UA,
                    },
                )
                try:
                    with urllib.request.urlopen(req, timeout=30) as resp:
                        raw = resp.read().decode('utf-8')
                        body = _json.loads(raw)
                        content_blocks = body.get('content') or []
                        text_out = ''
                        for b in content_blocks:
                            if b.get('type') == 'text':
                                text_out += b.get('text', '')
                        tasks = _parse_task_array(text_out)
                        if tasks:
                            return {'status': True, 'source': 'ai-claude', 'tasks': tasks}
                except Exception as ai_err:
                    _log.warning(f"suggest_tasks Claude error: {ai_err}")

            # Heuristic fallback: split lines, dedupe, drop very short or numbered-only lines.
            import re
            candidates = []
            for line in text.splitlines():
                s = line.strip()
                if not s:
                    continue
                # Strip leading numbers, bullets
                s = re.sub(r'^[\-\*•\d\.\)\s]+', '', s).strip()
                if len(s) < 5 or len(s) > 200:
                    continue
                candidates.append(s)
            # Dedupe while preserving order
            seen = set()
            tasks = []
            for c in candidates:
                if c.lower() in seen:
                    continue
                seen.add(c.lower())
                tasks.append(c)
                if len(tasks) >= 30:
                    break

            any_key = bool(api_key or gemini_key or groq_key)
            return {
                'status': True,
                'source': 'heuristic-fallback' if any_key else 'heuristic',
                'tasks': tasks,
                'message': ('AI providers failed — used heuristic fallback. See server log.'
                            if any_key else
                            'AI not configured (set ir.config_parameter kra_kpi.groq_api_key '
                            'or kra_kpi.gemini_api_key or kra_kpi.anthropic_api_key).'),
            }
        except Exception as e:
            import logging
            logging.getLogger(__name__).error(f"suggest_tasks_from_text: {str(e)}")
            return {'status': False, 'message': str(e), 'tasks': []}

    @http.route('/kpi_completion_cert/create_task_from_doc', type='json', auth='user', methods=['POST'], csrf=False)
    def create_task_from_doc(self, **params):
        """Create a new KPI from an uploaded requirement document.
        Required: client_kra_id, sub_kra_id, task_name. Optional: file_data, file_name,
        requirement_version, estimate_hours, estimate_minutes, primary_user_id."""
        try:
            sub_kra_id = int(params.get('sub_kra_id') or params.get('client_kra_id'))
            sub_kra = request.env['kra.master'].browse(sub_kra_id)
            if not sub_kra.exists():
                return {'status': False, 'message': 'Sub-KRA / project not found.'}
            vals = {
                'name': params.get('task_name') or 'New Task',
                'kra_id': sub_kra_id,
                'estimate_hours': int(params.get('estimate_hours') or 0),
                'estimate_minutes': int(params.get('estimate_minutes') or 0),
                'priority': 'regular',
                'task_state': 'assigned',
                'active': True,
            }
            if params.get('primary_user_id'):
                vals['user_id'] = int(params['primary_user_id'])
            if params.get('requirement_version'):
                vals['requirement_version'] = params['requirement_version']
            if params.get('file_data'):
                vals['requirement_document'] = params['file_data']
                vals['requirement_document_name'] = params.get('file_name') or 'requirement'
            if params.get('description'):
                vals['description'] = params['description']
            kpi = request.env['kra.kpi'].create(vals)
            return {'status': True, 'kpi_id': kpi.id, 'name': kpi.name}
        except Exception as e:
            import logging
            logging.getLogger(__name__).error(f"create_task_from_doc: {str(e)}")
            return {'status': False, 'message': str(e)}

    @http.route('/kpi_completion_cert/get_sub_kras', type='json', auth='user', methods=['POST'], csrf=False)
    def get_sub_kras_for_client(self, **params):
        """For a chosen client root KRA, return descendant KRAs to populate the project dropdown."""
        try:
            client = request.env['kra.master'].browse(int(params['client_kra_id']))
            if not client.exists():
                return {'status': False, 'kras': []}
            kra_ids = client._get_descendant_ids()
            kras = request.env['kra.master'].browse(kra_ids)
            result = []
            for k in kras:
                # Build a path like "Nexgenn > Oman Photo House"
                path = []
                cur = k
                while cur:
                    path.insert(0, cur.name)
                    cur = cur.parent_id
                result.append({'id': k.id, 'name': k.name, 'display': ' > '.join(path)})
            return {'status': True, 'kras': result}
        except Exception as e:
            import logging
            logging.getLogger(__name__).error(f"get_sub_kras_for_client: {str(e)}")
            return {'status': False, 'message': str(e), 'kras': []}

    @http.route('/kpi_completion_cert/get_users', type='json', auth='user', methods=['POST'], csrf=False)
    def get_internal_users_for_picker(self, **params):
        """Return active internal users for the role / assignee dropdowns."""
        try:
            users = request.env['res.users'].search(
                [('active', '=', True), ('share', '=', False)], order='name')
            return {
                'status': True,
                'users': [{'id': u.id, 'name': u.name or u.login or '', 'login': u.login or ''} for u in users],
            }
        except Exception as e:
            import logging
            logging.getLogger(__name__).error(f"get_internal_users_for_picker: {str(e)}")
            return {'status': False, 'users': []}

    @http.route('/kpi_client_workspace/generate_zip', type='http', auth='user')
    def generate_client_workspace_zip(self, client_kra_id=None, project_kra_id=None,
                                       version='v1.0', include_data='0', **kw):
        """Generate a ZIP containing 3 XLSX files (Requirements / Updates / Bugs)
        pre-filled with the chosen client's name in the title row.

        If include_data='1', export existing KPIs under the selected project (or client) as rows.
        """
        import io, re, zipfile, base64
        try:
            try:
                import openpyxl
                from openpyxl.styles import Font, PatternFill, Alignment
                from openpyxl.utils import get_column_letter
            except Exception:
                return request.make_response(
                    b'openpyxl not available', headers=[('Content-Type', 'text/plain')])

            company_name = request.env.company.name or 'Service Provider'
            client_name = ''
            project_name = ''
            client_kra = None
            project_kra = None
            if client_kra_id:
                client_kra = request.env['kra.master'].browse(int(client_kra_id))
                if client_kra.exists():
                    client_name = client_kra.name or ''
            if project_kra_id:
                project_kra = request.env['kra.master'].browse(int(project_kra_id))
                if project_kra.exists():
                    project_name = project_kra.name or ''

            include = include_data == '1' or include_data is True

            # Pre-load existing KPI data if requested
            existing_by_kind = {'requirement': [], 'update': [], 'bug': []}
            if include and client_kra:
                scope_kra = project_kra or client_kra
                kra_ids = scope_kra._get_descendant_ids()
                kpis = request.env['kra.kpi'].search([('kra_id', 'in', kra_ids), ('active', '=', True)])
                for k in kpis:
                    nm = (k.name or '')
                    ref = (k.external_ref or '').upper()
                    if nm.startswith('[Update]') or ref.startswith('UPD'):
                        kind = 'update'
                    elif nm.startswith('[Bug]') or ref.startswith('BUG'):
                        kind = 'bug'
                    else:
                        kind = 'requirement'
                    existing_by_kind[kind].append(k)

            # Build a workbook for each template
            def _build_workbook(kind):
                wb = openpyxl.Workbook()
                ws = wb.active
                if kind == 'requirement':
                    ws.title = 'Requirements'
                    title = f'Requirements — {client_name or company_name}'
                    if project_name:
                        title += f' / {project_name}'
                    if version:
                        title += f' ({version})'
                    headers = ['Req ID (ID)', 'Module / Feature', 'Description (Task Name)',
                               'Acceptance Criteria', 'Priority (Urgent/Important/Regular)',
                               'Expected Delivery', 'Notes']
                    widths = [12, 22, 40, 32, 24, 16, 25]
                    sample = ['REQ-001', 'POS Discount', '(describe the requirement)',
                              '', 'Important', '', '']
                    fill_color = '305496'
                elif kind == 'update':
                    ws.title = 'Updates'
                    title = f'Updates / Amendments — {client_name or company_name}'
                    if project_name:
                        title += f' / {project_name}'
                    if version:
                        title += f' ({version})'
                    headers = ['Update ID (ID)', 'Linked Req (Linked Req)',
                               'Description of Change (Task Name)',
                               'Reason for Change', 'Priority (Urgent/Important/Regular)',
                               'Requested Delivery', 'Notes']
                    widths = [14, 20, 40, 30, 24, 16, 25]
                    sample = ['UPD-001', 'REQ-001', '(describe the change)',
                              '', 'Regular', '', '']
                    fill_color = 'BF8F00'
                else:  # bug
                    ws.title = 'Bug Report'
                    title = f'Bug Report — {client_name or company_name}'
                    if project_name:
                        title += f' / {project_name}'
                    if version:
                        title += f' ({version})'
                    headers = ['Bug ID (ID)', 'Linked Req (Linked Req)',
                               'Module / Screen', 'Description (Task Name)',
                               'Steps to Reproduce', 'Expected Behavior', 'Actual Behavior',
                               'Severity (High/Med/Low)', 'Priority (Urgent/Important/Regular)',
                               'Status (Open/Fixed/Closed)', 'Reported By', 'Reported Date', 'Notes']
                    widths = [12, 18, 18, 35, 35, 25, 25, 20, 22, 20, 18, 14, 25]
                    sample = ['BUG-001', 'REQ-001', 'Login Screen', '(describe the bug)',
                              '', '', '', 'Medium', 'Regular', 'Open', '', '', '']
                    fill_color = 'C00000'

                n_cols = len(headers)

                # Title row
                ws.cell(row=1, column=1, value=title)
                ws.cell(row=1, column=1).font = Font(bold=True, size=14, color='FFFFFF')
                ws.cell(row=1, column=1).fill = PatternFill('solid', fgColor=fill_color)
                ws.cell(row=1, column=1).alignment = Alignment(horizontal='center', vertical='center')
                ws.merge_cells(start_row=1, end_row=1, start_column=1, end_column=n_cols)
                ws.row_dimensions[1].height = 26

                # Header row
                for col, h in enumerate(headers, start=1):
                    c = ws.cell(row=2, column=col, value=h)
                    c.font = Font(bold=True, color='FFFFFF')
                    c.fill = PatternFill('solid', fgColor='4F81BD')
                    c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                ws.row_dimensions[2].height = 32

                # Data rows
                row = 3
                if include and existing_by_kind.get(kind):
                    for k in existing_by_kind[kind]:
                        if kind == 'requirement':
                            data = [
                                k.external_ref or '',
                                '',  # Module/Feature — not currently a KPI field
                                k.name or '',
                                '',  # Acceptance Criteria
                                (k.priority or '').title(),
                                str(k.deadline) if k.deadline else '',
                                '',
                            ]
                        elif kind == 'update':
                            data = [
                                k.external_ref or '',
                                k.related_req_ref or '',
                                k.name.replace('[Update] ', '') if k.name else '',
                                '',
                                (k.priority or '').title(),
                                str(k.deadline) if k.deadline else '',
                                '',
                            ]
                        else:  # bug
                            # Try to split "Module: Description" back
                            raw = (k.name or '').replace('[Bug] ', '')
                            module = ''
                            desc = raw
                            if ': ' in raw:
                                module, desc = raw.split(': ', 1)
                            data = [
                                k.external_ref or '',
                                k.related_req_ref or '',
                                module,
                                desc,
                                '', '', '',
                                'Medium', (k.priority or '').title(),
                                'Fixed' if k.task_state == 'completed' else 'Open',
                                '', '', '',
                            ]
                        for col, val in enumerate(data, start=1):
                            ws.cell(row=row, column=col, value=val)
                        row += 1
                else:
                    # Sample placeholder row
                    for col, val in enumerate(sample, start=1):
                        c = ws.cell(row=row, column=col, value=val)
                        c.alignment = Alignment(wrap_text=True, vertical='top')

                # Column widths
                for i, w in enumerate(widths, start=1):
                    ws.column_dimensions[get_column_letter(i)].width = w
                ws.freeze_panes = 'A3'

                buf = io.BytesIO()
                wb.save(buf)
                return buf.getvalue()

            def _safe(s):
                return re.sub(r'[^A-Za-z0-9_-]+', '_', s or '').strip('_') or 'Client'

            client_safe = _safe(client_name or 'Client')
            ver_safe = _safe(version or 'v1')

            files_to_zip = [
                (f'{client_safe}_Requirements_{ver_safe}.xlsx', _build_workbook('requirement')),
                (f'{client_safe}_Updates_{ver_safe}.xlsx',      _build_workbook('update')),
                (f'{client_safe}_Bugs_{ver_safe}.xlsx',         _build_workbook('bug')),
            ]

            zip_buf = io.BytesIO()
            with zipfile.ZipFile(zip_buf, 'w', zipfile.ZIP_DEFLATED) as z:
                for name, data in files_to_zip:
                    z.writestr(name, data)

            zip_name = f'{client_safe}_DocumentSet_{ver_safe}.zip'
            return request.make_response(zip_buf.getvalue(), headers=[
                ('Content-Type', 'application/zip'),
                ('Content-Disposition', f'attachment; filename="{zip_name}"'),
            ])
        except Exception as e:
            import logging
            logging.getLogger(__name__).error(f"generate_client_workspace_zip: {str(e)}")
            return request.make_response(
                b'Error generating workspace: ' + str(e).encode(),
                headers=[('Content-Type', 'text/plain')])

    # 🆕 JSON sibling of the ZIP route above — for the MOBILE APP.
    # The app can't use the type='http' GET: opening that URL in an external
    # browser loses the app's Odoo session cookie, so it lands on the login page
    # instead of downloading. This route rides the same authenticated JSON-RPC
    # channel every other app call already uses and returns the identical ZIP as
    # base64, which the app writes to a file and hands to the share/save sheet.
    @http.route('/kpi_client_workspace/generate_zip_b64', type='json', auth='user', methods=['POST'], csrf=False)
    def generate_client_workspace_zip_b64(self, **params):
        """Admin-only: the Generate-Client-Files ZIP, base64-encoded for the app."""
        import base64 as _b64
        user = request.env.user
        if not (user.has_group('kra_kpi_module.group_kra_admin')
                or user.has_group('kra_kpi_module.group_kra_owner')
                or user.has_group('base.group_system')):
            return {'status': False, 'message': 'Not authorized'}
        if not params.get('client_kra_id'):
            return {'status': False, 'message': 'Please select a client'}
        include = params.get('include_data')
        # Reuse the EXACT same builder as the web download — no duplicated logic.
        resp = self.generate_client_workspace_zip(
            client_kra_id=params.get('client_kra_id'),
            project_kra_id=params.get('project_kra_id') or None,
            version=params.get('version') or 'v1.0',
            include_data='1' if include in (True, 1, '1') else '0',
        )
        data = resp.data or b''
        ctype = (resp.headers.get('Content-Type') or '').lower()
        if 'zip' not in ctype:
            # The builder returned a text/plain error (e.g. openpyxl missing).
            msg = data[:200].decode('utf-8', 'ignore') or 'Could not generate the ZIP'
            return {'status': False, 'message': msg}
        disp = resp.headers.get('Content-Disposition') or ''
        m = re.search(r'filename="([^"]+)"', disp)
        return {
            'status': True,
            'filename': m.group(1) if m else 'client_files.zip',
            'data_b64': _b64.b64encode(data).decode('ascii'),
            'size': len(data),
        }

    @http.route('/kpi_completion_cert/template/<string:template_name>', type='http', auth='user')
    def download_template(self, template_name='requirement', **kw):
        """Serve dynamically-generated blank XLSX templates for the client to fill in.

        All three templates (requirement / update / bug_report) are XLSX so that the
        admin can re-import them via the same XLSX/CSV import flow on the upload screens.
        """
        try:
            company_name = request.env.company.name or 'Service Provider'
            try:
                import openpyxl
                from openpyxl.styles import Font, PatternFill, Alignment
                from openpyxl.utils import get_column_letter
                from io import BytesIO
            except Exception:
                return request.make_response(
                    b'openpyxl not available', headers=[('Content-Type', 'text/plain')])

            # ── Per-template config ──────────────────────────────────────────
            # IMPORTANT: the column tagged "(ID)" is read as the KPI's external_ref.
            # The column tagged "(Linked Req)" is read as related_req_ref (for updates/bugs only).
            configs = {
                'requirement': {
                    'sheet_title': 'Requirements',
                    'title': f'Requirements Template — {company_name}',
                    'headers': ['Req ID (ID)', 'Module / Feature', 'Description (Task Name)',
                                'Acceptance Criteria', 'Priority (Urgent/Important/Regular)',
                                'Expected Delivery', 'Notes'],
                    'widths':  [12,            22,                  40,
                                32,                   24,
                                16,                   25],
                    'sample':  ['REQ-001',     'POS Discount',     '(describe the requirement)',
                                '',                            'Important',
                                '',                            ''],
                    'filename': 'Requirements_Template.xlsx',
                },
                'update': {
                    'sheet_title': 'Updates',
                    'title': f'Requirements Update / Amendment — {company_name}',
                    'headers': ['Update ID (ID)', 'Linked Req (Linked Req)',
                                'Description of Change (Task Name)',
                                'Reason for Change', 'Priority (Urgent/Important/Regular)',
                                'Requested Delivery', 'Notes'],
                    'widths':  [14,               20,
                                40,
                                30,                  24,
                                16,                  25],
                    'sample':  ['UPD-001',        'REQ-001',
                                '(describe the change)',
                                '',                              'Regular',
                                '',                              ''],
                    'filename': 'Updates_Template.xlsx',
                },
                'bug_report': {
                    'sheet_title': 'Bug Report',
                    'title': f'Bug Report Template — {company_name}',
                    'headers': ['Bug ID (ID)', 'Linked Req (Linked Req)',
                                'Module / Screen', 'Description (Task Name)',
                                'Steps to Reproduce', 'Expected Behavior', 'Actual Behavior',
                                'Severity (High/Med/Low)', 'Priority (Urgent/Important/Regular)',
                                'Status (Open/Fixed/Closed)', 'Reported By', 'Reported Date', 'Notes'],
                    'widths':  [12,            18,
                                18,                  35,
                                35,                    25,                  25,
                                20,                          22,
                                20,                            18,            14,              25],
                    'sample':  ['BUG-001',     'REQ-001',
                                'Login Screen',      '(describe the bug)',
                                '',                    '',                  '',
                                'Medium',                    'Regular',
                                'Open',                       '',            '',              ''],
                    'filename': 'Bug_Report_Template.xlsx',
                },
            }
            cfg = configs.get(template_name)
            if not cfg:
                return request.make_response(
                    f'Unknown template: {template_name}'.encode(),
                    headers=[('Content-Type', 'text/plain')])

            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = cfg['sheet_title']
            n_cols = len(cfg['headers'])

            # Title row
            title_font = Font(bold=True, size=14, color='FFFFFF')
            title_fill = PatternFill('solid', fgColor='305496')
            ws.cell(row=1, column=1, value=cfg['title'])
            ws.cell(row=1, column=1).font = title_font
            ws.cell(row=1, column=1).fill = title_fill
            ws.cell(row=1, column=1).alignment = Alignment(horizontal='center', vertical='center')
            ws.merge_cells(start_row=1, end_row=1, start_column=1, end_column=n_cols)
            ws.row_dimensions[1].height = 26

            # Header row (row 2)
            hf = Font(bold=True, color='FFFFFF')
            fill = PatternFill('solid', fgColor='4F81BD')
            for col, h in enumerate(cfg['headers'], start=1):
                c = ws.cell(row=2, column=col, value=h)
                c.font = hf
                c.fill = fill
                c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            ws.row_dimensions[2].height = 32

            # Sample row (row 3)
            for col, v in enumerate(cfg['sample'], start=1):
                c = ws.cell(row=3, column=col, value=v)
                c.alignment = Alignment(wrap_text=True, vertical='top')

            # Column widths
            for i, w in enumerate(cfg['widths'], start=1):
                ws.column_dimensions[get_column_letter(i)].width = w

            # Freeze top 2 rows
            ws.freeze_panes = 'A3'

            buf = BytesIO()
            wb.save(buf)
            return request.make_response(buf.getvalue(), headers=[
                ('Content-Type', 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'),
                ('Content-Disposition', f'attachment; filename="{cfg["filename"]}"'),
            ])
        except Exception as e:
            import logging
            logging.getLogger(__name__).error(f"download_template: {str(e)}")
            return request.make_response(b'Error generating template: ' + str(e).encode(),
                headers=[('Content-Type', 'text/plain')])

    @http.route('/kpi_completion_cert/download_doc', type='http', auth='user')
    def download_kpi_doc(self, kpi_id=None, doc_type='requirement', **kw):
        """Stream the stored doc back as a file download. doc_type='requirement'|'updates'."""
        import base64
        try:
            if not kpi_id:
                return request.not_found()
            kpi = request.env['kra.kpi'].browse(int(kpi_id))
            if not kpi.exists():
                return request.not_found()
            if doc_type == 'updates':
                blob = kpi.updates_document
                fname = kpi.updates_document_name or f'updates_{kpi.id}.bin'
            elif doc_type == 'signed':
                blob = kpi.signed_certificate
                fname = kpi.signed_certificate_name or f'signed_cert_{kpi.id}.pdf'
            else:
                blob = kpi.requirement_document
                fname = kpi.requirement_document_name or f'requirement_{kpi.id}.bin'
            if not blob:
                return request.not_found()
            data = base64.b64decode(blob)
            return request.make_response(data, headers=[
                ('Content-Type', 'application/octet-stream'),
                ('Content-Disposition', f'attachment; filename="{fname}"'),
            ])
        except Exception as e:
            import logging
            logging.getLogger(__name__).error(f"download_kpi_doc: {str(e)}")
            return request.not_found()

    @http.route('/kpi_client_invoice/reset_draft', type='json', auth='user', methods=['POST'], csrf=False)
    def reset_client_invoice_draft(self, **params):
        try:
            inv = request.env['kpi.client.invoice'].browse(int(params['invoice_id']))
            if not inv.exists():
                return {'status': False, 'message': 'Invoice not found.'}
            inv.action_reset_draft()
            return {'status': True, 'data': inv._serialize_for_ui()}
        except Exception as e:
            import logging
            logging.getLogger(__name__).error(f"reset_client_invoice_draft: {str(e)}")
            return {'status': False, 'message': str(e)}